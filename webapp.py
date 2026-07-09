"""
Drop Bot — Web Dashboard
========================

A lightweight web interface for managing the DB-backed parts of a Drop Bot
server without using Discord: payment methods, managers, drop history, orders,
and shipping tracking. It talks to the SAME PostgreSQL database as the bot.

Scope note: staging stock, taking claims, and closing a drop still happen in
Discord, because that live state lives in the bot's memory. The dashboard can
now *watch* a live drop and *mark buyers paid/unpaid* on it: the bot mirrors the
live drop into the live_orders/live_drops tables, and this app writes buyer
paid/unpaid actions to a pending_actions outbox the bot applies back into its
in-memory state (~15s). Config changes made here (payment methods, managers,
channels) are picked up by the running bot within ~60s via its config refresh.

Auth: per-server access key. A manager runs `!webkey` in Discord to get the
key, then pastes it here to sign in. The signed session cookie remembers which
guild they're managing.

Run:  uvicorn webapp:app --host 0.0.0.0 --port $PORT
Env:  DATABASE_URL (required, shared with the bot)
      WEB_SECRET   (required in prod — signs session cookies)
      WEB_SECURE_COOKIES  ("true"/"false", default "true")
      PORT         (provided by the host)
"""

import datetime
import io
import json
import os
import secrets
from contextlib import asynccontextmanager
from urllib.parse import quote, quote_plus

import asyncpg
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from fastapi import FastAPI, Form, Request
from fastapi.responses import RedirectResponse, Response
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from starlette.middleware.sessions import SessionMiddleware

DATABASE_URL = os.getenv("DATABASE_URL")
WEB_SECRET = os.getenv("WEB_SECRET")
if not WEB_SECRET:
    # Ephemeral secret keeps things working but logs everyone out on restart.
    WEB_SECRET = secrets.token_urlsafe(32)
    print("⚠️  WEB_SECRET not set — using a temporary secret. "
          "Set WEB_SECRET in the environment so sessions survive restarts.")
SECURE_COOKIES = os.getenv("WEB_SECURE_COOKIES", "true").lower() == "true"
# Optional master key for the bot creator to oversee ALL servers. If unset,
# creator login is disabled and only per-server access keys work.
CREATOR_WEB_KEY = os.getenv("CREATOR_WEB_KEY", "")


async def ensure_schema(pool):
    """Make sure the columns the dashboard reads exist.

    The bot's init_db() normally creates these, but the dashboard shouldn't
    depend on the bot having redeployed first. These statements are idempotent
    and safe to run on every startup.
    """
    statements = [
        "ALTER TABLE server_settings ADD COLUMN IF NOT EXISTS web_access_key TEXT",
        "ALTER TABLE server_settings ADD COLUMN IF NOT EXISTS guild_name TEXT",
        "ALTER TABLE user_claims ADD COLUMN IF NOT EXISTS tracking TEXT",
        """CREATE TABLE IF NOT EXISTS bot_guilds (
               guild_id   BIGINT PRIMARY KEY,
               guild_name TEXT,
               active     BOOLEAN NOT NULL DEFAULT TRUE,
               last_seen  TIMESTAMP NOT NULL DEFAULT NOW()
           )""",
        """CREATE TABLE IF NOT EXISTS pending_notifications (
               id         SERIAL PRIMARY KEY,
               guild_id   BIGINT NOT NULL,
               user_id    BIGINT NOT NULL,
               message    TEXT NOT NULL,
               created_at TIMESTAMP NOT NULL DEFAULT NOW(),
               sent_at    TIMESTAMP
           )""",
        # Live-drop mirror (written by the bot) + write-back outbox (written here).
        """CREATE TABLE IF NOT EXISTS live_drops (
               guild_id    BIGINT PRIMARY KEY,
               drop_number INT,
               is_live     BOOLEAN NOT NULL DEFAULT FALSE,
               updated_at  TIMESTAMP NOT NULL DEFAULT NOW()
           )""",
        """CREATE TABLE IF NOT EXISTS live_orders (
               guild_id        BIGINT NOT NULL,
               user_id         BIGINT NOT NULL,
               user_name       TEXT NOT NULL,
               drop_number     INT,
               items           JSONB NOT NULL,
               total           NUMERIC NOT NULL,
               confirmed_total NUMERIC NOT NULL DEFAULT 0,
               paid            BOOLEAN NOT NULL DEFAULT FALSE,
               updated_at      TIMESTAMP NOT NULL DEFAULT NOW(),
               PRIMARY KEY (guild_id, user_id)
           )""",
        """CREATE TABLE IF NOT EXISTS pending_actions (
               id         SERIAL PRIMARY KEY,
               guild_id   BIGINT NOT NULL,
               user_id    BIGINT NOT NULL,
               action     TEXT NOT NULL,
               created_at TIMESTAMP NOT NULL DEFAULT NOW(),
               applied_at TIMESTAMP
           )""",
        """CREATE TABLE IF NOT EXISTS raffles (
               guild_id   BIGINT  NOT NULL,
               name       TEXT    NOT NULL,
               spots      INTEGER NOT NULL,
               price      TEXT    NOT NULL,
               channel_id BIGINT  NOT NULL,
               message_id BIGINT,
               status     TEXT    NOT NULL DEFAULT 'open',
               PRIMARY KEY (guild_id, name)
           )""",
        "ALTER TABLE raffles ADD COLUMN IF NOT EXISTS host_num INTEGER DEFAULT 0",
        """CREATE TABLE IF NOT EXISTS raffle_slots (
               guild_id    BIGINT  NOT NULL,
               raffle_name TEXT    NOT NULL,
               spot_num    INTEGER NOT NULL,
               user_id     BIGINT,
               username    TEXT,
               paid        BOOLEAN NOT NULL DEFAULT FALSE,
               PRIMARY KEY (guild_id, raffle_name, spot_num)
           )""",
        """CREATE TABLE IF NOT EXISTS raffle_hosts (
               guild_id  BIGINT NOT NULL,
               host_num  INTEGER NOT NULL,
               name      TEXT,
               venmo     TEXT,
               zelle     TEXT,
               cashapp   TEXT,
               applepay  TEXT,
               PRIMARY KEY (guild_id, host_num)
           )""",
    ]
    async with pool.acquire() as conn:
        for stmt in statements:
            try:
                await conn.execute(stmt)
            except Exception as e:  # table may not exist yet on a brand-new DB
                print(f"⚠️  ensure_schema skipped: {stmt} — {e}")


@asynccontextmanager
async def lifespan(app: FastAPI):
    if not DATABASE_URL:
        raise RuntimeError("DATABASE_URL is not set.")
    app.state.pool = await asyncpg.create_pool(DATABASE_URL)
    await ensure_schema(app.state.pool)
    yield
    await app.state.pool.close()


app = FastAPI(lifespan=lifespan, title="Drop Bot Dashboard")
app.add_middleware(
    SessionMiddleware,
    secret_key=WEB_SECRET,
    same_site="lax",
    https_only=SECURE_COOKIES,
    max_age=60 * 60 * 24 * 7,  # 7 days
)
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")


# ── Helpers ───────────────────────────────────────────────────────────────────

def _money(value) -> str:
    try:
        return f"${float(value):,.2f}"
    except (TypeError, ValueError):
        return "$0.00"


templates.env.filters["money"] = _money


def _asset_version() -> str:
    """Short hash of the static asset mtimes, recomputed at startup. Appended
    to CSS/JS URLs so browsers fetch fresh files after each deploy instead of
    serving a stale cached copy."""
    import hashlib
    import pathlib

    h = hashlib.md5()
    for name in ("static/style.css", "static/app.js"):
        try:
            h.update(str(pathlib.Path(name).stat().st_mtime_ns).encode())
        except OSError:
            pass
    return h.hexdigest()[:8]


templates.env.globals["asset_ver"] = _asset_version()


def _session_guild(request: Request):
    """Return (guild_id, guild_name) from the session, or (None, None)."""
    gid = request.session.get("guild_id")
    if gid is None:
        return None, None
    return gid, request.session.get("guild_name") or str(gid)


def _redirect_login():
    return RedirectResponse("/login", status_code=303)


def _ctx(request: Request, gid, gname, **extra):
    base = {
        "request": request,
        "guild_id": gid,
        "guild_name": gname,
        "is_creator": bool(request.session.get("is_creator")),
    }
    base.update(extra)
    return base


async def _drop_number_for(conn, guild_id, closed_at):
    """Map a drop_history row to its 1-based drop number (by close order)."""
    row = await conn.fetchrow(
        """SELECT COUNT(*) AS n FROM drop_history
           WHERE guild_id = $1 AND closed_at <= $2""",
        guild_id, closed_at,
    )
    return row["n"] if row else 0


async def _load_drop_orders(conn, guild_id, drop_number):
    """Return per-buyer orders for a given drop from user_claims."""
    rows = await conn.fetch(
        """SELECT user_id, user_name, item_display, qty, price, subtotal,
                  confirmed, tracking
           FROM user_claims
           WHERE guild_id = $1 AND drop_number = $2
           ORDER BY user_name, item_display""",
        guild_id, drop_number,
    )
    buyers = {}
    for r in rows:
        uid = r["user_id"]
        if uid not in buyers:
            buyers[uid] = {
                "user_id": uid,
                "name": r["user_name"],
                "items": [],
                "total": 0.0,
                "confirmed": False,
                "tracking": r["tracking"] or "",
            }
        buyers[uid]["items"].append({
            "display": r["item_display"],
            "qty": r["qty"],
            "subtotal": float(r["subtotal"]),
        })
        buyers[uid]["total"] += float(r["subtotal"])
        if r["confirmed"]:
            buyers[uid]["confirmed"] = True
        if r["tracking"] and not buyers[uid]["tracking"]:
            buyers[uid]["tracking"] = r["tracking"]
    return list(buyers.values())


# ── Auth ──────────────────────────────────────────────────────────────────────

@app.get("/login")
async def login_form(request: Request):
    if request.session.get("guild_id"):
        return RedirectResponse("/", status_code=303)
    return templates.TemplateResponse(
        request, "login.html", {"error": None}
    )


@app.post("/login")
async def login_submit(request: Request, key: str = Form("")):
    key = (key or "").strip()
    if key:
        # Creator master key → oversee all servers.
        if CREATOR_WEB_KEY and secrets.compare_digest(key, CREATOR_WEB_KEY):
            request.session.clear()
            request.session["is_creator"] = True
            return RedirectResponse("/admin", status_code=303)
        # Otherwise a per-server access key.
        async with request.app.state.pool.acquire() as conn:
            row = await conn.fetchrow(
                "SELECT guild_id, guild_name FROM server_settings "
                "WHERE web_access_key = $1",
                key,
            )
        if row:
            request.session.clear()
            request.session["guild_id"] = row["guild_id"]
            request.session["guild_name"] = row["guild_name"] or str(row["guild_id"])
            return RedirectResponse("/", status_code=303)
    return templates.TemplateResponse(
        request,
        "login.html",
        {"error": "Invalid access key. Run !webkey in Discord."},
        status_code=401,
    )


@app.get("/logout")
async def logout(request: Request):
    request.session.clear()
    return RedirectResponse("/login", status_code=303)


# ── Creator (all-servers) oversight ───────────────────────────────────────────

@app.get("/admin")
async def admin_overview(request: Request):
    if not request.session.get("is_creator"):
        return _redirect_login()
    async with request.app.state.pool.acquire() as conn:
        settings = await conn.fetch(
            "SELECT guild_id, guild_name FROM server_settings"
        )
        agg = await conn.fetch(
            """SELECT guild_id, COUNT(*) AS drops,
                      COALESCE(SUM(total_revenue), 0) AS revenue,
                      COALESCE(SUM(total_items), 0) AS items,
                      COALESCE(SUM(unique_buyers), 0) AS buyers,
                      MAX(closed_at) AS last_drop
               FROM drop_history GROUP BY guild_id"""
        )
        outstanding = await conn.fetch(
            """SELECT guild_id, COALESCE(SUM(subtotal), 0) AS outstanding
               FROM user_claims WHERE confirmed = FALSE GROUP BY guild_id"""
        )
        presence = await conn.fetch(
            "SELECT guild_id, guild_name, active FROM bot_guilds"
        )
    names = {r["guild_id"]: r["guild_name"] for r in settings}
    agg_map = {r["guild_id"]: r for r in agg}
    out_map = {r["guild_id"]: float(r["outstanding"]) for r in outstanding}
    presence_map = {r["guild_id"]: r for r in presence}

    stores = []
    for gid in set(names) | set(agg_map) | set(out_map) | set(presence_map):
        a = agg_map.get(gid)
        p = presence_map.get(gid)
        # Bot presence is only confirmed by an explicit active=TRUE row,
        # written on join and reconciled on every startup. No row at all
        # (data predates this feature) is treated the same as "left" —
        # we can't confirm the bot is still there.
        bot_present = bool(p and p["active"])
        display_name = names.get(gid) or (p["guild_name"] if p else None) or str(gid)
        stores.append({
            "guild_id": gid,
            "name": display_name,
            "drops": a["drops"] if a else 0,
            "revenue": float(a["revenue"]) if a else 0.0,
            "items": a["items"] if a else 0,
            "buyers": a["buyers"] if a else 0,
            "last_drop": a["last_drop"] if a else None,
            "outstanding": out_map.get(gid, 0.0),
            "bot_present": bot_present,
        })
    stores.sort(
        key=lambda s: s["last_drop"] or datetime.datetime.min, reverse=True
    )
    totals = {
        "stores": len(stores),
        "active_stores": sum(1 for s in stores if s["bot_present"]),
        "drops": sum(s["drops"] for s in stores),
        "revenue": sum(s["revenue"] for s in stores),
        "outstanding": sum(s["outstanding"] for s in stores),
    }
    return templates.TemplateResponse(request, "admin.html", _ctx(
        request, None, None, stores=stores, totals=totals,
    ))


@app.get("/admin/select/{guild_id}")
async def admin_select(request: Request, guild_id: int):
    if not request.session.get("is_creator"):
        return _redirect_login()
    async with request.app.state.pool.acquire() as conn:
        row = await conn.fetchrow(
            "SELECT guild_name FROM server_settings WHERE guild_id = $1", guild_id
        )
    request.session["guild_id"] = guild_id
    request.session["guild_name"] = (
        row["guild_name"] if row and row["guild_name"] else str(guild_id)
    )
    return RedirectResponse("/", status_code=303)


# ── Dashboard ─────────────────────────────────────────────────────────────────

@app.get("/")
async def dashboard(request: Request):
    gid, gname = _session_guild(request)
    if gid is None:
        return _redirect_login()
    async with request.app.state.pool.acquire() as conn:
        stats = await conn.fetchrow(
            """SELECT COUNT(*) AS drops,
                      COALESCE(SUM(total_revenue), 0) AS revenue,
                      COALESCE(SUM(total_items), 0) AS items_sold,
                      COALESCE(SUM(unique_buyers), 0) AS buyers
               FROM drop_history WHERE guild_id = $1""",
            gid,
        )
        recent = await conn.fetch(
            """SELECT closed_at, total_revenue, total_items, unique_buyers,
                      ROW_NUMBER() OVER (ORDER BY closed_at) AS drop_number
               FROM drop_history WHERE guild_id = $1
               ORDER BY closed_at DESC LIMIT 5""",
            gid,
        )
        untracked = await conn.fetchrow(
            """SELECT COUNT(*) AS n FROM (
                   SELECT DISTINCT drop_number, user_id
                   FROM user_claims
                   WHERE guild_id = $1 AND (tracking IS NULL OR tracking = '')
               ) t""",
            gid,
        )
        outstanding = await conn.fetchrow(
            """SELECT COALESCE(SUM(subtotal), 0) AS amount
               FROM user_claims WHERE guild_id = $1 AND confirmed = FALSE""",
            gid,
        )
    return templates.TemplateResponse(request, "dashboard.html", _ctx(
        request, gid, gname,
        stats=dict(stats), recent=[dict(r) for r in recent],
        untracked=untracked["n"] if untracked else 0,
        outstanding=float(outstanding["amount"]) if outstanding else 0.0,
    ))


# ── Settings ──────────────────────────────────────────────────────────────────

@app.get("/settings")
async def settings_form(request: Request, saved: int = 0):
    gid, gname = _session_guild(request)
    if gid is None:
        return _redirect_login()
    async with request.app.state.pool.acquire() as conn:
        row = await conn.fetchrow(
            "SELECT * FROM server_settings WHERE guild_id = $1", gid
        )
    return templates.TemplateResponse(request, "settings.html", _ctx(
        request, gid, gname, s=dict(row) if row else None, saved=bool(saved),
    ))


@app.post("/settings")
async def settings_save(
    request: Request,
    venmo: str = Form(""),
    zelle: str = Form(""),
    cashapp: str = Form(""),
    applepay: str = Form(""),
    drop_channel_id: str = Form(""),
    raffle_channel_id: str = Form(""),
):
    gid, _ = _session_guild(request)
    if gid is None:
        return _redirect_login()

    def _clean(v):
        v = (v or "").strip()
        return v or None

    def _clean_int(v):
        v = (v or "").strip()
        return int(v) if v.isdigit() else None

    async with request.app.state.pool.acquire() as conn:
        await conn.execute(
            """INSERT INTO server_settings
                   (guild_id, venmo, zelle, cashapp, applepay,
                    drop_channel_id, raffle_channel_id)
               VALUES ($1, $2, $3, $4, $5, $6, $7)
               ON CONFLICT (guild_id) DO UPDATE SET
                   venmo = $2, zelle = $3, cashapp = $4, applepay = $5,
                   drop_channel_id = $6, raffle_channel_id = $7""",
            gid, _clean(venmo), _clean(zelle), _clean(cashapp), _clean(applepay),
            _clean_int(drop_channel_id), _clean_int(raffle_channel_id),
        )
    return RedirectResponse("/settings?saved=1", status_code=303)


# ── Managers ──────────────────────────────────────────────────────────────────

@app.get("/managers")
async def managers_view(request: Request, msg: str = ""):
    gid, gname = _session_guild(request)
    if gid is None:
        return _redirect_login()
    async with request.app.state.pool.acquire() as conn:
        mrows = await conn.fetch(
            "SELECT user_id FROM server_managers WHERE guild_id = $1 ORDER BY user_id",
            gid,
        )
        arow = await conn.fetchrow(
            "SELECT user_id FROM server_admins WHERE guild_id = $1", gid
        )
    return templates.TemplateResponse(request, "managers.html", _ctx(
        request, gid, gname,
        managers=[r["user_id"] for r in mrows],
        admin=arow["user_id"] if arow else None,
        msg=msg,
    ))


@app.post("/managers/add")
async def managers_add(request: Request, user_id: str = Form("")):
    gid, _ = _session_guild(request)
    if gid is None:
        return _redirect_login()
    user_id = (user_id or "").strip()
    if not user_id.isdigit():
        return RedirectResponse("/managers?msg=A+numeric+Discord+user+ID+is+required.",
                                status_code=303)
    async with request.app.state.pool.acquire() as conn:
        await conn.execute(
            """INSERT INTO server_managers (guild_id, user_id)
               VALUES ($1, $2) ON CONFLICT DO NOTHING""",
            gid, int(user_id),
        )
    return RedirectResponse("/managers?msg=Manager+added.", status_code=303)


@app.post("/managers/remove")
async def managers_remove(request: Request, user_id: str = Form("")):
    gid, _ = _session_guild(request)
    if gid is None:
        return _redirect_login()
    user_id = (user_id or "").strip()
    if user_id.isdigit():
        async with request.app.state.pool.acquire() as conn:
            await conn.execute(
                "DELETE FROM server_managers WHERE guild_id = $1 AND user_id = $2",
                gid, int(user_id),
            )
    return RedirectResponse("/managers?msg=Manager+removed.", status_code=303)


# ── Live drop ─────────────────────────────────────────────────────────────────

@app.get("/live")
async def live_drop(request: Request, msg: str = ""):
    gid, gname = _session_guild(request)
    if gid is None:
        return _redirect_login()
    async with request.app.state.pool.acquire() as conn:
        meta = await conn.fetchrow(
            "SELECT drop_number, is_live, updated_at FROM live_drops WHERE guild_id = $1",
            gid,
        )
        rows = await conn.fetch(
            """SELECT user_id, user_name, items, total, confirmed_total, paid
               FROM live_orders WHERE guild_id = $1
               ORDER BY user_name""",
            gid,
        )
    is_live = bool(meta and meta["is_live"])
    orders = []
    for r in rows:
        items = r["items"]
        if isinstance(items, str):
            items = json.loads(items)
        orders.append({
            "user_id": r["user_id"],
            "name": r["user_name"],
            "items": items,
            "total": float(r["total"]),
            "confirmed_total": float(r["confirmed_total"]),
            "paid": r["paid"],
        })
    total = sum(o["total"] for o in orders)
    unpaid_count = sum(1 for o in orders if not o["paid"])
    return templates.TemplateResponse(request, "live.html", _ctx(
        request, gid, gname,
        is_live=is_live,
        drop_number=meta["drop_number"] if meta else None,
        updated_at=meta["updated_at"] if meta else None,
        orders=orders, total=total, unpaid_count=unpaid_count, msg=msg,
    ))


async def _queue_live_action(request: Request, user_id: str, action: str):
    gid, _ = _session_guild(request)
    if gid is None:
        return _redirect_login()
    user_id = (user_id or "").strip()
    if user_id.isdigit():
        async with request.app.state.pool.acquire() as conn:
            await conn.execute(
                """INSERT INTO pending_actions (guild_id, user_id, action)
                   VALUES ($1, $2, $3)""",
                gid, int(user_id), action,
            )
    verb = "Marked+paid" if action == "confirm" else "Marked+unpaid"
    return RedirectResponse(
        f"/live?msg={verb}+-+syncing+to+Discord+in+~15s.", status_code=303,
    )


@app.post("/live/confirm")
async def live_confirm(request: Request, user_id: str = Form("")):
    return await _queue_live_action(request, user_id, "confirm")


@app.post("/live/unconfirm")
async def live_unconfirm(request: Request, user_id: str = Form("")):
    return await _queue_live_action(request, user_id, "unconfirm")


# ── Drops & orders ────────────────────────────────────────────────────────────

@app.get("/drops")
async def drops_list(request: Request):
    gid, gname = _session_guild(request)
    if gid is None:
        return _redirect_login()
    async with request.app.state.pool.acquire() as conn:
        rows = await conn.fetch(
            """SELECT closed_at, total_revenue, total_items, unique_buyers,
                      ROW_NUMBER() OVER (ORDER BY closed_at) AS drop_number
               FROM drop_history WHERE guild_id = $1
               ORDER BY closed_at DESC""",
            gid,
        )
    return templates.TemplateResponse(request, "drops.html", _ctx(
        request, gid, gname, drops=[dict(r) for r in rows],
    ))


@app.get("/drops/{drop_number}")
async def drop_detail(request: Request, drop_number: int, msg: str = "", view: str = ""):
    gid, gname = _session_guild(request)
    if gid is None:
        return _redirect_login()
    async with request.app.state.pool.acquire() as conn:
        orders_all = await _load_drop_orders(conn, gid, drop_number)
        meta = await conn.fetchrow(
            """SELECT closed_at FROM drop_history WHERE guild_id = $1
               ORDER BY closed_at OFFSET $2 LIMIT 1""",
            gid, max(drop_number - 1, 0),
        )
    total = sum(o["total"] for o in orders_all)
    unpaid_count = sum(1 for o in orders_all if not o["confirmed"])
    tracked_count = sum(1 for o in orders_all if o["tracking"])
    show_unpaid = (view == "unpaid")
    orders = [o for o in orders_all if not o["confirmed"]] if show_unpaid else orders_all
    return templates.TemplateResponse(request, "drop_detail.html", _ctx(
        request, gid, gname,
        drop_number=drop_number, orders=orders, total=total,
        closed_at=meta["closed_at"] if meta else None, msg=msg,
        show_unpaid=show_unpaid, unpaid_count=unpaid_count, all_count=len(orders_all),
        tracked_count=tracked_count,
    ))


def _drop_redirect(drop_number, msg, view=""):
    suffix = f"&view={view}" if view in ("unpaid",) else ""
    return RedirectResponse(f"/drops/{drop_number}?msg={msg}{suffix}", status_code=303)


def _tracking_dm(drop_number, tracking):
    """The DM text a buyer gets when their tracking number is delivered.

    Kept in one place so the single-save flow and the bulk push below stay
    identical to what `!addtracking` sends.
    """
    return (
        f"📦  Your order from **Drop #{drop_number}** has shipped! "
        f"Here is your tracking number:\n"
        f"**{tracking}**\n\n"
        f"You can use `!myhistory` to view your full order history."
    )


@app.post("/drops/{drop_number}/tracking")
async def drop_set_tracking(
    request: Request, drop_number: int,
    user_id: str = Form(""), tracking: str = Form(""), view: str = Form(""),
):
    gid, _ = _session_guild(request)
    if gid is None:
        return _redirect_login()
    user_id = (user_id or "").strip()
    tracking = (tracking or "").strip() or None
    notified = False
    if user_id.isdigit():
        uid = int(user_id)
        async with request.app.state.pool.acquire() as conn:
            async with conn.transaction():
                old_row = await conn.fetchrow(
                    """SELECT tracking FROM user_claims
                       WHERE guild_id = $1 AND user_id = $2 AND drop_number = $3
                       LIMIT 1""",
                    gid, uid, drop_number,
                )
                old_tracking = old_row["tracking"] if old_row else None
                await conn.execute(
                    """UPDATE user_claims SET tracking = $4
                       WHERE guild_id = $1 AND user_id = $2 AND drop_number = $3""",
                    gid, uid, drop_number, tracking,
                )
                # Only DM the buyer on a real, new tracking number — not when
                # clearing it or re-saving the same value.
                if tracking and tracking != old_tracking:
                    message = _tracking_dm(drop_number, tracking)
                    await conn.execute(
                        """INSERT INTO pending_notifications (guild_id, user_id, message)
                           VALUES ($1, $2, $3)""",
                        gid, uid, message,
                    )
                    notified = True
    verb = "Tracking+updated+-+buyer+notified+in+Discord." if notified else "Tracking+updated."
    return _drop_redirect(drop_number, verb, view)


@app.post("/drops/{drop_number}/confirm")
async def drop_set_confirmed(
    request: Request, drop_number: int,
    user_id: str = Form(""), confirmed: str = Form(""), view: str = Form(""),
):
    gid, _ = _session_guild(request)
    if gid is None:
        return _redirect_login()
    user_id = (user_id or "").strip()
    is_confirmed = (confirmed == "1")
    if user_id.isdigit():
        async with request.app.state.pool.acquire() as conn:
            await conn.execute(
                """UPDATE user_claims SET confirmed = $4
                   WHERE guild_id = $1 AND user_id = $2 AND drop_number = $3""",
                gid, int(user_id), drop_number, is_confirmed,
            )
    verb = "Marked+paid." if is_confirmed else "Marked+unpaid."
    return _drop_redirect(drop_number, verb, view)


@app.post("/drops/{drop_number}/confirm_all")
async def drop_confirm_all(request: Request, drop_number: int):
    gid, _ = _session_guild(request)
    if gid is None:
        return _redirect_login()
    async with request.app.state.pool.acquire() as conn:
        result = await conn.execute(
            """UPDATE user_claims SET confirmed = TRUE
               WHERE guild_id = $1 AND drop_number = $2 AND confirmed = FALSE""",
            gid, drop_number,
        )
    # result like "UPDATE <n>"
    try:
        n_rows = int(result.split()[-1])
    except (ValueError, IndexError):
        n_rows = 0
    label = "All+orders+marked+paid." if n_rows else "Nothing+to+update."
    return _drop_redirect(drop_number, label)


@app.post("/drops/{drop_number}/notify_tracking")
async def drop_notify_tracking(request: Request, drop_number: int):
    """Re-send the tracking DM to every buyer in a drop that already has a
    saved tracking number.

    The per-buyer Save only DMs when a tracking value *changes* to something
    new, so trackings entered before the notify feature existed — or any that
    a buyer never got — leave no way to reach the buyer. This enqueues one DM
    per buyer with a saved tracking number, using the same outbox the bot
    polls for the single-save flow.
    """
    gid, _ = _session_guild(request)
    if gid is None:
        return _redirect_login()
    async with request.app.state.pool.acquire() as conn:
        rows = await conn.fetch(
            """SELECT DISTINCT ON (user_id) user_id, tracking
               FROM user_claims
               WHERE guild_id = $1 AND drop_number = $2
                 AND tracking IS NOT NULL AND tracking <> ''
               ORDER BY user_id, tracking""",
            gid, drop_number,
        )
        for r in rows:
            await conn.execute(
                """INSERT INTO pending_notifications (guild_id, user_id, message)
                   VALUES ($1, $2, $3)""",
                gid, r["user_id"], _tracking_dm(drop_number, r["tracking"]),
            )
    n = len(rows)
    label = (
        f"Pushed+tracking+to+{n}+buyer{'s' if n != 1 else ''}+-+DMs+arrive+in+~15-30s."
        if n else "No+saved+tracking+numbers+to+push."
    )
    return _drop_redirect(drop_number, label)


@app.get("/orders")
async def orders_search(request: Request, q: str = ""):
    gid, gname = _session_guild(request)
    if gid is None:
        return _redirect_login()
    q = (q or "").strip()
    results = []
    if q:
        async with request.app.state.pool.acquire() as conn:
            if q.isdigit():
                rows = await conn.fetch(
                    """SELECT drop_number, user_id, user_name, item_display,
                              qty, subtotal, confirmed, tracking, closed_at
                       FROM user_claims
                       WHERE guild_id = $1 AND user_id = $2
                       ORDER BY drop_number DESC, item_display""",
                    gid, int(q),
                )
            else:
                rows = await conn.fetch(
                    """SELECT drop_number, user_id, user_name, item_display,
                              qty, subtotal, confirmed, tracking, closed_at
                       FROM user_claims
                       WHERE guild_id = $1 AND user_name ILIKE $2
                       ORDER BY drop_number DESC, item_display""",
                    gid, f"%{q}%",
                )
        results = [dict(r) for r in rows]
    return templates.TemplateResponse(request, "orders.html", _ctx(
        request, gid, gname, q=q, results=results,
    ))


# ── Raffles ───────────────────────────────────────────────────────────────────
#
# Raffle records are DB-backed, so the dashboard can read and edit them like
# drops. The one thing it can't do is talk to Discord — so a raffle created
# here is saved with status 'pending', and the bot's ~15s sync loop posts the
# embed + claim buttons and flips it to 'open'. Paid toggles work the same
# way in reverse: the dashboard updates raffle_slots directly (and queues the
# buyer's confirmation DM), and the bot folds the change into its memory and
# refreshes the Discord embed on its next pass.

RAFFLE_STATUS_LABELS = {
    "pending":  "Posting to Discord…",
    "open":     "Open",
    "closed":   "Full — awaiting payments",
    "complete": "Complete",
}


def _raffles_redirect(msg="", err=""):
    q = f"?msg={quote_plus(msg)}" if msg else (f"?err={quote_plus(err)}" if err else "")
    return RedirectResponse(f"/raffles{q}", status_code=303)


def _raffle_redirect(name, msg=""):
    q = f"?msg={quote_plus(msg)}" if msg else ""
    return RedirectResponse(f"/raffles/{quote(name, safe='')}{q}", status_code=303)


@app.get("/raffles")
async def raffles_list(request: Request, msg: str = "", err: str = ""):
    gid, gname = _session_guild(request)
    if gid is None:
        return _redirect_login()
    async with request.app.state.pool.acquire() as conn:
        rows = await conn.fetch(
            "SELECT * FROM raffles WHERE guild_id = $1", gid)
        slot_rows = await conn.fetch(
            """SELECT raffle_name,
                      COUNT(user_id) FILTER (WHERE user_id IS NOT NULL) AS claimed,
                      COUNT(*) FILTER (WHERE paid) AS paid
               FROM raffle_slots WHERE guild_id = $1 GROUP BY raffle_name""", gid)
        settings = await conn.fetchrow(
            "SELECT raffle_channel_id FROM server_settings WHERE guild_id = $1", gid)
        hosts = await conn.fetch(
            "SELECT host_num, name FROM raffle_hosts WHERE guild_id = $1 "
            "ORDER BY host_num", gid)
    counts = {r["raffle_name"]: r for r in slot_rows}
    order = {"pending": 0, "open": 1, "closed": 2, "complete": 3}
    raffles = sorted((dict(r) for r in rows),
                     key=lambda r: (order.get(r["status"], 9), r["name"].lower()))
    for r in raffles:
        c = counts.get(r["name"])
        r["claimed"] = c["claimed"] if c else 0
        r["paid"] = c["paid"] if c else 0
        r["status_label"] = RAFFLE_STATUS_LABELS.get(r["status"], r["status"])
    return templates.TemplateResponse(request, "raffles.html", _ctx(
        request, gid, gname, raffles=raffles, msg=msg, err=err,
        raffle_channel_id=settings["raffle_channel_id"] if settings else None,
        hosts=[dict(h) for h in hosts],
    ))


@app.post("/raffles/create")
async def raffle_create(
    request: Request,
    name: str = Form(""), spots: str = Form(""), price: str = Form(""),
    host: str = Form("0"),
):
    gid, _ = _session_guild(request)
    if gid is None:
        return _redirect_login()

    name = (name or "").strip()
    price = (price or "").strip()
    # ":" is the separator inside the claim buttons' custom_id, and Discord
    # caps custom_ids at 100 chars — keep names simple like /raffle create does.
    if not name or len(name) > 60 or ":" in name:
        return _raffles_redirect(err="Raffle name is required — up to 60 characters, no ':'.")
    try:
        n_spots = int(spots)
    except (TypeError, ValueError):
        n_spots = 0
    if not 2 <= n_spots <= 10:
        return _raffles_redirect(err="Spots must be between 2 and 10.")
    if not price:
        return _raffles_redirect(err="Price per spot is required (e.g. $25).")
    if not price.startswith("$"):
        price = f"${price}"
    try:
        host_num = int(host)
    except (TypeError, ValueError):
        host_num = -1
    if host_num not in (0, 1, 2):
        return _raffles_redirect(err="Host must be server default, 1 or 2.")

    async with request.app.state.pool.acquire() as conn:
        settings = await conn.fetchrow(
            "SELECT raffle_channel_id FROM server_settings WHERE guild_id = $1", gid)
        channel_id = settings["raffle_channel_id"] if settings else None
        if not channel_id:
            return _raffles_redirect(
                err="Set the raffle channel ID in Settings first.")
        if host_num in (1, 2):
            h = await conn.fetchrow(
                "SELECT 1 FROM raffle_hosts WHERE guild_id = $1 AND host_num = $2",
                gid, host_num)
            if not h:
                return _raffles_redirect(
                    err=f"Host {host_num} has no payment info yet — "
                        f"run /raffle sethost in Discord first.")
        exists = await conn.fetchval(
            "SELECT 1 FROM raffles WHERE guild_id = $1 AND name = $2", gid, name)
        if exists:
            return _raffles_redirect(
                err=f"A raffle named “{name}” already exists.")
        async with conn.transaction():
            await conn.execute(
                """INSERT INTO raffles
                       (guild_id, name, spots, price, channel_id,
                        message_id, status, host_num)
                   VALUES ($1, $2, $3, $4, $5, NULL, 'pending', $6)""",
                gid, name, n_spots, price, channel_id, host_num)
            for num in range(1, n_spots + 1):
                await conn.execute(
                    """INSERT INTO raffle_slots
                           (guild_id, raffle_name, spot_num, user_id, username, paid)
                       VALUES ($1, $2, $3, NULL, NULL, FALSE)
                       ON CONFLICT DO NOTHING""",
                    gid, name, num)
    return _raffles_redirect(
        msg=f"Raffle “{name}” created — the bot posts it in Discord "
            f"within ~15-30 seconds.")


@app.post("/raffles/paid")
async def raffle_set_paid(
    request: Request,
    name: str = Form(""), spot: str = Form(""), paid: str = Form(""),
):
    gid, _ = _session_guild(request)
    if gid is None:
        return _redirect_login()
    name = (name or "").strip()
    try:
        spot_num = int(spot)
    except (TypeError, ValueError):
        return _raffles_redirect(err="Bad spot number.")
    is_paid = (paid == "1")

    async with request.app.state.pool.acquire() as conn:
        async with conn.transaction():
            row = await conn.fetchrow(
                """SELECT user_id, username, paid FROM raffle_slots
                   WHERE guild_id = $1 AND raffle_name = $2 AND spot_num = $3""",
                gid, name, spot_num)
            if not row:
                return _raffle_redirect(name, "That spot doesn't exist.")
            if row["user_id"] is None:
                return _raffle_redirect(name, "That spot is unclaimed.")
            await conn.execute(
                """UPDATE raffle_slots SET paid = $4
                   WHERE guild_id = $1 AND raffle_name = $2 AND spot_num = $3""",
                gid, name, spot_num, is_paid)
            notified = False
            # Same DM /raffle confirm sends — only on a genuinely new confirm.
            if is_paid and not row["paid"]:
                await conn.execute(
                    """INSERT INTO pending_notifications (guild_id, user_id, message)
                       VALUES ($1, $2, $3)""",
                    gid, row["user_id"],
                    f"Your payment for Spot #{spot_num} in the **{name}** raffle "
                    f"is confirmed! Watch for the live spin announcement.")
                notified = True
    verb = ("Marked paid — buyer DM'd, Discord board updates in ~15s."
            if notified else
            ("Marked paid." if is_paid else "Marked unpaid — Discord board updates in ~15s."))
    return _raffle_redirect(name, verb)


@app.get("/raffles/{name:path}")
async def raffle_detail(request: Request, name: str, msg: str = ""):
    gid, gname = _session_guild(request)
    if gid is None:
        return _redirect_login()
    async with request.app.state.pool.acquire() as conn:
        raffle = await conn.fetchrow(
            "SELECT * FROM raffles WHERE guild_id = $1 AND name = $2", gid, name)
        if not raffle:
            return _raffles_redirect(err=f"No raffle named “{name}”.")
        slots = await conn.fetch(
            """SELECT spot_num, user_id, username, paid FROM raffle_slots
               WHERE guild_id = $1 AND raffle_name = $2 ORDER BY spot_num""",
            gid, name)
        host = None
        if raffle["host_num"] in (1, 2):
            host = await conn.fetchrow(
                "SELECT host_num, name FROM raffle_hosts "
                "WHERE guild_id = $1 AND host_num = $2", gid, raffle["host_num"])
    slot_list = [dict(s) for s in slots]
    claimed = sum(1 for s in slot_list if s["user_id"] is not None)
    paid = sum(1 for s in slot_list if s["paid"])
    return templates.TemplateResponse(request, "raffle_detail.html", _ctx(
        request, gid, gname,
        raffle=dict(raffle), slots=slot_list, claimed=claimed, paid_count=paid,
        status_label=RAFFLE_STATUS_LABELS.get(raffle["status"], raffle["status"]),
        host=dict(host) if host else None, msg=msg,
    ))


# ── Excel export ──────────────────────────────────────────────────────────────

@app.get("/drops/{drop_number}/export.xlsx")
async def drop_export(request: Request, drop_number: int):
    gid, gname = _session_guild(request)
    if gid is None:
        return _redirect_login()
    async with request.app.state.pool.acquire() as conn:
        orders = await _load_drop_orders(conn, gid, drop_number)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Drop {drop_number}"
    header_fill = PatternFill("solid", fgColor="1E1E2E")
    header_font = Font(bold=True, color="FFFFFF")
    headers = ["Buyer", "User ID", "Item", "Qty", "Subtotal",
               "Order Total", "Paid", "Tracking #"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center")

    for o in orders:
        first = True
        for item in o["items"]:
            ws.append([
                o["name"] if first else "",
                str(o["user_id"]) if first else "",
                item["display"],
                item["qty"],
                round(item["subtotal"], 2),
                round(o["total"], 2) if first else "",
                ("Yes" if o["confirmed"] else "No") if first else "",
                o["tracking"] if first else "",
            ])
            first = False

    for col_cells in ws.columns:
        width = max((len(str(c.value)) for c in col_cells if c.value is not None),
                    default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(width + 3, 45)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_name = (gname or str(gid)).replace(" ", "_")
    filename = f"{safe_name}_Drop_{drop_number}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/healthz")
async def healthz():
    return {"ok": True}
