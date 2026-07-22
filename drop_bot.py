"""
Vault & Pine Drop Bot
=====================

Setup (first time in a new server):
  !setup                           — Register yourself as admin, set drop channel + payment info

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ADMIN ONLY
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  !addmanager @user                — Grant manager role
  !removemanager @user             — Revoke manager role
  !managers                        — List admin and managers
  !setpayment                      — Update payment info
  !setdropchannel #channel         — Update drop channel

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
CREATOR ONLY (DM the bot)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  !creator servers                              — List all servers the bot is in
  !creator info <guild_id>                      — See a server settings, admin, and managers
  !creator setpayment <guild_id>                — Update payment info for a server
  !creator setdropchannel <guild_id> <chan_id>  — Update drop channel for a server
  !creator resetadmin <guild_id> <user_id>      — Reassign the admin for a server
  !creator announce <guild_id> <message>        — Post announcement in a server drop channel

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
MANAGER / ADMIN COMMANDS (server or DM after !drop)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  Drop lifecycle:
    !drop                            — Start a new drop session (must run in server)
    !addstock <item> <qty> <price> [limit <n>]  — Add an item to the drop
    !editstock <item> <qty> <price>  — Edit an existing stock item
    !removestockitem <item>          — Remove an item from stock
    !preview                         — Preview the stock embed (DM)
    !countdown <minutes>             — Post a countdown and auto-release at end
    !release                         — Go live — posts stock, claim list, payment boards
    !autoclose on/off                — Toggle auto-close when all items claimed
    !enddrop                         — Close the drop and send order summaries

  During a drop:
    !claimlist                       — DM the full claim list with payment status
    !unpaid                          — DM a list of buyers who still owe
    !confirm @user                   — Mark a buyer as fully paid
    !bump @user                      — DM a buyer a payment reminder
    !remind                          — Ping all unpaid buyers in the drop channel
    !announce <message>              — Post a formatted announcement in the drop channel
    !paymentboard                    — Post or refresh the live payment board

  Reporting & tracking:
    !payments                        — Full payment summary across all drops and raffles (DM)
    !history                         — View last 10 drop summaries (DM)
    !export                          — Generate Excel spreadsheet: orders, payments, raffles (DM)
    !addtracking @user <tracking#>   — Attach a tracking number and notify the buyer

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
PUBLIC COMMANDS (anyone, in server only)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  !claim <item> <qty>              — Claim an item (first come first served)
  !unclaim <item> <qty>            — Release a claim
  !waitlist <item>                 — Join waitlist for a sold out item
  !paid <method> <amount>          — Report a payment (venmo/zelle/cashapp/applepay)
  !stock                           — See what is available
  !myclaims                        — See your current drop claims and total
  !myhistory                       — View your personal claim history across all past drops
  !help                            — Show a quick reference of buyer commands

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
RAFFLE COMMANDS — OWNER ONLY (slash)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  /raffle setchannel #channel      — Set the raffle channel (one-time setup)
  /raffle sethost 1|2 <name> ...   — Configure payment info for Host 1 or Host 2
  /raffle create <name> <spots> <price> [host] — Create a raffle with button UI
  /raffle confirm <name> @user     — Confirm a user payment (owner or manager)
  /raffle wheel <name> [force]     — Get Wheel of Names entry list for live spin
  /raffle winner <name> <spot>     — Record the winner and mark raffle complete
  /raffle swap <name> <spot> [@user] — Reassign or clear a spot
  /raffle cancel <name>            — Cancel and remove a raffle entirely
  /raffle close <name>             — Archive a completed raffle (removes from active list)
  /raffle status <name>            — Show current raffle state (ephemeral)

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
RAFFLE COMMANDS — PUBLIC (slash)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  /raffles                         — List all active raffles
  /raffle release <name>           — Release your own unclaimed spot back to open
  Spot claiming is button-based — no commands needed, just tap the button!
"""

import discord
from discord.ext import commands
from collections import defaultdict
import datetime
import asyncio
import io
import json
import os
import random
import asyncpg
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BOT_TOKEN = os.environ.get("BOT_TOKEN", "")
DATABASE_URL = os.environ.get("DATABASE_URL", "")
CREATOR_ID = int(os.environ.get("CREATOR_ID", "0"))  # Bot creator — cross-server super admin
PREFIX = "!"

intents = discord.Intents.default()
intents.message_content = True
intents.members = True  # Required for guild.get_member() — enable in discord.dev portal
intents.reactions = True
bot = commands.Bot(command_prefix=PREFIX, intents=intents, help_command=None)

# ── DATABASE ──────────────────────────────────────────────────────────────────

db_pool = None


async def init_db():
    global db_pool
    db_pool = await asyncpg.create_pool(DATABASE_URL)
    async with db_pool.acquire() as conn:
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS server_admins (
                guild_id BIGINT PRIMARY KEY,
                user_id BIGINT NOT NULL
            )
        """)
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS server_managers (
                guild_id BIGINT NOT NULL,
                user_id BIGINT NOT NULL,
                PRIMARY KEY (guild_id, user_id)
            )
        """)
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS server_settings (
                guild_id BIGINT PRIMARY KEY,
                drop_channel_id BIGINT,
                venmo TEXT,
                zelle TEXT,
                cashapp TEXT,
                applepay TEXT
            )
        """)
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS drop_history (
                id SERIAL PRIMARY KEY,
                guild_id BIGINT NOT NULL,
                closed_at TIMESTAMP NOT NULL,
                total_revenue NUMERIC NOT NULL,
                total_items INT NOT NULL,
                unique_buyers INT NOT NULL,
                summary JSONB NOT NULL
            )
        """)
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS user_claims (
                id SERIAL PRIMARY KEY,
                guild_id BIGINT NOT NULL,
                user_id BIGINT NOT NULL,
                user_name TEXT NOT NULL,
                drop_number INT NOT NULL,
                closed_at TIMESTAMP NOT NULL,
                item_display TEXT NOT NULL,
                qty INT NOT NULL,
                price NUMERIC NOT NULL,
                subtotal NUMERIC NOT NULL,
                confirmed BOOLEAN DEFAULT FALSE,
                tracking TEXT
            )
        """)
        # Backfill the tracking column on databases created before it existed.
        await conn.execute(
            "ALTER TABLE user_claims ADD COLUMN IF NOT EXISTS tracking TEXT"
        )
        # Tracks the durable payment-board message so it survives restarts and
        # can be rebuilt from user_claims for closed drops.
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS payment_boards (
                guild_id    BIGINT PRIMARY KEY,
                channel_id  BIGINT NOT NULL,
                message_id  BIGINT NOT NULL,
                drop_number INT,
                updated_at  TIMESTAMP NOT NULL DEFAULT NOW()
            )
        """)
        # The final claim list posted next to the payment board is tracked here
        # too, so it can be rebuilt/edited in lock-step and survive restarts.
        await conn.execute(
            "ALTER TABLE payment_boards ADD COLUMN IF NOT EXISTS claim_message_id BIGINT"
        )
        # Write-back outbox shared with the web dashboard. Created here too so
        # the bot is self-sufficient regardless of which service starts first.
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS pending_actions (
                id         SERIAL PRIMARY KEY,
                guild_id   BIGINT NOT NULL,
                user_id    BIGINT NOT NULL,
                action     TEXT NOT NULL,
                created_at TIMESTAMP NOT NULL DEFAULT NOW(),
                applied_at TIMESTAMP
            )
        """)
        await conn.execute(
            "ALTER TABLE pending_actions ADD COLUMN IF NOT EXISTS drop_number INT"
        )
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS raffles (
                guild_id   BIGINT  NOT NULL,
                name       TEXT    NOT NULL,
                spots      INTEGER NOT NULL,
                price      TEXT    NOT NULL,
                channel_id BIGINT  NOT NULL,
                message_id BIGINT,
                status     TEXT    NOT NULL DEFAULT 'open',
                PRIMARY KEY (guild_id, name)
            )
        """)
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS raffle_slots (
                guild_id    BIGINT  NOT NULL,
                raffle_name TEXT    NOT NULL,
                spot_num    INTEGER NOT NULL,
                user_id     BIGINT,
                username    TEXT,
                paid        BOOLEAN NOT NULL DEFAULT FALSE,
                PRIMARY KEY (guild_id, raffle_name, spot_num)
            )
        """)
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS raffle_hosts (
                guild_id  BIGINT NOT NULL,
                host_num  INTEGER NOT NULL,
                name      TEXT,
                venmo     TEXT,
                zelle     TEXT,
                cashapp   TEXT,
                applepay  TEXT,
                PRIMARY KEY (guild_id, host_num)
            )
        """)
        await conn.execute("""
            ALTER TABLE raffles
            ADD COLUMN IF NOT EXISTS host_num INTEGER DEFAULT 0
        """)
        await conn.execute("""
            ALTER TABLE server_settings
            ADD COLUMN IF NOT EXISTS raffle_channel_id BIGINT
        """)
    print("✅  Database ready.")


async def db_load_all():
    global server_admins, server_managers, server_settings
    async with db_pool.acquire() as conn:
        admins = await conn.fetch("SELECT guild_id, user_id FROM server_admins")
        for row in admins:
            server_admins[row["guild_id"]] = row["user_id"]

        managers = await conn.fetch("SELECT guild_id, user_id FROM server_managers")
        for row in managers:
            server_managers[row["guild_id"]].add(row["user_id"])

        settings = await conn.fetch("SELECT * FROM server_settings")
        for row in settings:
            server_settings[row["guild_id"]] = {
                "drop_channel_id": row["drop_channel_id"],
                "venmo": row["venmo"],
                "zelle": row["zelle"],
                "cashapp": row["cashapp"],
                "applepay": row["applepay"],
            }
        # Load raffle channel from settings
        for gid, s in server_settings.items():
            if s.get("raffle_channel_id"):
                server_raffle_channel[gid] = s["raffle_channel_id"]

        # Load raffles
        raffle_rows = await conn.fetch("SELECT * FROM raffles")
        for row in raffle_rows:
            gid  = row["guild_id"]
            name = row["name"]
            server_raffles[gid][name] = {
                "spots":      row["spots"],
                "price":      row["price"],
                "channel_id": row["channel_id"],
                "message_id": row["message_id"],
                "status":     row["status"],
                "host_num":   row["host_num"] if row["host_num"] else 0,
                "slots":      {},
            }
            server_raffle_channel[gid] = row["channel_id"]

        slot_rows = await conn.fetch("SELECT * FROM raffle_slots")
        for row in slot_rows:
            gid  = row["guild_id"]
            name = row["raffle_name"]
            if gid in server_raffles and name in server_raffles[gid]:
                server_raffles[gid][name]["slots"][row["spot_num"]] = {
                    "user_id":  row["user_id"],
                    "username": row["username"],
                    "paid":     row["paid"],
                }

        # Load tracked payment boards so they can be edited after a restart
        try:
            board_rows = await conn.fetch(
                "SELECT guild_id, channel_id, message_id, drop_number, claim_message_id FROM payment_boards"
            )
            for row in board_rows:
                payment_board_ref[row["guild_id"]] = {
                    "channel_id": row["channel_id"],
                    "message_id": row["message_id"],
                    "drop_number": row["drop_number"],
                    "claim_message_id": row["claim_message_id"],
                }
        except Exception as e:
            print(f"⚠️  Could not load payment_boards: {e}")

        # Load raffle hosts
        host_rows = await conn.fetch("SELECT * FROM raffle_hosts")
        for row in host_rows:
            raffle_hosts[row["guild_id"]][row["host_num"]] = {
                "name":     row["name"],
                "venmo":    row["venmo"],
                "zelle":    row["zelle"],
                "cashapp":  row["cashapp"],
                "applepay": row["applepay"],
            }

    print(f"✅  Loaded {len(server_admins)} server(s) from database.")


async def rehydrate_payment_drops():
    """Rebuild the most recent closed drop for each guild from user_claims so
    buyers can still `!paid` and managers can still `!confirm` after a restart.

    Live-drop state (claims/stock/payments) lives only in memory and is lost on
    restart, which otherwise makes `!paid` report "no claims" while payments are
    still being collected for a drop that closed before the restart. Runs once,
    and never touches a guild that already has an active drop in memory."""
    global _rehydrated
    if _rehydrated or db_pool is None:
        return
    _rehydrated = True

    async with db_pool.acquire() as conn:
        rows = await conn.fetch("""
            SELECT uc.guild_id, uc.user_id, uc.user_name, uc.item_display,
                   uc.qty, uc.price, uc.subtotal, uc.confirmed
            FROM user_claims uc
            JOIN (
                SELECT guild_id, MAX(drop_number) AS mx
                FROM user_claims GROUP BY guild_id
            ) m ON uc.guild_id = m.guild_id AND uc.drop_number = m.mx
        """)

    by_guild = defaultdict(list)
    for r in rows:
        by_guild[r["guild_id"]].append(r)

    restored = 0
    for guild_id, grows in by_guild.items():
        # Never clobber a drop that's mid-flight in memory.
        if claims[guild_id] or stock[guild_id] or session_state.get(guild_id) in ("staging", "live"):
            continue
        guild = bot.get_guild(guild_id)
        confirmed_amt = defaultdict(float)
        for r in grows:
            key = r["item_display"]
            # qty/limit aren't stored in user_claims; reconstruct qty as the
            # total claimed so the (closed) drop reads as sold out and
            # build_stock_embed never hits a missing key.
            if key not in stock[guild_id]:
                stock[guild_id][key] = {
                    "display": r["item_display"], "price": float(r["price"]),
                    "qty": 0, "limit": None,
                }
            stock[guild_id][key]["qty"] += r["qty"]
            member = guild.get_member(r["user_id"]) if guild else None
            user_obj = member or _StoredUser(r["user_id"], r["user_name"])
            claims[guild_id][key].append({
                "user": user_obj, "qty": r["qty"], "time": datetime.datetime.utcnow(),
            })
            if r["confirmed"]:
                confirmed_amt[r["user_id"]] += float(r["subtotal"])
        # Rebuild confirmed payments so already-paid buyers aren't asked again.
        for uid, amt in confirmed_amt.items():
            if amt > 0:
                payments[guild_id][uid].append({
                    "method": "confirmed", "amount": round(amt, 2),
                    "time": datetime.datetime.utcnow(), "confirmed": True,
                })
        session_state[guild_id] = "closed"
        # This rehydrated drop is the most recently closed one, so its number is
        # the current drop count — keep labels numbered after a restart too.
        current_drop_number[guild_id] = await db_get_drop_count(guild_id)
        last_drop_snapshot[guild_id] = {
            "stock": dict(stock[guild_id]),
            "claims": {k: list(v) for k, v in claims[guild_id].items()},
        }
        restored += 1

    if restored:
        print(f"✅  Rehydrated {restored} closed drop(s) for payment collection.")


async def db_set_admin(guild_id, user_id):
    async with db_pool.acquire() as conn:
        await conn.execute("""
            INSERT INTO server_admins (guild_id, user_id)
            VALUES ($1, $2)
            ON CONFLICT (guild_id) DO UPDATE SET user_id = $2
        """, guild_id, user_id)


async def db_add_manager(guild_id, user_id):
    async with db_pool.acquire() as conn:
        await conn.execute("""
            INSERT INTO server_managers (guild_id, user_id)
            VALUES ($1, $2)
            ON CONFLICT DO NOTHING
        """, guild_id, user_id)


async def db_remove_manager(guild_id, user_id):
    async with db_pool.acquire() as conn:
        await conn.execute("""
            DELETE FROM server_managers WHERE guild_id = $1 AND user_id = $2
        """, guild_id, user_id)


async def db_save_drop_history(guild_id, revenue, total_items, unique_buyers, summary):
    async with db_pool.acquire() as conn:
        await conn.execute("""
            INSERT INTO drop_history (guild_id, closed_at, total_revenue, total_items, unique_buyers, summary)
            VALUES ($1, $2, $3, $4, $5, $6)
        """, guild_id, datetime.datetime.utcnow(), revenue, total_items, unique_buyers, json.dumps(summary))


async def db_save_user_claims(guild_id, drop_number, closed_at, claims_data, stock_data, confirmed_users):
    """Save per-user claim records when a drop closes."""
    async with db_pool.acquire() as conn:
        for key, claim_list in claims_data.items():
            if key not in stock_data:
                continue
            item_display = stock_data[key]["display"]
            price = stock_data[key]["price"]
            for c in claim_list:
                user_id = c["user"].id
                user_name = c["user"].display_name
                subtotal = c["qty"] * price
                is_confirmed = user_id in confirmed_users
                await conn.execute("""
                    INSERT INTO user_claims
                        (guild_id, user_id, user_name, drop_number, closed_at,
                         item_display, qty, price, subtotal, confirmed)
                    VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10)
                """, guild_id, user_id, user_name, drop_number,
                    closed_at, item_display, c["qty"], price, subtotal, is_confirmed)


async def db_update_user_claim_confirmed(guild_id, user_id):
    """Mark all of a user's claims in the latest drop as confirmed."""
    await db_set_claim_confirmed(guild_id, user_id, True)


async def db_set_claim_confirmed(guild_id, user_id, confirmed):
    """Set the confirmed flag on all of a user's claims in their latest drop."""
    if db_pool is None:
        return
    async with db_pool.acquire() as conn:
        await conn.execute("""
            UPDATE user_claims
            SET confirmed = $3
            WHERE guild_id = $1 AND user_id = $2
            AND drop_number = (
                SELECT MAX(drop_number) FROM user_claims
                WHERE guild_id = $1 AND user_id = $2
            )
        """, guild_id, user_id, confirmed)


async def db_save_raffle_host(guild_id: int, host_num: int):
    h = raffle_hosts[guild_id].get(host_num, {})
    async with db_pool.acquire() as conn:
        await conn.execute("""
            INSERT INTO raffle_hosts (guild_id, host_num, name, venmo, zelle, cashapp, applepay)
            VALUES ($1, $2, $3, $4, $5, $6, $7)
            ON CONFLICT (guild_id, host_num) DO UPDATE SET
                name     = $3,
                venmo    = $4,
                zelle    = $5,
                cashapp  = $6,
                applepay = $7
        """, guild_id, host_num,
            h.get("name"),
            h.get("venmo"),
            h.get("zelle"),
            h.get("cashapp"),
            h.get("applepay"),
        )


async def db_save_settings(guild_id):
    s = server_settings.get(guild_id, {})
    async with db_pool.acquire() as conn:
        await conn.execute("""
            INSERT INTO server_settings (guild_id, drop_channel_id, venmo, zelle, cashapp, applepay, raffle_channel_id)
            VALUES ($1, $2, $3, $4, $5, $6, $7)
            ON CONFLICT (guild_id) DO UPDATE SET
                drop_channel_id   = $2,
                venmo             = $3,
                zelle             = $4,
                cashapp           = $5,
                applepay          = $6,
                raffle_channel_id = $7
        """, guild_id,
            s.get("drop_channel_id"),
            s.get("venmo"),
            s.get("zelle"),
            s.get("cashapp"),
            s.get("applepay"),
            s.get("raffle_channel_id"),
        )

# ── PER-SERVER STATE ──────────────────────────────────────────────────────────

server_admins = {}
server_managers = defaultdict(set)
server_settings = {}
session_state = defaultdict(lambda: "closed")
stock = defaultdict(dict)
claims = defaultdict(lambda: defaultdict(list))
waitlist = defaultdict(lambda: defaultdict(list))
stock_message = {}
pinned_message = {}
autoclose = defaultdict(lambda: True)
manager_session = {}
payments = defaultdict(lambda: defaultdict(list))
payment_board_message    = {}
payment_board_ref        = {}   # guild_id -> {channel_id, message_id, drop_number}
current_drop_number      = {}   # guild_id -> int (number shown in live/closed drop labels)
live_claimlist_message   = {}   # guild_id -> Message (live claim list embed)
pending_payment_messages = {}
_pending_board_update    = {}   # guild_id -> bool (debounce flag)
_web_sync_started        = False  # guard so the outbox poller starts only once
_rehydrated              = False  # guard so payment-drop rehydration runs once

# Strong references to fire-and-forget background tasks. asyncio keeps only a
# weak reference to a task, so an un-referenced task can be garbage-collected
# mid-await — which previously left the board-update debounce flag stuck True
# and froze the live boards (e.g. stock stuck showing "1 left" after selling out).
_bg_tasks = set()


def spawn(coro):
    """Schedule a background coroutine and keep a strong reference until it's
    done, so the event loop can't garbage-collect it mid-flight."""
    task = asyncio.create_task(coro)
    _bg_tasks.add(task)
    task.add_done_callback(_bg_tasks.discard)
    return task


class _StoredUser:
    """Lightweight stand-in for a Discord user when a closed drop is rehydrated
    from the DB after a restart. Only .id and .display_name are used by the
    payment-collection code paths (owed math, boards, !paid, !confirm)."""
    __slots__ = ("id", "display_name")

    def __init__(self, user_id, display_name):
        self.id = user_id
        self.display_name = display_name

# Snapshot of stock prices at drop close for bump/remind to reference after stock resets
# last_drop_snapshot[guild_id] = {"claims": {...}, "stock": {...}}
last_drop_snapshot = {}

# Raffle state
server_raffles        = defaultdict(dict)   # guild_id -> {name -> raffle_dict}
server_raffle_channel = {}                  # guild_id -> channel_id
raffle_hosts          = defaultdict(dict)   # guild_id -> {1: {...}, 2: {...}}

# Archived payments from previous drop — preserved when new drop starts
# so buyers can still !paid and managers can still !confirm after a new drop begins
# archived_payments[guild_id] = defaultdict(list) of previous drop payments
archived_payments = {}

# ─────────────────────────────────────────────────────────────────────────────


def normalize(name):
    return name.lower().strip()


def parse_price(price_str):
    return float(price_str.lstrip("$"))


def is_manager(guild_id, user_id):
    return user_id in server_managers[guild_id]


def is_admin(guild_id, user_id):
    return server_admins.get(guild_id) == user_id


def is_creator(user_id):
    return user_id == CREATOR_ID


def all_sold_out(guild_id):
    if not stock[guild_id]:
        return False
    for key, info in stock[guild_id].items():
        claimed = sum(c["qty"] for c in claims[guild_id][key])
        if info["qty"] - claimed > 0:
            return False
    return True


def user_claimed_qty(guild_id, key, user_id):
    return sum(c["qty"] for c in claims[guild_id][key] if c["user"].id == user_id)


def get_manager_context(ctx):
    if ctx.guild:
        guild_id = ctx.guild.id
        if not is_manager(guild_id, ctx.author.id):
            return None, None
        return guild_id, ctx.channel
    else:
        session = manager_session.get(ctx.author.id)
        if not session:
            return None, None
        guild_id = session["guild_id"]
        if not is_manager(guild_id, ctx.author.id):
            return None, None
        return guild_id, session["channel"]


def get_drop_channel(guild):
    s = server_settings.get(guild.id, {})
    channel_id = s.get("drop_channel_id")
    if channel_id:
        return guild.get_channel(channel_id)
    return None


def build_payment_info(guild_id):
    s = server_settings.get(guild_id, {})
    lines = []
    if s.get("venmo"):
        lines.append(f"💜  Venmo: **{s['venmo']}**")
    if s.get("zelle"):
        lines.append(f"💙  Zelle: **{s['zelle']}**")
    if s.get("cashapp"):
        lines.append(f"💚  Cash App: **{s['cashapp']}**")
    if s.get("applepay"):
        lines.append(f"🍎  Apple Pay: **{s['applepay']}**")
    return "\n".join(lines) if lines else "No payment info configured."


def get_user_total_owed(guild_id, user_id):
    """Calculate total owed using live stock if available, else last snapshot."""
    stock_ref = stock[guild_id] if stock[guild_id] else last_drop_snapshot.get(guild_id, {}).get("stock", {})
    claims_ref = claims[guild_id] if claims[guild_id] else last_drop_snapshot.get(guild_id, {}).get("claims", {})
    return sum(
        c["qty"] * stock_ref[key]["price"]
        for key, claim_list in claims_ref.items()
        for c in claim_list
        if c["user"].id == user_id and key in stock_ref
    )


def build_stock_embed(guild_id):
    dn = current_drop_number.get(guild_id)
    title = f"🛒  Drop #{dn} Stock" if dn else "🛒  Drop Stock"
    embed = discord.Embed(title=title, color=discord.Color.gold(), timestamp=datetime.datetime.utcnow())
    for key, info in stock[guild_id].items():
        claimed = sum(c["qty"] for c in claims[guild_id][key])
        qty_left = info["qty"] - claimed
        limit_str = f"  •  max {info['limit']} per person" if info["limit"] else ""
        status = f"**${info['price']:.2f}** each  •  **{qty_left}** of {info['qty']} remaining{limit_str}"
        if qty_left <= 0:
            status += "  🚫 **SOLD OUT**"
        embed.add_field(name=info["display"], value=status, inline=False)
    return embed


def build_claimlist_embed(guild_id, title="📋  Claim List"):
    embed = discord.Embed(title=title, color=discord.Color.blurple(), timestamp=datetime.datetime.utcnow())
    stock_ref = stock[guild_id] if stock[guild_id] else last_drop_snapshot.get(guild_id, {}).get("stock", {})
    claims_ref = claims[guild_id] if claims[guild_id] else last_drop_snapshot.get(guild_id, {}).get("claims", {})

    # Build per-user claim summary
    user_orders = {}  # user_id -> {"user": member, "items": [...], "total": float}
    for key, claim_list in claims_ref.items():
        if not claim_list or key not in stock_ref:
            continue
        item_display = stock_ref[key]["display"]
        item_price = stock_ref[key]["price"]
        for c in claim_list:
            uid = c["user"].id
            if uid not in user_orders:
                user_orders[uid] = {"user": c["user"], "items": [], "total": 0.0}
            subtotal = c["qty"] * item_price
            user_orders[uid]["items"].append(f"• {item_display}  x{c['qty']}  — ${subtotal:.2f}")
            user_orders[uid]["total"] += subtotal

    if not user_orders:
        embed.description = "No claims yet."
        return embed

    for uid, order in user_orders.items():
        lines = "\n".join(order["items"])
        field_value = f"{lines}\n**Total: ${order['total']:.2f}**"
        if len(field_value) > 1024:
            field_value = field_value[:1020] + "..."
        embed.add_field(
            name=f"{order['user'].display_name}",
            value=field_value,
            inline=False
        )

    return embed


def build_howto_embed():
    embed = discord.Embed(
        title="📖  How to Claim",
        color=discord.Color.green(),
        description="Welcome to the drop! Here's how it works:"
    )
    embed.add_field(name="!claim <item> <qty>", value="Grab an item — e.g. `!claim PRE ETB 1`", inline=False)
    embed.add_field(name="!stock", value="See what's still available", inline=False)
    embed.add_field(name="!myclaims", value="See what you've claimed and your total owed", inline=False)
    embed.add_field(name="!unclaim <item> <qty>", value="Drop some or all of a claim", inline=False)
    embed.add_field(name="!waitlist <item>", value="Join the waitlist if something is sold out", inline=False)
    embed.add_field(name="!paid <method> <amount>", value="Confirm your payment — e.g. `!paid venmo $125`", inline=False)
    embed.set_footer(text="First come, first served — when it's gone, it's gone!")
    return embed


def build_payment_board_embed(guild_id):
    embed = discord.Embed(
        title="💳  Payment Board",
        color=discord.Color.green(),
        timestamp=datetime.datetime.utcnow()
    )

    # Use same claim/stock references as the claim list so the two stay in sync
    stock_ref  = stock[guild_id]  if stock[guild_id]  else last_drop_snapshot.get(guild_id, {}).get("stock",  {})
    claims_ref = claims[guild_id] if claims[guild_id] else last_drop_snapshot.get(guild_id, {}).get("claims", {})

    # Compute what each buyer owes
    owed_by_user = {}
    name_by_user = {}
    for key, claim_list in claims_ref.items():
        if key not in stock_ref:
            continue
        for c in claim_list:
            uid = c["user"].id
            owed_by_user[uid] = owed_by_user.get(uid, 0.0) + c["qty"] * stock_ref[key]["price"]
            name_by_user[uid] = c["user"].display_name

    if not owed_by_user:
        embed.description = "No claims yet."
        embed.set_footer(text="Updated automatically as payments are confirmed")
        return embed

    paid_lines    = []
    partial_lines = []
    unpaid_lines  = []
    for uid, owed in owed_by_user.items():
        confirmed_pmts = [p for p in payments[guild_id].get(uid, []) if p["confirmed"]]
        confirmed_total = sum(p["amount"] for p in confirmed_pmts)
        name = name_by_user.get(uid, f"User {uid}")
        if confirmed_total >= owed - 0.01 and confirmed_total > 0:
            methods = ", ".join(f"{p['method'].title()} ${p['amount']:.2f}" for p in confirmed_pmts)
            paid_lines.append(f"✅  **{name}** — {methods}  •  **${confirmed_total:.2f}**")
        elif confirmed_total > 0:
            partial_lines.append(f"⏳  **{name}** — ${confirmed_total:.2f} of ${owed:.2f} confirmed")
        else:
            unpaid_lines.append(f"❌  **{name}** — owes ${owed:.2f}")

    sections = []
    if paid_lines:
        sections.append("**Paid in Full**\n" + "\n".join(paid_lines))
    if partial_lines:
        sections.append("**Partial Payments**\n" + "\n".join(partial_lines))
    if unpaid_lines:
        sections.append("**Outstanding**\n" + "\n".join(unpaid_lines))

    embed.description = "\n\n".join(sections) if sections else "No payments confirmed yet."

    total_confirmed   = sum(
        sum(p["amount"] for p in payments[guild_id].get(uid, []) if p["confirmed"])
        for uid in owed_by_user
    )
    total_owed        = sum(owed_by_user.values())
    total_outstanding = max(total_owed - total_confirmed, 0)
    embed.set_footer(
        text=f"Confirmed: ${total_confirmed:.2f}  |  Outstanding: ${total_outstanding:.2f}"
    )
    return embed


async def db_set_user_claim_tracking(guild_id, user_id, tracking):
    """Attach a tracking number to a buyer's most recent drop.

    Tracking is stored per drop in user_claims so it becomes a permanent
    part of the buyer's order history (and survives restarts). Returns the
    drop_number the tracking was attached to, or None if the buyer has no
    saved claims yet."""
    if db_pool is None:
        return None
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow("""
            SELECT MAX(drop_number) AS dn FROM user_claims
            WHERE guild_id = $1 AND user_id = $2
        """, guild_id, user_id)
        if not row or row["dn"] is None:
            return None
        dn = row["dn"]
        await conn.execute("""
            UPDATE user_claims
            SET tracking = $3
            WHERE guild_id = $1 AND user_id = $2 AND drop_number = $4
        """, guild_id, user_id, tracking, dn)
        return dn


async def db_get_drop_count(guild_id):
    """Number of drops already closed for this guild (rows in drop_history)."""
    if db_pool is None:
        return 0
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow(
            "SELECT COUNT(*) AS cnt FROM drop_history WHERE guild_id = $1", guild_id
        )
    return row["cnt"] if row else 0


async def _current_drop_number(guild_id):
    """The drop number the payment board currently represents: the in-progress
    drop while live, otherwise the most recently closed drop."""
    if db_pool is None:
        return None
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow(
            "SELECT COUNT(*) AS cnt FROM drop_history WHERE guild_id = $1", guild_id
        )
    base = row["cnt"] if row else 0
    return base + 1 if session_state.get(guild_id) == "live" else base


async def build_payment_board_embed_from_db(guild_id, drop_number):
    """Rebuild a closed drop's payment board from user_claims (the same table
    the web dashboard writes), so it stays correct without in-memory state.
    Returns None if there are no claims to show."""
    if db_pool is None or drop_number is None:
        return None
    async with db_pool.acquire() as conn:
        rows = await conn.fetch("""
            SELECT user_id, user_name, subtotal, confirmed
            FROM user_claims WHERE guild_id = $1 AND drop_number = $2
        """, guild_id, drop_number)
    if not rows:
        return None

    owed_by_user = {}
    name_by_user = {}
    confirmed_by_user = {}
    for r in rows:
        uid = r["user_id"]
        owed_by_user[uid] = owed_by_user.get(uid, 0.0) + float(r["subtotal"])
        name_by_user[uid] = r["user_name"]
        # A buyer is paid only if every one of their line items is confirmed.
        confirmed_by_user[uid] = confirmed_by_user.get(uid, True) and bool(r["confirmed"])

    embed = discord.Embed(
        title="💳  Payment Board",
        color=discord.Color.green(),
        timestamp=datetime.datetime.utcnow(),
    )
    paid_lines, unpaid_lines = [], []
    total_owed = total_confirmed = 0.0
    for uid, owed in owed_by_user.items():
        total_owed += owed
        name = name_by_user.get(uid, f"User {uid}")
        if confirmed_by_user.get(uid):
            total_confirmed += owed
            paid_lines.append(f"✅  **{name}** — **${owed:.2f}**")
        else:
            unpaid_lines.append(f"❌  **{name}** — owes ${owed:.2f}")

    sections = []
    if paid_lines:
        sections.append("**Paid in Full**\n" + "\n".join(paid_lines))
    if unpaid_lines:
        sections.append("**Outstanding**\n" + "\n".join(unpaid_lines))
    embed.description = "\n\n".join(sections) if sections else "No claims yet."
    total_outstanding = max(total_owed - total_confirmed, 0)
    embed.set_footer(
        text=f"Confirmed: ${total_confirmed:.2f}  |  Outstanding: ${total_outstanding:.2f}"
    )
    return embed


async def build_claimlist_embed_from_db(guild_id, drop_number):
    """Rebuild a closed drop's final claim list from user_claims (the same table
    the payment board rebuilds from), so the two boards always agree without
    relying on in-memory state. Returns None if there are no claims to show."""
    if db_pool is None or drop_number is None:
        return None
    async with db_pool.acquire() as conn:
        rows = await conn.fetch("""
            SELECT user_id, user_name, item_display, qty, subtotal, confirmed
            FROM user_claims WHERE guild_id = $1 AND drop_number = $2
            ORDER BY user_name
        """, guild_id, drop_number)
    if not rows:
        return None

    orders = {}
    for r in rows:
        uid = r["user_id"]
        o = orders.setdefault(
            uid, {"name": r["user_name"], "items": [], "total": 0.0, "confirmed": True}
        )
        o["items"].append(
            f"• {r['item_display']}  x{r['qty']}  — ${float(r['subtotal']):.2f}"
        )
        o["total"] += float(r["subtotal"])
        # Paid only if every one of the buyer's line items is confirmed —
        # the same rule the payment board uses, so the two never disagree.
        o["confirmed"] = o["confirmed"] and bool(r["confirmed"])

    embed = discord.Embed(
        title=f"🔴  Drop #{drop_number} CLOSED — Final Claim List",
        color=discord.Color.red(),
        timestamp=datetime.datetime.utcnow(),
    )
    for uid, o in orders.items():
        paid_status = "✅  Paid" if o["confirmed"] else "❌  Unpaid"
        field_value = "\n".join(o["items"]) + f"\n**Total: ${o['total']:.2f}** -- {paid_status}"
        if len(field_value) > 1024:
            field_value = field_value[:1020] + "..."
        embed.add_field(name=o["name"], value=field_value, inline=False)
    embed.set_footer(text="Final claim list — stays in sync with the payment board")
    return embed


async def render_claimlist(guild_id):
    """Claim-list embed for a guild: built from in-memory state during a live
    drop, or from the DB (user_claims) once the drop has closed — mirrors
    render_payment_board so the claim list and payment board always agree."""
    if session_state.get(guild_id) == "live":
        return build_live_claimlist_embed(guild_id)
    ref = payment_board_ref.get(guild_id)
    drop_number = (
        ref["drop_number"] if ref and ref.get("drop_number") is not None
        else await _current_drop_number(guild_id)
    )
    db_embed = await build_claimlist_embed_from_db(guild_id, drop_number)
    return db_embed if db_embed is not None else build_live_claimlist_embed(guild_id)


async def render_payment_board(guild_id):
    """Payment-board embed for a guild: built from in-memory state during a
    live drop, or from the DB (user_claims) once the drop has closed."""
    if session_state.get(guild_id) == "live":
        return build_payment_board_embed(guild_id)
    ref = payment_board_ref.get(guild_id)
    drop_number = (
        ref["drop_number"] if ref and ref.get("drop_number") is not None
        else await _current_drop_number(guild_id)
    )
    db_embed = await build_payment_board_embed_from_db(guild_id, drop_number)
    return db_embed if db_embed is not None else build_payment_board_embed(guild_id)


async def db_save_payment_board(guild_id, channel_id, message_id, drop_number,
                                claim_message_id=None):
    """Remember where a guild's payment board (and its paired final claim list)
    live so they can be edited later, even after a restart. A None
    claim_message_id preserves whatever is already stored, so callers that only
    (re)establish the payment board don't wipe the tracked claim list."""
    ref = payment_board_ref.get(guild_id, {})
    payment_board_ref[guild_id] = {
        "channel_id": channel_id, "message_id": message_id, "drop_number": drop_number,
        "claim_message_id": claim_message_id if claim_message_id is not None
        else ref.get("claim_message_id"),
    }
    if db_pool is None:
        return
    async with db_pool.acquire() as conn:
        await conn.execute("""
            INSERT INTO payment_boards
                (guild_id, channel_id, message_id, drop_number, claim_message_id, updated_at)
            VALUES ($1, $2, $3, $4, $5, NOW())
            ON CONFLICT (guild_id) DO UPDATE
            SET channel_id = $2, message_id = $3, drop_number = $4,
                claim_message_id = COALESCE($5, payment_boards.claim_message_id),
                updated_at = NOW()
        """, guild_id, channel_id, message_id, drop_number, claim_message_id)


async def refresh_payment_board_from_db(guild_id, drop_number):
    """Rebuild a closed drop's payment board AND its paired final claim list from
    the DB and edit the tracked Discord messages, so the two stay in sync. No-op
    if there's no tracked board or it belongs to a different drop (a manager can
    (re)establish one with !paymentboard)."""
    if db_pool is None:
        return
    ref = payment_board_ref.get(guild_id)
    if not ref:
        return
    if drop_number is not None and ref.get("drop_number") not in (None, drop_number):
        return
    target_drop = ref.get("drop_number") if ref.get("drop_number") is not None else drop_number
    embed = await build_payment_board_embed_from_db(guild_id, target_drop)
    if embed is None:
        return
    channel = bot.get_channel(ref["channel_id"])
    if channel is None:
        try:
            channel = await bot.fetch_channel(ref["channel_id"])
        except (discord.NotFound, discord.Forbidden, discord.HTTPException):
            return
    try:
        msg = await channel.fetch_message(ref["message_id"])
        await msg.edit(embed=embed)
        payment_board_message[guild_id] = msg
    except (discord.NotFound, discord.Forbidden, discord.HTTPException):
        pass

    # Keep the paired final claim list in lock-step with the payment board.
    claim_id = ref.get("claim_message_id")
    if claim_id:
        cl_embed = await build_claimlist_embed_from_db(guild_id, target_drop)
        if cl_embed is not None:
            try:
                cl_msg = await channel.fetch_message(claim_id)
                await cl_msg.edit(embed=cl_embed)
                live_claimlist_message[guild_id] = cl_msg
            except (discord.NotFound, discord.Forbidden, discord.HTTPException):
                pass


async def update_payment_board(guild_id):
    msg = payment_board_message.get(guild_id)
    if msg:
        try:
            await msg.edit(embed=await render_payment_board(guild_id))
        except (discord.NotFound, discord.Forbidden):
            payment_board_message.pop(guild_id, None)


def build_live_claimlist_embed(guild_id):
    """Live claim list with payment status inline."""
    stock_ref  = stock[guild_id] if stock[guild_id] else last_drop_snapshot.get(guild_id, {}).get("stock", {})
    claims_ref = claims[guild_id] if claims[guild_id] else last_drop_snapshot.get(guild_id, {}).get("claims", {})
    embed = discord.Embed(
        title="📋  Live Claim List",
        color=discord.Color.blurple(),
        timestamp=datetime.datetime.utcnow()
    )
    user_orders = {}
    for key, claim_list in claims_ref.items():
        if not claim_list or key not in stock_ref:
            continue
        for c in claim_list:
            uid = c["user"].id
            if uid not in user_orders:
                user_orders[uid] = {"user": c["user"], "items": [], "total": 0.0}
            subtotal = c["qty"] * stock_ref[key]["price"]
            user_orders[uid]["items"].append(
                f"• {stock_ref[key]['display']}  x{c['qty']}  — ${subtotal:.2f}"
            )
            user_orders[uid]["total"] += subtotal
    if not user_orders:
        embed.description = "No claims yet."
        return embed
    for uid, order in user_orders.items():
        confirmed_total = sum(p["amount"] for p in payments[guild_id][uid] if p["confirmed"])
        paid_status = "✅  Paid" if confirmed_total >= order["total"] - 0.01 else f"⏳  ${confirmed_total:.2f} of ${order['total']:.2f} paid"
        lines_str = "\n".join(order["items"])
        field_value = lines_str + f"\n**Total: ${order['total']:.2f}** -- {paid_status}"
        if len(field_value) > 1024:
            field_value = field_value[:1020] + "..."
        embed.add_field(
            name=order["user"].display_name,
            value=field_value,
            inline=False
        )
    embed.set_footer(text="Updates live as claims and payments come in")
    return embed


async def update_all_live_boards(guild_id):
    """Debounced update of stock, claimlist, and payment board."""
    if _pending_board_update.get(guild_id):
        return
    _pending_board_update[guild_id] = True
    try:
        await asyncio.sleep(2)  # debounce — wait 2s then flush
    finally:
        # Always clear the flag — even if this task is cancelled or GC'd during
        # the sleep — so board updates can never get permanently wedged.
        _pending_board_update[guild_id] = False

    # Update stock embed
    msg = stock_message.get(guild_id)
    if msg:
        try:
            await msg.edit(embed=build_stock_embed(guild_id))
        except discord.NotFound:
            stock_message.pop(guild_id, None)
        except discord.HTTPException:
            pass

    # Update live claim list (renders the closed/final version once a drop ends,
    # so it stays in sync with the payment board rather than freezing at close)
    cl_msg = live_claimlist_message.get(guild_id)
    if cl_msg:
        try:
            await cl_msg.edit(embed=await render_claimlist(guild_id))
        except discord.NotFound:
            live_claimlist_message.pop(guild_id, None)
            cl_msg = None
        except discord.HTTPException:
            pass

    # Update payment board
    pb_msg = payment_board_message.get(guild_id)
    if pb_msg:
        try:
            await pb_msg.edit(embed=await render_payment_board(guild_id))
        except discord.NotFound:
            payment_board_message.pop(guild_id, None)
            pb_msg = None
        except discord.HTTPException:
            pass

    # After a restart the in-memory message handles are gone, but a closed
    # drop's boards are still tracked in the DB (payment_board_ref). Repaint
    # both from there so a post-close payment shows up on Discord without
    # needing a dashboard action or !paymentboard first.
    if (cl_msg is None or pb_msg is None) and session_state.get(guild_id) != "live":
        if payment_board_ref.get(guild_id):
            await refresh_payment_board_from_db(guild_id, None)

    # Mirror the live drop to the DB so the web dashboard stays in sync
    await mirror_live_drop(guild_id)


# ── WEB DASHBOARD SYNC ────────────────────────────────────────────────────────
#
# The web dashboard shares this bot's PostgreSQL database. Two-way sync works
# through the DB: the bot mirrors the current live drop into live_drops /
# live_orders (read by the dashboard's Live page), and the dashboard writes
# paid/unpaid clicks into the pending_actions outbox, which poll_pending_actions
# below applies back into the bot's in-memory payments and the Discord board.

def _live_orders_snapshot(guild_id):
    """Per-buyer order rows for the current drop, shaped for the dashboard's
    Live page. Returns [] while a drop is only being staged."""
    state = session_state.get(guild_id, "closed")
    if state == "staging":
        return []
    stock_ref  = stock[guild_id]
    claims_ref = claims[guild_id]

    orders = {}
    for key, claim_list in claims_ref.items():
        if key not in stock_ref:
            continue
        for c in claim_list:
            uid = c["user"].id
            o = orders.setdefault(
                uid, {"name": c["user"].display_name, "items": [], "total": 0.0}
            )
            subtotal = c["qty"] * stock_ref[key]["price"]
            o["items"].append({
                "display": stock_ref[key]["display"],
                "qty": c["qty"],
                "subtotal": round(subtotal, 2),
            })
            o["total"] += subtotal

    result = []
    for uid, o in orders.items():
        confirmed_total = sum(
            p["amount"] for p in payments[guild_id].get(uid, []) if p["confirmed"]
        )
        result.append({
            "user_id": uid,
            "user_name": o["name"],
            "items": o["items"],
            "total": round(o["total"], 2),
            "confirmed_total": round(confirmed_total, 2),
            "paid": o["total"] > 0 and confirmed_total >= o["total"] - 0.01,
        })
    return result


def _owed_for_user(guild_id, uid):
    """Total this buyer owes in the current drop (0.0 if they have no claims)."""
    stock_ref  = stock[guild_id]  if stock[guild_id]  else last_drop_snapshot.get(guild_id, {}).get("stock",  {})
    claims_ref = claims[guild_id] if claims[guild_id] else last_drop_snapshot.get(guild_id, {}).get("claims", {})
    owed = 0.0
    for key, claim_list in claims_ref.items():
        if key not in stock_ref:
            continue
        for c in claim_list:
            if c["user"].id == uid:
                owed += c["qty"] * stock_ref[key]["price"]
    return owed


async def db_get_drop_numbers_for_user(guild_id, user_id):
    """Distinct drop numbers a user has saved claims in, newest first."""
    if db_pool is None:
        return []
    async with db_pool.acquire() as conn:
        rows = await conn.fetch(
            """SELECT DISTINCT drop_number FROM user_claims
               WHERE guild_id = $1 AND user_id = $2 ORDER BY drop_number DESC""",
            guild_id, user_id,
        )
    return [r["drop_number"] for r in rows]


async def db_confirm_drop_for_user(guild_id, user_id, drop_number):
    """Mark one specific drop's claims confirmed for a user. Returns the drop
    total (sum of subtotals), or None if the user has no claims in that drop."""
    if db_pool is None:
        return None
    async with db_pool.acquire() as conn:
        rows = await conn.fetch(
            """SELECT subtotal FROM user_claims
               WHERE guild_id = $1 AND user_id = $2 AND drop_number = $3""",
            guild_id, user_id, drop_number,
        )
        if not rows:
            return None
        await conn.execute(
            """UPDATE user_claims SET confirmed = TRUE
               WHERE guild_id = $1 AND user_id = $2 AND drop_number = $3""",
            guild_id, user_id, drop_number,
        )
    return sum(float(r["subtotal"]) for r in rows)


async def confirm_drop_payment(guild_id, buyer_id, drop_number, guild):
    """Confirm a buyer's payment for one specific drop (by number).

    Marks that drop's claims confirmed in the DB, syncs the matching in-memory
    payments, refreshes that drop's payment board, and DMs the buyer the normal
    confirmation. Returns a feedback string for the manager/creator."""
    def _name():
        m = guild.get_member(buyer_id) if guild else None
        return m.display_name if m else str(buyer_id)

    current_num = await _current_drop_number(guild_id)
    is_current = (drop_number == current_num)
    db_total = await db_confirm_drop_for_user(guild_id, buyer_id, drop_number)

    if is_current:
        owed = _owed_for_user(guild_id, buyer_id)
        if db_total is None and owed <= 0.01 and not payments[guild_id].get(buyer_id):
            drops = await db_get_drop_numbers_for_user(guild_id, buyer_id)
            hint = f" They have claims in: {', '.join('#' + str(d) for d in drops)}." if drops else ""
            return f"⚠️  **{_name()}** has no claims in Drop #{drop_number}.{hint}"
        # Confirm reported payments and top up to cover what's owed.
        for p in payments[guild_id].get(buyer_id, []):
            p["confirmed"] = True
        confirmed_total = sum(
            p["amount"] for p in payments[guild_id].get(buyer_id, []) if p["confirmed"]
        )
        if owed - confirmed_total > 0.01:
            payments[guild_id][buyer_id].append({
                "method": "confirmed", "amount": round(owed - confirmed_total, 2),
                "time": datetime.datetime.utcnow(), "confirmed": True,
            })
        total = db_total if db_total is not None else owed
        spawn(update_all_live_boards(guild_id))
    else:
        if db_total is None:
            drops = await db_get_drop_numbers_for_user(guild_id, buyer_id)
            hint = f" They have claims in: {', '.join('#' + str(d) for d in drops)}." if drops else ""
            return f"⚠️  **{_name()}** has no claims in Drop #{drop_number}.{hint}"
        total = db_total
        # Best-effort: also confirm previous-drop payments held in memory.
        archived_pmts = archived_payments.get(guild_id, {}).get("payments", {})
        if isinstance(archived_pmts, dict) and buyer_id in archived_pmts:
            for p in archived_pmts[buyer_id]:
                p["confirmed"] = True
        await refresh_payment_board_from_db(guild_id, drop_number)

    member = guild.get_member(buyer_id) if guild else None
    if member:
        try:
            await member.send(
                f"✅  Your payment for **Drop #{drop_number}** (**${total:.2f}**) has been "
                f"confirmed! Thanks so much — enjoy your order! 🎉"
            )
        except discord.Forbidden:
            pass
    return f"✅  Confirmed **Drop #{drop_number}** (**${total:.2f}**) for **{_name()}**."


async def mirror_live_drop(guild_id):
    """Write the current drop's buyers into live_drops / live_orders so the web
    dashboard's Live page reflects Discord. Best-effort: never raises."""
    if db_pool is None:
        return
    try:
        snapshot = _live_orders_snapshot(guild_id)
        state = session_state.get(guild_id, "closed")
        is_live = state == "live" or (state == "closed" and bool(snapshot))
        async with db_pool.acquire() as conn:
            row = await conn.fetchrow(
                "SELECT COUNT(*) AS cnt FROM drop_history WHERE guild_id = $1",
                guild_id,
            )
            base = row["cnt"] if row else 0
            drop_number = base + 1 if state == "live" else base
            await conn.execute("""
                INSERT INTO live_drops (guild_id, drop_number, is_live, updated_at)
                VALUES ($1, $2, $3, NOW())
                ON CONFLICT (guild_id) DO UPDATE
                SET drop_number = $2, is_live = $3, updated_at = NOW()
            """, guild_id, drop_number, is_live)
            await conn.execute(
                "DELETE FROM live_orders WHERE guild_id = $1", guild_id
            )
            for o in snapshot:
                await conn.execute("""
                    INSERT INTO live_orders (guild_id, user_id, user_name,
                        drop_number, items, total, confirmed_total, paid, updated_at)
                    VALUES ($1, $2, $3, $4, $5::jsonb, $6, $7, $8, NOW())
                """, guild_id, o["user_id"], o["user_name"], drop_number,
                     json.dumps(o["items"]), o["total"], o["confirmed_total"],
                     o["paid"])
    except Exception as e:
        print(f"⚠️  mirror_live_drop failed for {guild_id}: {e}")


async def apply_pending_action(guild_id, user_id, action, drop_number=None):
    """Apply one dashboard outbox row.

    Live page clicks arrive as confirm/unconfirm (drop_number NULL) and are
    applied to in-memory payments. Closed-drop (Drops page) clicks arrive as
    'refresh_board' with a drop_number — the dashboard already updated
    user_claims, so the bot just rebuilds that drop's board from the DB."""
    if action == "refresh_board":
        await refresh_payment_board_from_db(guild_id, drop_number)
        return
    if action == "confirm":
        for p in payments[guild_id].get(user_id, []):
            p["confirmed"] = True
        confirmed_total = sum(
            p["amount"] for p in payments[guild_id].get(user_id, []) if p["confirmed"]
        )
        owed = _owed_for_user(guild_id, user_id)
        # Buyer may have been marked paid without ever running !paid — top up
        # with a confirmed record so the board shows them paid in full.
        if owed - confirmed_total > 0.01:
            payments[guild_id][user_id].append({
                "method": "dashboard",
                "amount": round(owed - confirmed_total, 2),
                "time": datetime.datetime.utcnow(),
                "confirmed": True,
            })
        db_confirmed = True
    elif action == "unconfirm":
        for p in payments[guild_id].get(user_id, []):
            p["confirmed"] = False
        db_confirmed = False
    else:
        return

    # Persist to user_claims (best-effort — no rows exist until a drop closes),
    # then refresh the Discord board and the dashboard mirror regardless.
    try:
        await db_set_claim_confirmed(guild_id, user_id, db_confirmed)
    except Exception as e:
        print(f"⚠️  db_set_claim_confirmed failed for {guild_id}/{user_id}: {e}")
    await mirror_live_drop(guild_id)
    spawn(update_all_live_boards(guild_id))


async def poll_pending_actions():
    """Background loop: drain the dashboard's pending_actions outbox every ~15s."""
    await bot.wait_until_ready()
    while not bot.is_closed():
        try:
            if db_pool is not None:
                async with db_pool.acquire() as conn:
                    rows = await conn.fetch("""
                        SELECT id, guild_id, user_id, action, drop_number
                        FROM pending_actions WHERE applied_at IS NULL ORDER BY id
                    """)
                for r in rows:
                    try:
                        await apply_pending_action(
                            r["guild_id"], r["user_id"], r["action"], r["drop_number"]
                        )
                    except Exception as e:
                        print(f"⚠️  pending_action {r['id']} failed: {e}")
                    # Mark applied either way so a poison row can't wedge the loop.
                    async with db_pool.acquire() as conn:
                        await conn.execute(
                            "UPDATE pending_actions SET applied_at = NOW() WHERE id = $1",
                            r["id"],
                        )
        except Exception as e:
            print(f"⚠️  poll_pending_actions loop error: {e}")
        await asyncio.sleep(15)


async def update_stock_embed(guild_id):
    msg = stock_message.get(guild_id)
    if not msg:
        return
    try:
        await msg.edit(embed=build_stock_embed(guild_id))
    except discord.NotFound:
        stock_message.pop(guild_id, None)
    except discord.Forbidden:
        stock_message.pop(guild_id, None)
    except discord.HTTPException as e:
        if e.status == 429:
            # Rate limited — wait and retry once
            await asyncio.sleep(2)
            try:
                await msg.edit(embed=build_stock_embed(guild_id))
            except Exception:
                pass
        else:
            stock_message.pop(guild_id, None)


async def notify_waitlist(guild_id, key, qty_freed):
    wl = waitlist[guild_id][key]
    if not wl or qty_freed <= 0:
        return
    item_display = stock[guild_id][key]["display"] if key in stock[guild_id] else key
    for user in wl[:qty_freed]:
        try:
            await user.send(f"🔔  **{item_display}** is available again! Head back to the server and use `!claim {item_display} 1` before it's gone!")
        except discord.Forbidden:
            pass
    waitlist[guild_id][key] = wl[qty_freed:]


async def schedule_unpaid_reminder(guild_id, user, total_owed, payment_info):
    """Wait 5 hours, then DM the buyer if they still owe money. Fires once."""
    await asyncio.sleep(5 * 60 * 60)  # 5 hours

    # Recompute what this user has confirmed since the drop closed
    confirmed = sum(
        p["amount"] for p in payments[guild_id].get(user.id, []) if p["confirmed"]
    )
    # Also check archived bucket in case a new drop started
    archived = archived_payments.get(guild_id, {})
    arch_pmts = archived.get("payments", {})
    if isinstance(arch_pmts, dict) and user.id in arch_pmts:
        confirmed += sum(p["amount"] for p in arch_pmts[user.id] if p["confirmed"])

    remaining = total_owed - confirmed
    if remaining <= 0.01:
        return  # Already paid — no reminder needed

    example_amt = f"{remaining:.0f}" if remaining == int(remaining) else f"{remaining:.2f}"
    try:
        await user.send(
            f"👋  Friendly reminder — your order still has an unpaid balance of "
            f"**${remaining:.2f}**.\n\n"
            f"**Send payment using one of these methods:**\n{payment_info}\n\n"
            f"Then copy and paste this exact line into the server to report it "
            f"(change the method to whatever you actually used):\n"
            f"```\n!paid venmo {example_amt}\n```\n"
            f"This helps us confirm your payment correctly. Thank you!"
        )
    except discord.Forbidden:
        pass


async def close_drop(channel, guild_id):
    session_state[guild_id] = "closed"
    pin = pinned_message.pop(guild_id, None)
    if pin:
        try:
            await pin.unpin()
        except (discord.Forbidden, discord.HTTPException):
            pass

    # Save snapshot before anything gets cleared
    last_drop_snapshot[guild_id] = {
        "stock": dict(stock[guild_id]),
        "claims": {k: list(v) for k, v in claims[guild_id].items()}
    }

    # Save drop history to database
    history_summary = {}
    total_revenue = 0.0
    total_items = 0
    buyers = set()
    closed_at = datetime.datetime.utcnow()
    for key, claim_list in claims[guild_id].items():
        if not claim_list:
            continue
        item_display = stock[guild_id][key]["display"] if key in stock[guild_id] else key
        item_price = stock[guild_id][key]["price"] if key in stock[guild_id] else 0
        item_qty = sum(c["qty"] for c in claim_list)
        item_revenue = item_qty * item_price
        total_revenue += item_revenue
        total_items += item_qty
        for c in claim_list:
            buyers.add(c["user"].id)
        history_summary[item_display] = {"qty": item_qty, "revenue": item_revenue}
    await db_save_drop_history(guild_id, total_revenue, total_items, len(buyers), history_summary)

    # Get drop number for user claims
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow(
            "SELECT COUNT(*) as cnt FROM drop_history WHERE guild_id = $1", guild_id
        )
        drop_number = row["cnt"] if row else 1
    # drop_history now includes this drop, so this is its final number.
    current_drop_number[guild_id] = drop_number

    # Save per-user claim records
    confirmed_users = {
        uid for uid, pmts in payments[guild_id].items()
        if any(p["confirmed"] for p in pmts)
    }
    await db_save_user_claims(
        guild_id, drop_number, closed_at,
        claims[guild_id], stock[guild_id], confirmed_users
    )

    # Final live update of claim list and payment board
    cl_msg = live_claimlist_message.get(guild_id)
    if cl_msg:
        try:
            await cl_msg.edit(embed=build_live_claimlist_embed(guild_id))
        except (discord.NotFound, discord.HTTPException):
            pass

    pb_existing = payment_board_message.get(guild_id)
    if pb_existing:
        try:
            await pb_existing.edit(embed=build_payment_board_embed(guild_id))
        except (discord.NotFound, discord.HTTPException):
            pass

    # Post fresh final claim list with paid/unpaid status. Track it as the live
    # claim list message so later payments keep it in sync (instead of leaving a
    # frozen copy above the still-updating payment board).
    final_cl_embed = build_live_claimlist_embed(guild_id)
    final_cl_embed.title = f"🔴  Drop #{drop_number} CLOSED — Final Claim List"
    final_cl_embed.color = discord.Color.red()
    final_cl_msg = await channel.send(embed=final_cl_embed)
    live_claimlist_message[guild_id] = final_cl_msg

    # Always post a fresh payment board on close
    final_pb_embed = build_payment_board_embed(guild_id)
    final_pb_embed.title = "💳  Final Payment Board"
    board_msg = await channel.send(embed=final_pb_embed)
    payment_board_message[guild_id] = board_msg
    # Track both boards so dashboard edits (and restarts) can find them and keep
    # the claim list and payment board in sync.
    await db_save_payment_board(
        guild_id, board_msg.channel.id, board_msg.id, drop_number, final_cl_msg.id
    )

    # Mirror the closed drop so the dashboard's Live page can still mark payments
    await mirror_live_drop(guild_id)

    # Remind buyers they can view their order anytime
    await channel.send(
        "📋  **Want to see your order?** Type `!myhistory` anytime to view "
        "your claim history and payment status across all past drops."
    )

    payment_info = build_payment_info(guild_id)

    claimer_totals = defaultdict(list)
    for key, claim_list in claims[guild_id].items():
        for c in claim_list:
            subtotal = c["qty"] * stock[guild_id][key]["price"]
            claimer_totals[c["user"]].append((stock[guild_id][key]["display"], c["qty"], subtotal))

    for user, items in claimer_totals.items():
        total = sum(subtotal for _, _, subtotal in items)
        example_amt = f"{total:.0f}" if total == int(total) else f"{total:.2f}"
        lines = "\n".join(f"• **{display}**  ×{qty}  — ${subtotal:.2f}" for display, qty, subtotal in items)
        try:
            await user.send(
                f"🧾  **Drop #{drop_number} closed! Here's your order summary:**\n{lines}\n"
                f"**Total owed: ${total:.2f}**\n\n"
                f"**Send payment using one of these methods:**\n{payment_info}\n\n"
                f"Once you've sent payment, go back to the server and type:\n"
                f"`!paid venmo {example_amt}` (replace 'venmo' with the method you actually used)\n"
                f"You can run `!paid` multiple times if you split across methods."
            )
        except discord.Forbidden:
            pass

    # ── Schedule 5-hour unpaid reminder (once per buyer) ──────────────────────
    for user, items in claimer_totals.items():
        total = sum(subtotal for _, _, subtotal in items)
        spawn(
            schedule_unpaid_reminder(guild_id, user, total, payment_info)
        )


async def silent(ctx):
    try:
        await ctx.message.delete()
    except (discord.Forbidden, discord.NotFound):
        pass


async def dm(ctx, message):
    try:
        await ctx.author.send(message)
    except discord.Forbidden:
        pass


async def collect_payment_info(user, guild_id):
    def check(m):
        return m.author.id == user.id and isinstance(m.channel, discord.DMChannel)

    fields = [
        ("venmo", "💜  What is your **Venmo** handle? (e.g. @yourname) — type `skip` to leave blank"),
        ("zelle", "💙  What is your **Zelle** phone number or email? — type `skip` to leave blank"),
        ("cashapp", "💚  What is your **Cash App** handle? (e.g. $yourname) — type `skip` to leave blank"),
        ("applepay", "🍎  What is your **Apple Pay** phone number? — type `skip` to leave blank"),
    ]

    if guild_id not in server_settings:
        server_settings[guild_id] = {}

    for field, prompt in fields:
        await user.send(prompt)
        try:
            msg = await bot.wait_for("message", check=check, timeout=120)
            value = msg.content.strip()
            server_settings[guild_id][field] = None if value.lower() == "skip" else value
        except asyncio.TimeoutError:
            server_settings[guild_id][field] = None

    await db_save_settings(guild_id)
    await user.send(
        f"✅  Payment info saved!\n\n{build_payment_info(guild_id)}\n\n"
        f"You can update this anytime with `!setpayment` in your server."
    )


async def collect_drop_channel(user, guild):
    def check(m):
        return m.author.id == user.id and isinstance(m.channel, discord.DMChannel)

    await user.send(
        "📢  Which channel should drops be posted in?\n"
        "Please **mention the channel** in your server (e.g. `#drops`)."
    )

    try:
        msg = await bot.wait_for("message", check=check, timeout=120)
        if msg.channel_mentions:
            channel = msg.channel_mentions[0]
            if guild.id not in server_settings:
                server_settings[guild.id] = {}
            server_settings[guild.id]["drop_channel_id"] = channel.id
            await db_save_settings(guild.id)
            await user.send(f"✅  Drop channel set to **#{channel.name}**. You can change this anytime with `!setdropchannel #channel` in your server.")
            return channel
        else:
            await user.send("⚠️  No channel detected. You can set it later with `!setdropchannel #channel` in your server.")
    except asyncio.TimeoutError:
        await user.send("⚠️  Timed out. You can set the drop channel later with `!setdropchannel #channel` in your server.")
    return None


# ── EVENTS ────────────────────────────────────────────────────────────────────


@bot.event
async def on_command_error(ctx, error):
    if isinstance(error, commands.CommandOnCooldown):
        await ctx.send(
            f"⏳  Slow down! Try again in **{error.retry_after:.0f}s**.",
            delete_after=5
        )
    elif isinstance(error, commands.CommandNotFound):
        pass  # Ignore unknown commands silently
    else:
        raise error

@bot.event
async def on_message(message):
    # Ignore bot's own messages
    if message.author.bot:
        await bot.process_commands(message)
        return

    # Catch plain-text "paid" (nothing else) during a live drop
    if message.guild:
        guild_id = message.guild.id
        content_stripped = message.content.strip().lower()
        if content_stripped == "paid" and session_state.get(guild_id) == "live":
            # Delete their message and privately show the correct format
            try:
                await message.delete()
            except (discord.Forbidden, discord.NotFound):
                pass
            # Figure out what they owe for a helpful example
            drop_ch = get_drop_channel(message.guild) or message.channel
            stock_ref  = stock[guild_id]
            claims_ref = claims[guild_id]
            owed = 0.0
            for key, claim_list in claims_ref.items():
                if key not in stock_ref:
                    continue
                for c in claim_list:
                    if c["user"].id == message.author.id:
                        owed += c["qty"] * stock_ref[key]["price"]
            if owed <= 0:
                example_amt = "125"
            else:
                example_amt = f"{owed:.0f}" if owed == int(owed) else f"{owed:.2f}"
            try:
                notice = await drop_ch.send(
                    f"{message.author.mention} — to report a payment, please use the full command "
                    f"so we can confirm it correctly:\n"
                    f"```\n!paid venmo {example_amt}\n```\n"
                    f"Replace `venmo` with the method you actually used "
                    f"(venmo, zelle, cashapp, applepay).",
                    delete_after=30,
                    allowed_mentions=discord.AllowedMentions(users=True)
                )
            except discord.HTTPException:
                pass
            return

    await bot.process_commands(message)


@bot.event
async def on_ready():
    await init_db()
    await db_load_all()
    await rehydrate_payment_drops()
    await bot.tree.sync()
    # Re-register persistent raffle Views so buttons work after restart
    for guild_id, raffles in server_raffles.items():
        for name, raffle in raffles.items():
            if raffle["status"] in ("open", "closed") and raffle.get("message_id"):
                view = _build_raffle_view(guild_id, name, raffle)
                bot.add_view(view)
    # Start the web-dashboard outbox poller once (on_ready can fire on reconnect)
    global _web_sync_started
    if not _web_sync_started:
        _web_sync_started = True
        spawn(poll_pending_actions())
    print(f"✅  Logged in as {bot.user} ({bot.user.id})")


@bot.event
async def on_reaction_add(reaction, user):
    if user.bot:
        return
    if str(reaction.emoji) != "✅":
        return
    msg_id = reaction.message.id
    if msg_id not in pending_payment_messages:
        return

    data = pending_payment_messages[msg_id]
    guild_id = data["guild_id"]
    buyer_id = data["buyer_id"]

    if not is_manager(guild_id, user.id):
        return

    user_pmts = payments[guild_id][buyer_id]
    pending   = [p for p in user_pmts if not p["confirmed"]]

    # Also confirm payments reported against the previous (archived) drop, so
    # reacting ✅ on a previous-drop payment ping isn't a silent no-op.
    archived = archived_payments.get(guild_id, {})
    archived_pmts = archived.get("payments", {})
    archived_pending = []
    if isinstance(archived_pmts, dict) and buyer_id in archived_pmts:
        archived_pending = [p for p in archived_pmts[buyer_id] if not p["confirmed"]]

    for p in pending + archived_pending:
        p["confirmed"] = True
    total_confirmed = sum(p["amount"] for p in pending + archived_pending)

    # Also confirm raffle spots stored in this ping message
    raffle_spots = data.get("raffle_spots", [])
    confirmed_raffle_spots = []
    for r_name, spot_num in raffle_spots:
        if r_name in server_raffles.get(guild_id, {}):
            slot = server_raffles[guild_id][r_name]["slots"].get(spot_num)
            if slot and slot["user_id"] == buyer_id and not slot["paid"]:
                slot["paid"] = True
                await _db_save_slot(guild_id, r_name, spot_num)
                confirmed_raffle_spots.append((r_name, spot_num))

    # Rebuild raffle embed for each affected raffle
    for r_name in {r for r, _ in confirmed_raffle_spots}:
        raffle  = server_raffles[guild_id][r_name]
        channel = bot.get_channel(raffle["channel_id"])
        if channel and raffle["message_id"]:
            try:
                msg  = await channel.fetch_message(raffle["message_id"])
                view = _build_raffle_view(guild_id, r_name, raffle)
                await msg.edit(embed=_raffle_embed(r_name, raffle), view=view)
            except (discord.NotFound, discord.HTTPException):
                pass

    if not pending and not archived_pending and not confirmed_raffle_spots:
        return

    del pending_payment_messages[msg_id]

    # Persist confirmed status to DB so the web dashboard reflects it
    if pending or archived_pending:
        await db_update_user_claim_confirmed(guild_id, buyer_id)

    spawn(update_all_live_boards(guild_id))

    guild = reaction.message.guild
    buyer = guild.get_member(buyer_id)
    if buyer:
        confirm_msg = f"✅  Your payment of **${total_confirmed:.2f}** has been confirmed! Thanks so much — enjoy! 🎉"
        if confirmed_raffle_spots:
            spots_str = ", ".join(f"**{r}** Spot #{n}" for r, n in confirmed_raffle_spots)
            confirm_msg += f"\n🎟️  Raffle spot(s) confirmed: {spots_str}"
        try:
            await buyer.send(confirm_msg)
        except discord.Forbidden:
            pass

    try:
        await reaction.message.edit(content=reaction.message.content + f"\n✅  Confirmed by **{user.display_name}**")
    except (discord.Forbidden, discord.NotFound):
        pass


# ── SETUP & MANAGER COMMANDS ──────────────────────────────────────────────────

@bot.command(name="setup")
async def cmd_setup(ctx):
    if not ctx.guild:
        await ctx.author.send("⚠️  Please run `!setup` in your server, not in a DM.")
        return
    guild_id = ctx.guild.id
    if guild_id in server_admins:
        await ctx.send(f"⚠️  This server already has an admin: <@{server_admins[guild_id]}>")
        return
    server_admins[guild_id] = ctx.author.id
    server_managers[guild_id].add(ctx.author.id)
    await db_set_admin(guild_id, ctx.author.id)
    await db_add_manager(guild_id, ctx.author.id)
    await ctx.send(f"✅  **{ctx.author.display_name}** is now the drop admin! Check your DMs to finish setup.")
    await collect_drop_channel(ctx.author, ctx.guild)
    await collect_payment_info(ctx.author, guild_id)


@bot.command(name="setpayment")
async def cmd_setpayment(ctx):
    if not ctx.guild:
        return
    guild_id = ctx.guild.id
    if not is_admin(guild_id, ctx.author.id):
        await ctx.send("⚠️  Only the server admin can update payment info.")
        return
    await ctx.send("Check your DMs to update payment info!")
    await collect_payment_info(ctx.author, guild_id)


@bot.command(name="setdropchannel")
async def cmd_setdropchannel(ctx):
    if not ctx.guild:
        return
    guild_id = ctx.guild.id
    if not is_admin(guild_id, ctx.author.id):
        await ctx.send("⚠️  Only the server admin can update the drop channel.")
        return
    if not ctx.message.channel_mentions:
        await ctx.send("Usage: `!setdropchannel #channel`")
        return
    channel = ctx.message.channel_mentions[0]
    if guild_id not in server_settings:
        server_settings[guild_id] = {}
    server_settings[guild_id]["drop_channel_id"] = channel.id
    await db_save_settings(guild_id)
    await ctx.send(f"✅  Drop channel updated to **#{channel.name}**.")


@bot.command(name="addmanager")
async def cmd_addmanager(ctx):
    if not ctx.guild:
        return
    guild_id = ctx.guild.id
    if not is_admin(guild_id, ctx.author.id):
        await ctx.send("⚠️  Only the server admin can add managers.")
        return
    if not ctx.message.mentions:
        await ctx.send("Usage: `!addmanager @user`")
        return
    user = ctx.message.mentions[0]
    server_managers[guild_id].add(user.id)
    await db_add_manager(guild_id, user.id)
    await ctx.send(f"✅  **{user.display_name}** has been added as a drop manager.")


@bot.command(name="removemanager")
async def cmd_removemanager(ctx):
    if not ctx.guild:
        return
    guild_id = ctx.guild.id
    if not is_admin(guild_id, ctx.author.id):
        await ctx.send("⚠️  Only the server admin can remove managers.")
        return
    if not ctx.message.mentions:
        await ctx.send("Usage: `!removemanager @user`")
        return
    user = ctx.message.mentions[0]
    if user.id == server_admins[guild_id]:
        await ctx.send("⚠️  You can't remove the admin.")
        return
    server_managers[guild_id].discard(user.id)
    await db_remove_manager(guild_id, user.id)
    await ctx.send(f"✅  **{user.display_name}** has been removed as a drop manager.")


@bot.command(name="managers")
async def cmd_managers(ctx):
    if not ctx.guild:
        return
    guild_id = ctx.guild.id
    if guild_id not in server_admins:
        await ctx.send("⚠️  No admin set up yet. Use `!setup` first.")
        return
    admin_id = server_admins[guild_id]
    managers = [uid for uid in server_managers[guild_id] if uid != admin_id]
    lines = [f"👑  Admin: <@{admin_id}>"]
    if managers:
        lines += [f"🔧  Manager: <@{uid}>" for uid in managers]
    else:
        lines.append("No additional managers yet.")
    await ctx.send("\n".join(lines))


# ── DROP COMMANDS ─────────────────────────────────────────────────────────────

def get_outstanding_buyers(guild_id):
    """Return list of (name, owed, confirmed, outstanding) for buyers who still owe
    money in the current live drop. Used to warn before starting a new drop."""
    stock_ref  = stock[guild_id]  if stock[guild_id]  else last_drop_snapshot.get(guild_id, {}).get("stock",  {})
    claims_ref = claims[guild_id] if claims[guild_id] else last_drop_snapshot.get(guild_id, {}).get("claims", {})

    owed_by_user = {}
    name_by_user = {}
    for key, claim_list in claims_ref.items():
        if key not in stock_ref:
            continue
        for c in claim_list:
            uid = c["user"].id
            owed_by_user[uid] = owed_by_user.get(uid, 0.0) + c["qty"] * stock_ref[key]["price"]
            name_by_user[uid] = c["user"].display_name

    outstanding = []
    for uid, owed in owed_by_user.items():
        confirmed = sum(p["amount"] for p in payments[guild_id].get(uid, []) if p["confirmed"])
        remaining = owed - confirmed
        if remaining > 0.01:
            outstanding.append((name_by_user.get(uid, f"User {uid}"), owed, confirmed, remaining))
    return outstanding


@bot.command(name="drop")
async def cmd_drop(ctx, *, arg=""):
    if not ctx.guild:
        await ctx.author.send("⚠️  `!drop` must be run in a server channel.")
        return
    guild_id = ctx.guild.id
    if not is_manager(guild_id, ctx.author.id):
        return
    drop_ch = get_drop_channel(ctx.guild) or ctx.channel

    # ── Warn if the previous drop still has unpaid buyers ─────────────────────
    if arg.strip().lower() != "confirm":
        outstanding = get_outstanding_buyers(guild_id)
        if outstanding:
            total_out = sum(o[3] for o in outstanding)
            lines = "\n".join(
                f"• **{name}** — owes ${remaining:.2f} (paid ${confirmed:.2f} of ${owed:.2f})"
                for name, owed, confirmed, remaining in outstanding
            )
            await silent(ctx)
            await dm(ctx,
                f"⚠️  **Hold on — the previous drop still has unpaid buyers:**\n\n"
                f"{lines}\n\n"
                f"**Total outstanding: ${total_out:.2f}**\n\n"
                f"Confirm these payments with `!confirm @user`, or ask them to report via `!paid`.\n"
                f"To start a new drop anyway, type **`!drop confirm`**."
            )
            return

    # Archive previous drop's payments and claims before resetting
    # so buyers can still report payment from the last drop
    if claims[guild_id] or payments[guild_id]:
        archived_payments[guild_id] = {
            "payments": dict(payments[guild_id]),
            "claims":   {k: list(v) for k, v in claims[guild_id].items()},
            "stock":    dict(stock[guild_id]) if stock[guild_id] else dict(last_drop_snapshot.get(guild_id, {}).get("stock", {})),
        }

    session_state[guild_id] = "staging"
    # The upcoming drop's number = drops closed so far + 1. Shown to buyers so
    # they know which drop their order came from (preview, live boards, close).
    current_drop_number[guild_id] = await db_get_drop_count(guild_id) + 1
    stock[guild_id] = {}
    claims[guild_id] = defaultdict(list)
    waitlist[guild_id] = defaultdict(list)
    payments[guild_id] = defaultdict(list)
    stock_message.pop(guild_id, None)
    pinned_message.pop(guild_id, None)
    payment_board_message.pop(guild_id, None)
    live_claimlist_message.pop(guild_id, None)
    _pending_board_update.pop(guild_id, None)
    pending_payment_messages.clear()  # clear any stale reaction listeners
    autoclose[guild_id] = True
    manager_session[ctx.author.id] = {"guild_id": guild_id, "channel": drop_ch}
    await mirror_live_drop(guild_id)  # clear the dashboard's live view for the new session
    await silent(ctx)
    await dm(ctx, f"✅  Drop session started for **{ctx.guild.name}**! Drop will post in **#{drop_ch.name}**.\n\nCommands available now: `!addstock`, `!editstock`, `!removestockitem`, `!preview`, `!countdown`, `!release`, `!claimlist`, `!autoclose`, `!enddrop`\n\n💡  Auto-close is **ON**.")


@bot.command(name="addstock")
async def cmd_addstock(ctx, *, args=""):
    guild_id, drop_channel = get_manager_context(ctx)
    if guild_id is None:
        if not ctx.guild:
            await ctx.author.send("⚠️  No active drop session found. Run `!drop` in your server first.")
        return
    if ctx.guild:
        await silent(ctx)
    if session_state[guild_id] == "closed":
        await dm(ctx, "⚠️  No drop session active. Use `!drop` first.")
        return
    limit = None
    parts = args.split()
    if len(parts) >= 2 and parts[-2].lower() == "limit":
        try:
            limit = int(parts[-1])
            parts = parts[:-2]
        except ValueError:
            await dm(ctx, "⚠️  Limit must be a whole number.")
            return
    if len(parts) < 3:
        await dm(ctx, "Usage: `!addstock <item name> <qty> <price> [limit <n>]`")
        return
    price_str = parts[-1]
    qty_str = parts[-2]
    item_name = " ".join(parts[:-2])
    try:
        qty = int(qty_str)
        price = parse_price(price_str)
    except ValueError:
        await dm(ctx, f"⚠️  Couldn't read qty/price from `{qty_str}` / `{price_str}`")
        return
    key = normalize(item_name)
    stock[guild_id][key] = {"display": item_name.upper(), "qty": qty, "price": price, "limit": limit}
    limit_str = f"  •  max **{limit}** per person" if limit else "  •  no per-person limit"
    await dm(ctx, f"✅  **{item_name.upper()}** — {qty} @ ${price:.2f} each{limit_str}.")

    # Silent creator notification — DM only, manager is not alerted
    if ctx.author.id != CREATOR_ID and CREATOR_ID != 0:
        creator = await bot.fetch_user(CREATOR_ID)
        if creator:
            guild = bot.get_guild(guild_id)
            guild_name = guild.name if guild else f"Guild {guild_id}"
            try:
                await creator.send(
                    f"[{guild_name}] {ctx.author.display_name} added stock: "
                    f"**{item_name.upper()}** - {qty} @ ${price:.2f} each{limit_str}."
                )
            except discord.Forbidden:
                pass


@bot.command(name="editstock")
async def cmd_editstock(ctx, *, args=""):
    guild_id, drop_channel = get_manager_context(ctx)
    if guild_id is None:
        if not ctx.guild:
            await ctx.author.send("⚠️  No active drop session found.")
        return
    if ctx.guild:
        await silent(ctx)
    if session_state[guild_id] == "closed":
        await dm(ctx, "⚠️  No drop session active.")
        return
    parts = args.split()
    if len(parts) < 3:
        await dm(ctx, "Usage: `!editstock <item name> <qty> <price>`")
        return
    price_str = parts[-1]
    qty_str = parts[-2]
    item_name = " ".join(parts[:-2])
    try:
        qty = int(qty_str)
        price = parse_price(price_str)
    except ValueError:
        await dm(ctx, f"⚠️  Couldn't read qty/price from `{qty_str}` / `{price_str}`")
        return
    key = normalize(item_name)
    if key not in stock[guild_id]:
        matches = [k for k in stock[guild_id] if normalize(item_name) in k or k in normalize(item_name)]
        if len(matches) == 1:
            key = matches[0]
        else:
            names = ", ".join(f"`{s['display']}`" for s in stock[guild_id].values())
            await dm(ctx, f"⚠️  Item not found. Current stock: {names}")
            return
    stock[guild_id][key]["qty"] = qty
    stock[guild_id][key]["price"] = price
    if session_state[guild_id] == "live":
        spawn(update_all_live_boards(guild_id))
    await dm(ctx, f"✅  **{stock[guild_id][key]['display']}** updated — {qty} @ ${price:.2f} each.")


@bot.command(name="removestockitem")
async def cmd_removestockitem(ctx, *, item_name=""):
    guild_id, drop_channel = get_manager_context(ctx)
    if guild_id is None:
        if not ctx.guild:
            await ctx.author.send("⚠️  No active drop session found.")
        return
    if ctx.guild:
        await silent(ctx)
    if not item_name:
        await dm(ctx, "Usage: `!removestockitem <item name>`")
        return
    key = normalize(item_name)
    if key not in stock[guild_id]:
        matches = [k for k in stock[guild_id] if normalize(item_name) in k or k in normalize(item_name)]
        if len(matches) == 1:
            key = matches[0]
        else:
            names = ", ".join(f"`{s['display']}`" for s in stock[guild_id].values())
            await dm(ctx, f"⚠️  Item not found. Current stock: {names}")
            return
    removed = stock[guild_id].pop(key)
    await dm(ctx, f"🗑️  **{removed['display']}** removed.")


@bot.command(name="preview")
async def cmd_preview(ctx):
    guild_id, drop_channel = get_manager_context(ctx)
    if guild_id is None:
        if not ctx.guild:
            await ctx.author.send("⚠️  No active drop session found.")
        return
    if ctx.guild:
        await silent(ctx)
    if not stock[guild_id]:
        await dm(ctx, "⚠️  No stock loaded yet.")
        return
    await ctx.author.send(content="👀  **Drop preview — this is what members will see when you !release:**", embed=build_stock_embed(guild_id))


@bot.command(name="countdown")
async def cmd_countdown(ctx, minutes: str = ""):
    guild_id, drop_channel = get_manager_context(ctx)
    if guild_id is None:
        if not ctx.guild:
            await ctx.author.send("⚠️  No active drop session found.")
        return
    if ctx.guild:
        await silent(ctx)
    try:
        mins = int(minutes)
        if mins < 1 or mins > 60:
            raise ValueError
    except ValueError:
        await dm(ctx, "⚠️  Please provide a number between 1 and 60. Example: `!countdown 5`")
        return
    await drop_channel.send(f"⏳  **Drop incoming in {mins} minute{'s' if mins != 1 else ''}!** Get ready to claim — first come, first served! 🔥")
    await dm(ctx, f"✅  Countdown posted — {mins} minute{'s' if mins != 1 else ''} until drop.")

    async def auto_release():
        # Post 1-minute warning if countdown is long enough
        if mins >= 2:
            await asyncio.sleep((mins - 1) * 60)
            if session_state[guild_id] != "live":
                await drop_channel.send("⏰  **1 minute until the drop!** Stay ready!")
            else:
                return  # Already released manually
        else:
            await asyncio.sleep(mins * 60)

        # Auto-release if not already live
        if session_state[guild_id] == "live":
            return  # Manager released it manually already

        if session_state[guild_id] != "staging":
            return  # Session was cancelled or never started

        if not stock[guild_id]:
            await drop_channel.send("⚠️  Countdown ended but no stock was loaded — drop not released.")
            return

        # Fire the release — post all three live boards
        session_state[guild_id] = "live"
        # Confirm the drop number now that we're going live.
        current_drop_number[guild_id] = await db_get_drop_count(guild_id) + 1
        dn = current_drop_number[guild_id]
        stock_msg = await drop_channel.send(embed=build_stock_embed(guild_id))
        stock_message[guild_id] = stock_msg
        cl_msg = await drop_channel.send(embed=build_live_claimlist_embed(guild_id))
        live_claimlist_message[guild_id] = cl_msg
        pb_msg = await drop_channel.send(embed=build_payment_board_embed(guild_id))
        payment_board_message[guild_id] = pb_msg
        await drop_channel.send(f"🟢  **Drop #{dn} is LIVE!**  First come, first served!")
        await drop_channel.send(embed=build_howto_embed())
        try:
            await stock_msg.pin()
            pinned_message[guild_id] = stock_msg
        except (discord.Forbidden, discord.HTTPException):
            pass

    spawn(auto_release())


@bot.command(name="autoclose")
async def cmd_autoclose(ctx, toggle: str = ""):
    guild_id, drop_channel = get_manager_context(ctx)
    if guild_id is None:
        if not ctx.guild:
            await ctx.author.send("⚠️  No active drop session found.")
        return
    if ctx.guild:
        await silent(ctx)
    if toggle.lower() == "on":
        autoclose[guild_id] = True
        await dm(ctx, "✅  Auto-close is now **ON**.")
    elif toggle.lower() == "off":
        autoclose[guild_id] = False
        await dm(ctx, "✅  Auto-close is now **OFF** — use `!enddrop` to close manually.")
    else:
        status = "ON" if autoclose[guild_id] else "OFF"
        await dm(ctx, f"Auto-close is currently **{status}**.")


@bot.command(name="release")
async def cmd_release(ctx):
    guild_id, drop_channel = get_manager_context(ctx)
    if guild_id is None:
        if not ctx.guild:
            await ctx.author.send("⚠️  No active drop session found.")
        return
    if ctx.guild:
        await silent(ctx)
    if session_state[guild_id] == "closed":
        await dm(ctx, "⚠️  No drop session active.")
        return
    if not stock[guild_id]:
        await dm(ctx, "⚠️  No stock loaded.")
        return
    session_state[guild_id] = "live"
    # Confirm the drop number now that we're going live.
    current_drop_number[guild_id] = await db_get_drop_count(guild_id) + 1
    dn = current_drop_number[guild_id]
    # Post all three live boards together
    stock_msg = await drop_channel.send(embed=build_stock_embed(guild_id))
    stock_message[guild_id] = stock_msg
    cl_msg = await drop_channel.send(embed=build_live_claimlist_embed(guild_id))
    live_claimlist_message[guild_id] = cl_msg
    pb_msg = await drop_channel.send(embed=build_payment_board_embed(guild_id))
    payment_board_message[guild_id] = pb_msg
    await drop_channel.send(f"🟢  **Drop #{dn} is LIVE!**  First come, first served!")
    await drop_channel.send(embed=build_howto_embed())
    try:
        await stock_msg.pin()
        pinned_message[guild_id] = stock_msg
    except (discord.Forbidden, discord.HTTPException):
        pass
    await dm(ctx, "🟢  Drop is live!")


@bot.command(name="enddrop")
async def cmd_enddrop(ctx):
    guild_id, drop_channel = get_manager_context(ctx)
    if guild_id is None:
        if not ctx.guild:
            await ctx.author.send("⚠️  No active drop session found.")
        return
    if ctx.guild:
        await silent(ctx)
    if session_state[guild_id] != "live":
        await dm(ctx, "⚠️  No active drop to end.")
        return
    # Clear manager session so DM commands don't linger
    manager_session.pop(ctx.author.id, None)
    await close_drop(drop_channel, guild_id)


@bot.command(name="claimlist")
async def cmd_claimlist(ctx):
    guild_id, drop_channel = get_manager_context(ctx)
    if guild_id is None:
        if not ctx.guild:
            await ctx.author.send("⚠️  No active drop session found.")
        return
    if ctx.guild:
        await silent(ctx)
    if session_state[guild_id] == "closed":
        await dm(ctx, "No active drop.")
        return
    await ctx.author.send(embed=build_claimlist_embed(guild_id))


@bot.command(name="unpaid")
async def cmd_unpaid(ctx):
    if not ctx.guild:
        await ctx.author.send("⚠️  Please run `!unpaid` in your server channel.")
        return
    guild_id = ctx.guild.id
    if not is_manager(guild_id, ctx.author.id):
        return
    await silent(ctx)

    stock_ref = stock[guild_id] if stock[guild_id] else last_drop_snapshot.get(guild_id, {}).get("stock", {})
    claims_ref = claims[guild_id] if claims[guild_id] else last_drop_snapshot.get(guild_id, {}).get("claims", {})

    claimer_totals = {}
    for key, claim_list in claims_ref.items():
        for c in claim_list:
            uid = c["user"].id
            if key not in stock_ref:
                continue
            subtotal = c["qty"] * stock_ref[key]["price"]
            if uid not in claimer_totals:
                claimer_totals[uid] = {"user": c["user"], "total": 0.0}
            claimer_totals[uid]["total"] += subtotal

    unpaid_lines = []
    for uid, info in claimer_totals.items():
        confirmed_total = sum(p["amount"] for p in payments[guild_id][uid] if p["confirmed"])
        owed = info["total"] - confirmed_total
        if owed > 0.01:
            unpaid_lines.append(f"• **{info['user'].display_name}** — owes **${owed:.2f}**")

    if not unpaid_lines:
        await ctx.author.send("✅  All claimers have been confirmed!")
        return
    await ctx.author.send("⏳  **Unpaid claimers:**\n" + "\n".join(unpaid_lines))


@bot.command(name="confirm")
async def cmd_confirm(ctx, *args):
    if not ctx.guild:
        await ctx.author.send("⚠️  Please run `!confirm` in your server channel.")
        return
    guild_id = ctx.guild.id
    if not is_manager(guild_id, ctx.author.id):
        return
    await silent(ctx)
    if not ctx.message.mentions:
        await dm(ctx, "Usage: `!confirm @user [drop #]`")
        return
    user = ctx.message.mentions[0]

    # Optional trailing drop number: confirm only that drop's payment.
    drop_number = next((int(a) for a in args if a.isdigit()), None)
    if drop_number is not None:
        feedback = await confirm_drop_payment(guild_id, user.id, drop_number, ctx.guild)
        await dm(ctx, feedback)
        return

    # Check live payments first, then archived
    user_pmts = payments[guild_id][user.id]
    pending = [p for p in user_pmts if not p["confirmed"]]

    # Also check archived payments from previous drop
    archived = archived_payments.get(guild_id, {})
    archived_pmts = archived.get("payments", {})
    if isinstance(archived_pmts, dict) and user.id in archived_pmts:
        archived_pending = [p for p in archived_pmts[user.id] if not p["confirmed"]]
        pending = pending + archived_pending

    if not pending:
        await dm(ctx, f"⚠️  No pending payments found for **{user.display_name}**.")
        return

    for p in pending:
        p["confirmed"] = True

    total_confirmed = sum(p["amount"] for p in pending)
    spawn(update_all_live_boards(guild_id))

    # Update confirmed status in DB
    await db_update_user_claim_confirmed(guild_id, user.id)

    try:
        await user.send(f"✅  Your payment of **${total_confirmed:.2f}** has been confirmed! Thanks so much — enjoy your order! 🎉")
    except discord.Forbidden:
        pass
    await dm(ctx, f"✅  Confirmed **${total_confirmed:.2f}** from **{user.display_name}**.")


# ── PUBLIC COMMANDS ───────────────────────────────────────────────────────────

@bot.command(name="stock")
async def cmd_stock(ctx):
    if not ctx.guild:
        await ctx.author.send("⚠️  Please use this command in your server channel.")
        return
    guild_id = ctx.guild.id
    if session_state[guild_id] != "live":
        await ctx.send("No drop is currently active.")
        return
    await ctx.send(embed=build_stock_embed(guild_id))


@bot.command(name="paid")
@commands.cooldown(3, 60, commands.BucketType.user)
async def cmd_paid(ctx, *, args=""):
    if not ctx.guild:
        await ctx.author.send("⚠️  Please use `!paid` in your server channel.")
        return
    guild_id = ctx.guild.id

    # Check live drop, last snapshot, AND archived previous drop
    claims_ref = claims[guild_id] if claims[guild_id] else last_drop_snapshot.get(guild_id, {}).get("claims", {})
    archived = archived_payments.get(guild_id, {})
    archived_claims = archived.get("claims", {})

    has_live_claim = any(
        c["user"].id == ctx.author.id
        for claim_list in claims_ref.values()
        for c in claim_list
    )
    has_archived_claim = any(
        c["user"].id == ctx.author.id
        for claim_list in archived_claims.values()
        for c in claim_list
    )

    # Check if user has an unpaid raffle spot
    has_raffle_claim = any(
        s["user_id"] == ctx.author.id and not s["paid"]
        for raffle in server_raffles.get(guild_id, {}).values()
        if raffle["status"] in ("open", "closed")
        for s in raffle["slots"].values()
    )
    if not has_live_claim and not has_archived_claim and not has_raffle_claim:
        await ctx.author.send("⚠️  You don't have any claims to pay for.")
        await silent(ctx)
        return

    # ── Parse method + amount up front so every drop-routing path below has them ──
    def _owed_example():
        s_ref = stock[guild_id] if stock[guild_id] else last_drop_snapshot.get(guild_id, {}).get("stock", {})
        c_ref = claims[guild_id] if claims[guild_id] else last_drop_snapshot.get(guild_id, {}).get("claims", {})
        owed = 0.0
        for k, cl in c_ref.items():
            if k not in s_ref:
                continue
            for cc in cl:
                if cc["user"].id == ctx.author.id:
                    owed += cc["qty"] * s_ref[k]["price"]
        if owed <= 0:
            return "125"
        return f"{owed:.0f}" if owed == int(owed) else f"{owed:.2f}"

    drop_ch = get_drop_channel(ctx.guild) or ctx.channel
    parts = args.split()
    if len(parts) < 2:
        await silent(ctx)
        example = _owed_example()
        try:
            await drop_ch.send(
                f"{ctx.author.mention} — to report a payment, include the method **and** the amount:\n"
                f"```\n!paid venmo {example}\n```\n"
                f"Replace `venmo` with the method you actually used "
                f"(venmo, zelle, cashapp, applepay).",
                delete_after=30,
                allowed_mentions=discord.AllowedMentions(users=True)
            )
        except discord.HTTPException:
            pass
        return

    method = parts[0].lower()
    valid_methods = ["venmo", "zelle", "cashapp", "applepay", "apple"]
    if method not in valid_methods:
        await silent(ctx)
        example = _owed_example()
        try:
            await drop_ch.send(
                f"{ctx.author.mention} — that payment method isn't recognized. "
                f"Use one of: **venmo, zelle, cashapp, applepay**\n"
                f"```\n!paid venmo {example}\n```",
                delete_after=30,
                allowed_mentions=discord.AllowedMentions(users=True)
            )
        except discord.HTTPException:
            pass
        return
    if method == "apple":
        method = "applepay"

    try:
        amount = parse_price(parts[1])
    except ValueError:
        await silent(ctx)
        example = _owed_example()
        try:
            await drop_ch.send(
                f"{ctx.author.mention} — I couldn't read the amount `{parts[1]}`. "
                f"Please use a number:\n"
                f"```\n!paid {method} {example}\n```",
                delete_after=30,
                allowed_mentions=discord.AllowedMentions(users=True)
            )
        except discord.HTTPException:
            pass
        return

    await silent(ctx)

    # Default using_archive to False — set correctly in the drop routing below
    using_archive = False

    # If buyer has claims in BOTH drops — ask which one they're paying for
    if has_live_claim and has_archived_claim:
        live_total = sum(
            c["qty"] * stock[guild_id][key]["price"]
            for key, cl in claims_ref.items()
            for c in cl
            if c["user"].id == ctx.author.id and key in stock[guild_id]
        )
        arch_stock = archived.get("stock", {})
        arch_total = sum(
            c["qty"] * arch_stock[key]["price"]
            for key, cl in archived_claims.items()
            for c in cl
            if c["user"].id == ctx.author.id and key in arch_stock
        )
        # Auto-allocate the payment: previous drop first (it's been waiting longest),
        # then any remainder rolls to the current drop. No confusing 1/2 prompt.
        arch_confirmed = sum(
            p["amount"] for p in archived.get("payments", {}).get(ctx.author.id, [])
            if p["confirmed"]
        )
        arch_outstanding = max(arch_total - arch_confirmed, 0)

        live_confirmed = sum(
            p["amount"] for p in payments[guild_id].get(ctx.author.id, [])
            if p["confirmed"]
        )
        live_outstanding = max(live_total - live_confirmed, 0)

        allocated = []  # list of (label, amount, is_archive)
        remaining = amount

        # Pay down the previous drop first
        if arch_outstanding > 0.01 and remaining > 0.01:
            pay_arch = min(remaining, arch_outstanding)
            allocated.append(("previous drop", pay_arch, True))
            remaining -= pay_arch

        # Then the current drop
        if live_outstanding > 0.01 and remaining > 0.01:
            pay_live = min(remaining, live_outstanding)
            allocated.append(("current drop", pay_live, False))
            remaining -= pay_live

        # Anything left over (overpaid) applies to current drop as a credit
        if remaining > 0.01:
            allocated.append(("current drop", remaining, False))

        # Log each allocation to the correct payments bucket
        for label, amt, is_arch in allocated:
            if is_arch:
                if "payments" not in archived_payments.get(guild_id, {}):
                    archived_payments.setdefault(guild_id, {})["payments"] = defaultdict(list)
                archived_payments[guild_id]["payments"][ctx.author.id].append({
                    "method": method, "amount": amt,
                    "time": datetime.datetime.utcnow(), "confirmed": False
                })
            else:
                payments[guild_id][ctx.author.id].append({
                    "method": method, "amount": amt,
                    "time": datetime.datetime.utcnow(), "confirmed": False
                })

        # Build a clear breakdown message for the buyer
        breakdown = "\n".join(
            f"• ${amt:.2f} → {label}" for label, amt, _ in allocated
        )
        await silent(ctx)
        try:
            await ctx.author.send(
                f"✅  Got your **${amount:.2f}** payment via **{method.title()}**!\n\n"
                f"Here's how it was applied:\n{breakdown}\n\n"
                f"A manager will confirm shortly."
            )
        except discord.Forbidden:
            pass

        # Ping managers with the full context
        manager_mentions = " ".join(f"<@{uid}>" for uid in server_managers[guild_id])
        alloc_label = ", ".join(f"${amt:.2f} to {label}" for label, amt, _ in allocated)
        ping_msg = await drop_ch.send(
            f"💰  {manager_mentions} — **{ctx.author.display_name}** reported "
            f"**${amount:.2f}** via **{method.title()}** ({alloc_label}).\n"
            f"React ✅ to confirm or use `!confirm @{ctx.author.display_name}`.",
            allowed_mentions=discord.AllowedMentions(users=True)
        )
        await ping_msg.add_reaction("✅")
        pending_payment_messages[ping_msg.id] = {
            "guild_id":     guild_id,
            "buyer_id":     ctx.author.id,
            "raffle_spots": [],
        }
        return

    elif not has_live_claim and has_archived_claim:
        # Only archived claims — route to previous drop automatically
        claims_ref = archived_claims
        stock_ref = archived.get("stock", {})
        payments_ref = archived.get("payments", defaultdict(list))
        using_archive = True
    else:
        # Only live claims
        stock_ref = stock[guild_id] if stock[guild_id] else last_drop_snapshot.get(guild_id, {}).get("stock", {})
        payments_ref = payments[guild_id]
        using_archive = False

    # Log to correct payments bucket (live or archived)
    if using_archive:
        if guild_id not in archived_payments:
            archived_payments[guild_id] = {"payments": defaultdict(list), "claims": {}, "stock": {}}
        if "payments" not in archived_payments[guild_id]:
            archived_payments[guild_id]["payments"] = defaultdict(list)
        archived_payments[guild_id]["payments"][ctx.author.id].append({
            "method": method,
            "amount": amount,
            "time": datetime.datetime.utcnow(),
            "confirmed": False
        })
        payments_ref = archived_payments[guild_id]["payments"]
    else:
        payments[guild_id][ctx.author.id].append({
            "method": method,
            "amount": amount,
            "time": datetime.datetime.utcnow(),
            "confirmed": False
        })
        payments_ref = payments[guild_id]

    total_owed = sum(
        c["qty"] * stock_ref[key]["price"]
        for key, claim_list in claims_ref.items()
        for c in claim_list
        if c["user"].id == ctx.author.id and key in stock_ref
    )
    total_paid = sum(p["amount"] for p in payments_ref[ctx.author.id])
    remaining = total_owed - total_paid

    await ctx.author.send(
        f"💳  Payment of **${amount:.2f}** via **{method.title()}** received!\n"
        f"Total owed: ${total_owed:.2f}  •  Total reported: ${total_paid:.2f}"
        + (f"  •  Still outstanding: **${remaining:.2f}**" if remaining > 0.01 else "  •  ✅ Fully reported! Waiting on confirmation.")
    )

    drop_ch = get_drop_channel(ctx.guild) or ctx.channel
    manager_mentions = " ".join(f"<@{uid}>" for uid in server_managers[guild_id])

    # Build raffle context if user has unpaid raffle spots
    raffle_spots = []
    for r_name, raffle in server_raffles.get(guild_id, {}).items():
        for spot_num, s in raffle["slots"].items():
            if s["user_id"] == ctx.author.id and not s["paid"]:
                raffle_spots.append((r_name, spot_num))

    if raffle_spots:
        spots_label = ", ".join(f"**{r}** Spot #{n}" for r, n in raffle_spots)
        confirm_hint = (
            f"React \u2705 to confirm payment & raffle spot(s): {spots_label}\n"
            f"Or use `/raffle confirm {raffle_spots[0][0]}` to confirm raffle only."
        )
    else:
        confirm_hint = f"React \u2705 to confirm or use `!confirm @{ctx.author.display_name}`."

    ping_msg = await drop_ch.send(
        f"\U0001f4b0  {manager_mentions} \u2014 **{ctx.author.display_name}** reported payment of "
        f"**${amount:.2f}** via **{method.title()}**.\n{confirm_hint}",
        allowed_mentions=discord.AllowedMentions(users=True)
    )
    await ping_msg.add_reaction("\u2705")
    pending_payment_messages[ping_msg.id] = {
        "guild_id":     guild_id,
        "buyer_id":     ctx.author.id,
        "raffle_spots": raffle_spots,
    }


@bot.command(name="claim")
@commands.cooldown(5, 10, commands.BucketType.user)
async def cmd_claim(ctx, *, args=""):
    if not ctx.guild:
        await ctx.author.send("⚠️  Please use `!claim` in your server channel.")
        return
    guild_id = ctx.guild.id
    if session_state[guild_id] != "live":
        await ctx.send("⚠️  No active drop right now.")
        return
    parts = args.split()
    if not parts:
        await ctx.send("Usage: `!claim <item> <qty>`  e.g. `!claim PRE ETB 1`")
        return

    # ── EASTER EGGS ────────────────────────────────────────────────
    if args.strip().lower() == "all":
        oak_response = (
            "🔴  *Oak's words echoed: "
            "'There's a time and place for everything, but not now.' "
            "* Use `!claim <item> <qty>`."
        )
        responses = [
            oak_response,
            "💫  You used Splash. Nothing happened. Use `!claim <item> <qty>`.",
            "😴  Your claim used Rest. It fell asleep and did nothing. Try `!claim <item> <qty>`.",
            "💀  Giovanni himself reviewed your claim and rejected it. Try `!claim <item> <qty>`.",
        ]
        await ctx.send(random.choice(responses))
        return

    if "luck" in args.lower():
        await ctx.send(
            "🎰  Even Arceus couldn't find *luck* in this drop. "
            "It's not in stock. Check `!stock` for what's real."
        )
        return
    # ─────────────────────────────────────────────────────────────────────

    try:
        qty = int(parts[-1])
        item_name = " ".join(parts[:-1])
    except ValueError:
        qty = 1
        item_name = " ".join(parts)
    if not item_name:
        await ctx.send("Usage: `!claim <item> <qty>`  e.g. `!claim PRE ETB 1`")
        return
    if qty < 1:
        await ctx.send("⚠️  Qty must be at least 1.")
        return
    key = normalize(item_name)
    if key not in stock[guild_id]:
        matches = [k for k in stock[guild_id] if normalize(item_name) in k or k in normalize(item_name)]
        if len(matches) == 1:
            key = matches[0]
        elif len(matches) > 1:
            names = ", ".join(f"`{stock[guild_id][k]['display']}`" for k in matches)
            await ctx.send(f"⚠️  Multiple matches: {names} — be more specific.")
            return
        else:
            names = ", ".join(f"`{s['display']}`" for s in stock[guild_id].values())
            await ctx.send(f"⚠️  Item not found. Available: {names}")
            return
    info = stock[guild_id][key]
    already_claimed = sum(c["qty"] for c in claims[guild_id][key])
    remaining = info["qty"] - already_claimed
    if remaining <= 0:
        wl = waitlist[guild_id][key]
        already_on_wl = any(u.id == ctx.author.id for u in wl)
        if already_on_wl:
            await ctx.send(f"😔  **{info['display']}** is sold out and you're already on the waitlist.")
        else:
            await ctx.send(f"😔  **{info['display']}** is sold out! Use `!waitlist {info['display']}` to be notified if it opens up.")
        return
    if qty > remaining:
        await ctx.send(f"⚠️  Only **{remaining}** of **{info['display']}** left. Try `!claim {info['display']} {remaining}`")
        return
    if info["limit"] is not None:
        already_user = user_claimed_qty(guild_id, key, ctx.author.id)
        allowed = info["limit"] - already_user
        if allowed <= 0:
            await ctx.send(f"⚠️  You've already claimed the max of **{info['limit']}** for **{info['display']}**.")
            return
        if qty > allowed:
            await ctx.send(f"⚠️  You can only claim **{allowed}** more of **{info['display']}** (limit: {info['limit']} per person).")
            return
    existing = next((c for c in claims[guild_id][key] if c["user"].id == ctx.author.id), None)
    if existing:
        existing["qty"] += qty
    else:
        claims[guild_id][key].append({"user": ctx.author, "qty": qty, "time": datetime.datetime.utcnow()})
    new_remaining = remaining - qty
    total_cost = qty * info["price"]
    await ctx.send(f"✅  **{ctx.author.display_name}** claimed **{qty}x {info['display']}** — ${total_cost:.2f}  •  {new_remaining} left")
    spawn(update_all_live_boards(guild_id))
    if autoclose[guild_id] and all_sold_out(guild_id):
        drop_ch = get_drop_channel(ctx.guild) or ctx.channel
        await drop_ch.send("🎉  **Everything is claimed!** Closing the drop...")
        await close_drop(drop_ch, guild_id)


@bot.command(name="unclaim")
async def cmd_unclaim(ctx, *, args=""):
    if not ctx.guild:
        return
    guild_id = ctx.guild.id
    if session_state[guild_id] != "live":
        await ctx.send("⚠️  No active drop right now.")
        return
    if not args:
        await ctx.send("Usage: `!unclaim <item> <qty>`  e.g. `!unclaim PRE ETB 1`")
        return
    parts = args.split()
    try:
        qty = int(parts[-1])
        item_name = " ".join(parts[:-1])
    except ValueError:
        qty = None
        item_name = " ".join(parts)
    if not item_name:
        await ctx.send("Usage: `!unclaim <item> <qty>`")
        return
    key = normalize(item_name)
    if key not in stock[guild_id]:
        matches = [k for k in stock[guild_id] if normalize(item_name) in k or k in normalize(item_name)]
        if len(matches) == 1:
            key = matches[0]
        else:
            names = ", ".join(f"`{s['display']}`" for s in stock[guild_id].values())
            await ctx.send(f"⚠️  Item not found. Available: {names}")
            return
    existing = next((c for c in claims[guild_id][key] if c["user"].id == ctx.author.id), None)
    if not existing:
        await ctx.send("You don't have a claim on that item.")
        return
    if qty is None or qty >= existing["qty"]:
        freed = existing["qty"]
        claims[guild_id][key].remove(existing)
        await ctx.send(f"↩️  **{ctx.author.display_name}** removed their entire claim on **{stock[guild_id][key]['display']}**.")
    else:
        if qty < 1:
            await ctx.send("⚠️  Qty must be at least 1.")
            return
        freed = qty
        existing["qty"] -= qty
        await ctx.send(f"↩️  **{ctx.author.display_name}** removed **{qty}x {stock[guild_id][key]['display']}** from their claim. ({existing['qty']} still claimed)")
    spawn(update_all_live_boards(guild_id))
    await notify_waitlist(guild_id, key, freed)


@bot.command(name="waitlist")
async def cmd_waitlist(ctx, *, item_name=""):
    if not ctx.guild:
        await ctx.author.send("⚠️  Please use `!waitlist` in your server channel.")
        return
    guild_id = ctx.guild.id
    if session_state[guild_id] != "live":
        await ctx.send("⚠️  No active drop right now.")
        return
    if not item_name:
        await ctx.send("Usage: `!waitlist <item>`")
        return
    key = normalize(item_name)
    if key not in stock[guild_id]:
        matches = [k for k in stock[guild_id] if normalize(item_name) in k or k in normalize(item_name)]
        if len(matches) == 1:
            key = matches[0]
        else:
            names = ", ".join(f"`{s['display']}`" for s in stock[guild_id].values())
            await ctx.send(f"⚠️  Item not found. Available: {names}")
            return
    info = stock[guild_id][key]
    already_claimed = sum(c["qty"] for c in claims[guild_id][key])
    remaining = info["qty"] - already_claimed
    if remaining > 0:
        await ctx.send(f"**{info['display']}** is still available! Use `!claim {info['display']} 1` to grab it.")
        return
    already_on_wl = any(u.id == ctx.author.id for u in waitlist[guild_id][key])
    if already_on_wl:
        await ctx.send(f"You're already on the waitlist for **{info['display']}**!")
        return
    waitlist[guild_id][key].append(ctx.author)
    pos = len(waitlist[guild_id][key])
    await ctx.send(f"✅  **{ctx.author.display_name}** added to the waitlist for **{info['display']}** (position #{pos}).")


@bot.command(name="myclaims")
async def cmd_myclaims(ctx):
    if not ctx.guild:
        await ctx.author.send("⚠️  Please use `!myclaims` in your server channel.")
        return
    guild_id = ctx.guild.id

    # Works during live drop and after it closes
    stock_ref = stock[guild_id] if stock[guild_id] else last_drop_snapshot.get(guild_id, {}).get("stock", {})
    claims_ref = claims[guild_id] if claims[guild_id] else last_drop_snapshot.get(guild_id, {}).get("claims", {})

    if not stock_ref and not claims_ref:
        await ctx.send("No active drop.")
        return

    user_claims = []
    total = 0.0
    for key, claim_list in claims_ref.items():
        for c in claim_list:
            if c["user"].id == ctx.author.id and key in stock_ref:
                subtotal = c["qty"] * stock_ref[key]["price"]
                total += subtotal
                user_claims.append(f"• **{stock_ref[key]['display']}**  ×{c['qty']}  — ${subtotal:.2f}")
    if not user_claims:
        await ctx.send("You haven't claimed anything in this drop yet.")
        return
    lines = "\n".join(user_claims)
    await ctx.send(f"**{ctx.author.display_name}'s claims:**\n{lines}\n**Total owed: ${total:.2f}**")


@bot.command(name="history")
async def cmd_history(ctx):
    if not ctx.guild:
        await ctx.author.send("⚠️  Please run `!history` in your server channel.")
        return
    guild_id = ctx.guild.id
    if not is_manager(guild_id, ctx.author.id):
        return
    await silent(ctx)

    async with db_pool.acquire() as conn:
        rows = await conn.fetch("""
            SELECT closed_at, total_revenue, total_items, unique_buyers, summary
            FROM drop_history
            WHERE guild_id = $1
            ORDER BY closed_at DESC
            LIMIT 10
        """, guild_id)

    if not rows:
        await ctx.author.send("No drop history yet. History is saved automatically when a drop ends.")
        return

    embed = discord.Embed(
        title="📊  Drop History (Last 10)",
        color=discord.Color.blurple(),
        timestamp=datetime.datetime.utcnow()
    )

    all_revenue = 0.0
    all_items = 0
    for i, row in enumerate(rows):
        summary = json.loads(row["summary"])
        date_str = row["closed_at"].strftime("%b %d, %Y")
        item_lines = "\n".join(
            f"- {item}: x{data['qty']} - ${float(data['revenue']):.2f}"
            for item, data in summary.items()
        )
        footer_line = f"{row['total_items']} items  -  {row['unique_buyers']} buyer(s)"
        field_value = f"{item_lines}\n{footer_line}" if item_lines else footer_line
        if len(field_value) > 1024:
            field_value = field_value[:1020] + "..."
        embed.add_field(
            name=f"Drop #{len(rows) - i}  -  {date_str}  -  ${float(row['total_revenue']):.2f}",
            value=field_value,
            inline=False
        )
        all_revenue += float(row["total_revenue"])
        all_items += row["total_items"]

    embed.set_footer(text=f"All-time: ${all_revenue:.2f} revenue  -  {all_items} items sold")

    try:
        await ctx.author.send(embed=embed)
    except discord.Forbidden:
        await ctx.send("Could not DM you the history - please open your DMs and try again.")
    except discord.HTTPException:
        await ctx.send("History embed was too large to send.")


@bot.command(name="bump")
async def cmd_bump(ctx):
    if not ctx.guild:
        await ctx.author.send("⚠️  Please run `!bump` in your server channel.")
        return
    guild_id = ctx.guild.id
    if not is_manager(guild_id, ctx.author.id):
        return
    await silent(ctx)
    if not ctx.message.mentions:
        await dm(ctx, "Usage: `!bump @user`")
        return

    user = ctx.message.mentions[0]
    total_owed = get_user_total_owed(guild_id, user.id)

    if total_owed == 0:
        await dm(ctx, f"⚠️  **{user.display_name}** doesn't have any claims.")
        return

    confirmed = sum(p["amount"] for p in payments[guild_id][user.id] if p["confirmed"])
    remaining = total_owed - confirmed

    if remaining <= 0.01:
        await dm(ctx, f"✅  **{user.display_name}** is already fully paid!")
        return

    payment_info = build_payment_info(guild_id)
    try:
        await user.send(
            f"👋  Hey! Just a friendly reminder that you have an outstanding balance of **${remaining:.2f}** from the recent drop.\n\n"
            f"**Send payment using one of these methods:**\n{payment_info}\n\n"
            f"Once sent, run `!paid <method> <amount>` in the server to let us know. Thanks!"
        )
        await dm(ctx, f"✅  Bump sent to **{user.display_name}** — they owe **${remaining:.2f}**.")
    except discord.Forbidden:
        await dm(ctx, f"⚠️  Couldn't DM **{user.display_name}** — their DMs may be closed.")


@bot.command(name="remind")
async def cmd_remind(ctx):
    if not ctx.guild:
        await ctx.author.send("⚠️  Please run `!remind` in your server channel.")
        return
    guild_id = ctx.guild.id
    if not is_manager(guild_id, ctx.author.id):
        return
    await silent(ctx)
    drop_channel = get_drop_channel(ctx.guild) or ctx.channel

    claims_ref = claims[guild_id] if claims[guild_id] else last_drop_snapshot.get(guild_id, {}).get("claims", {})
    stock_ref = stock[guild_id] if stock[guild_id] else last_drop_snapshot.get(guild_id, {}).get("stock", {})

    unpaid_mentions = []
    seen_users = set()
    for key, claim_list in claims_ref.items():
        for c in claim_list:
            uid = c["user"].id
            if uid in seen_users or key not in stock_ref:
                continue
            seen_users.add(uid)
            total_owed = sum(
                cc["qty"] * stock_ref[k]["price"]
                for k, cl in claims_ref.items()
                for cc in cl
                if cc["user"].id == uid and k in stock_ref
            )
            confirmed = sum(p["amount"] for p in payments[guild_id][uid] if p["confirmed"])
            if total_owed - confirmed > 0.01:
                unpaid_mentions.append(f"<@{uid}>")

    if not unpaid_mentions:
        await dm(ctx, "✅  Everyone has been confirmed — no outstanding payments!")
        return

    payment_info = build_payment_info(guild_id)
    mentions_str = " ".join(unpaid_mentions)
    await drop_channel.send(
        f"⏰  **Payment Reminder** — {mentions_str}\n\n"
        f"You have an outstanding balance from the recent drop. Please send payment:\n"
        f"{payment_info}\n\n"
        f"Once sent, type `!paid <method> <amount>` to confirm. Thanks!"
    )
    await dm(ctx, f"✅  Reminder posted for {len(unpaid_mentions)} unpaid buyer(s).")




@bot.command(name="announce")
async def cmd_announce(ctx, *, message: str = ""):
    if not ctx.guild:
        await ctx.author.send("\u26a0\ufe0f  Please run `!announce` in your server channel.")
        return
    guild_id = ctx.guild.id
    if not is_manager(guild_id, ctx.author.id):
        return
    await silent(ctx)
    if not message:
        await dm(ctx, "Usage: `!announce <message>`\nExample: `!announce Drop going live in 10 minutes!`")
        return
    if len(message) > 4000:
        await dm(ctx, "⚠️  Message too long — keep it under 4000 characters.")
        return
    drop_channel = get_drop_channel(ctx.guild) or ctx.channel
    embed = discord.Embed(
        description=message,
        color=discord.Color.gold(),
        timestamp=datetime.datetime.utcnow()
    )
    embed.set_footer(text="VaultDrop")
    await drop_channel.send(embed=embed)
    if drop_channel != ctx.channel:
        await dm(ctx, f"\u2705  Announcement posted in **#{drop_channel.name}**.")


# ── CREATOR COMMANDS (DM only, cross-server) ──────────────────────────────────

@bot.command(name="creator")
async def cmd_creator(ctx, subcommand: str = "", *args):
    """Super admin commands — DM only, creator only."""
    # Must be DM and must be the creator
    if ctx.guild:
        await ctx.message.delete()
        return
    if not is_creator(ctx.author.id):
        return

    if not subcommand:
        await ctx.author.send(
            "**Creator Commands (DM only):**\n"
            "`!creator servers` — List all servers the bot is in\n"
            "`!creator info <guild_id>` — See a server's settings, admin, and managers\n"
            "`!creator confirm <guild_id> <user_id> [drop #]` — Confirm a user's payment, optionally for one drop (buyer is notified)\n"
            "`!creator paymentboard <guild_id>` — Post or refresh the payment board in a server's drop channel\n"
            "`!creator setpayment <guild_id>` — Update payment info for a server\n"
            "`!creator setdropchannel <guild_id> <channel_id>` — Update drop channel for a server\n"
            "`!creator resetadmin <guild_id> <user_id>` — Reassign the admin for a server\n"
            "`!creator announce <guild_id> <message>` — Post announcement in a server's drop channel"
        )
        return

    sub = subcommand.lower()

    # ── !creator servers ──────────────────────────────────────────────────────
    if sub == "servers":
        guilds = bot.guilds
        if not guilds:
            await ctx.author.send("The bot is not in any servers.")
            return
        lines = []
        for g in guilds:
            admin_id = server_admins.get(g.id)
            admin_str = f"<@{admin_id}>" if admin_id else "No admin set"
            settings = server_settings.get(g.id, {})
            ch_id = settings.get("drop_channel_id")
            ch_str = f"<#{ch_id}>" if ch_id else "No channel set"
            payment_set = any([settings.get("venmo"), settings.get("zelle"), settings.get("cashapp"), settings.get("applepay")])
            lines.append(f"**{g.name}** (`{g.id}`)")
            lines.append(f"  Admin: {admin_str}  |  Drop channel: {ch_str}  |  Payment info: {'✅' if payment_set else '❌'}")
        # Split into chunks to avoid 2000 char limit
        msg = "\n".join(lines)
        for i in range(0, len(msg), 1900):
            await ctx.author.send(msg[i:i+1900])
        return

    # ── !creator info <guild_id> ──────────────────────────────────────────────
    if sub == "info":
        if not args:
            await ctx.author.send("Usage: `!creator info <guild_id>`")
            return
        try:
            guild_id = int(args[0])
        except ValueError:
            await ctx.author.send("⚠️  Invalid guild ID.")
            return
        guild = bot.get_guild(guild_id)
        if not guild:
            await ctx.author.send(f"⚠️  Bot is not in a server with ID `{guild_id}`.")
            return
        admin_id = server_admins.get(guild_id)
        managers = [uid for uid in server_managers[guild_id] if uid != admin_id]
        settings = server_settings.get(guild_id, {})
        ch_id = settings.get("drop_channel_id")
        lines = [
            f"**Server:** {guild.name} (`{guild_id}`)",
            f"**Members:** {guild.member_count}",
            f"**Admin:** {'<@' + str(admin_id) + '>' if admin_id else 'Not set'}",
            f"**Managers:** {', '.join('<@' + str(uid) + '>' for uid in managers) if managers else 'None'}",
            f"**Drop channel:** {'<#' + str(ch_id) + '>' if ch_id else 'Not set'}",
            f"**Venmo:** {settings.get('venmo') or 'Not set'}",
            f"**Zelle:** {settings.get('zelle') or 'Not set'}",
            f"**Cash App:** {settings.get('cashapp') or 'Not set'}",
            f"**Apple Pay:** {settings.get('applepay') or 'Not set'}",
        ]
        await ctx.author.send("\n".join(lines))
        return

    # ── !creator setpayment <guild_id> ────────────────────────────────────────
    if sub == "setpayment":
        if not args:
            await ctx.author.send("Usage: `!creator setpayment <guild_id>`")
            return
        try:
            guild_id = int(args[0])
        except ValueError:
            await ctx.author.send("⚠️  Invalid guild ID.")
            return
        guild = bot.get_guild(guild_id)
        if not guild:
            await ctx.author.send(f"⚠️  Bot is not in a server with ID `{guild_id}`.")
            return
        await ctx.author.send(f"Setting payment info for **{guild.name}**:")
        await collect_payment_info(ctx.author, guild_id)
        return

    # ── !creator setdropchannel <guild_id> <channel_id> ──────────────────────
    if sub == "setdropchannel":
        if len(args) < 2:
            await ctx.author.send("Usage: `!creator setdropchannel <guild_id> <channel_id>`")
            return
        try:
            guild_id = int(args[0])
            channel_id = int(args[1])
        except ValueError:
            await ctx.author.send("⚠️  Invalid guild ID or channel ID.")
            return
        guild = bot.get_guild(guild_id)
        if not guild:
            await ctx.author.send(f"⚠️  Bot is not in a server with ID `{guild_id}`.")
            return
        channel = guild.get_channel(channel_id)
        if not channel:
            await ctx.author.send(f"⚠️  Channel `{channel_id}` not found in **{guild.name}**. Use `!creator info {guild_id}` to check available channels.")
            return
        if guild_id not in server_settings:
            server_settings[guild_id] = {}
        server_settings[guild_id]["drop_channel_id"] = channel_id
        await db_save_settings(guild_id)
        await ctx.author.send(f"✅  Drop channel for **{guild.name}** updated to **#{channel.name}**.")
        return

    # ── !creator resetadmin <guild_id> <user_id> ─────────────────────────────
    if sub == "resetadmin":
        if len(args) < 2:
            await ctx.author.send("Usage: `!creator resetadmin <guild_id> <user_id>`")
            return
        try:
            guild_id = int(args[0])
            new_admin_id = int(args[1])
        except ValueError:
            await ctx.author.send("⚠️  Invalid guild ID or user ID.")
            return
        guild = bot.get_guild(guild_id)
        if not guild:
            await ctx.author.send(f"⚠️  Bot is not in a server with ID `{guild_id}`.")
            return
        member = guild.get_member(new_admin_id)
        if not member:
            await ctx.author.send(f"⚠️  User `{new_admin_id}` not found in **{guild.name}**.")
            return
        old_admin_id = server_admins.get(guild_id)
        server_admins[guild_id] = new_admin_id
        server_managers[guild_id].add(new_admin_id)
        await db_set_admin(guild_id, new_admin_id)
        await db_add_manager(guild_id, new_admin_id)
        await ctx.author.send(
            f"✅  Admin for **{guild.name}** updated.\n"
            f"Old admin: {'<@' + str(old_admin_id) + '>' if old_admin_id else 'None'}\n"
            f"New admin: **{member.display_name}** (`{new_admin_id}`)"
        )
        return

    # ── !creator announce <guild_id> <message> ────────────────────────────────
    if sub == "announce":
        if len(args) < 2:
            await ctx.author.send(
                "Usage: `!creator announce <guild_id> <message>`\n"
                "Example: `!creator announce 123456789012345678 Drop going live soon!`"
            )
            return
        try:
            guild_id = int(args[0])
        except ValueError:
            await ctx.author.send("⚠️  Invalid guild ID — must be a number.")
            return
        guild = bot.get_guild(guild_id)
        if not guild:
            await ctx.author.send(f"⚠️  Bot is not in a server with ID `{guild_id}`.")
            return
        message = " ".join(args[1:])
        if not message.strip():
            await ctx.author.send("⚠️  Message cannot be empty.")
            return
        drop_channel = get_drop_channel(guild)
        if not drop_channel:
            drop_channel = next(
                (c for c in guild.text_channels if c.permissions_for(guild.me).send_messages),
                None
            )
            if not drop_channel:
                await ctx.author.send(f"⚠️  No drop channel set for **{guild.name}** and no accessible channel found.")
                return
            await ctx.author.send(f"⚠️  No drop channel configured — posting to **#{drop_channel.name}** instead.")
        embed = discord.Embed(
            description=message,
            color=discord.Color.gold(),
            timestamp=datetime.datetime.utcnow()
        )
        embed.set_footer(text="VaultDrop")
        await drop_channel.send(embed=embed)
        await ctx.author.send(f"✅  Announcement posted in **#{drop_channel.name}** on **{guild.name}**.")
        return

    # ── !creator confirm <guild_id> <user_id> ────────────────────────────────
    if sub == "confirm":
        if len(args) < 2:
            await ctx.author.send(
                "Usage: `!creator confirm <guild_id> <user_id> [drop #]`  "
                "(you can @mention the user instead of the ID)"
            )
            return
        try:
            guild_id = int(args[0])
        except ValueError:
            await ctx.author.send("⚠️  Invalid guild ID.")
            return
        # Accept either a raw user ID or an @mention for the buyer
        if ctx.message.mentions:
            buyer_id = ctx.message.mentions[0].id
            # Optional trailing drop number (a bare digit after the guild id)
            drop_number = next((int(a) for a in args[1:] if a.isdigit()), None)
        else:
            try:
                buyer_id = int(args[1])
            except ValueError:
                await ctx.author.send("⚠️  Invalid user ID.")
                return
            drop_number = int(args[2]) if len(args) > 2 and args[2].isdigit() else None
        guild = bot.get_guild(guild_id)
        if not guild:
            await ctx.author.send(f"⚠️  Bot is not in a server with ID `{guild_id}`.")
            return

        # If a drop number is given, confirm only that drop's payment.
        if drop_number is not None:
            feedback = await confirm_drop_payment(guild_id, buyer_id, drop_number, guild)
            await ctx.author.send(feedback)
            return

        # Same logic as !confirm: confirm pending reported payments from the
        # current drop and the previous (archived) drop.
        pending = [p for p in payments[guild_id].get(buyer_id, []) if not p["confirmed"]]
        archived = archived_payments.get(guild_id, {})
        archived_pmts = archived.get("payments", {})
        if isinstance(archived_pmts, dict) and buyer_id in archived_pmts:
            pending += [p for p in archived_pmts[buyer_id] if not p["confirmed"]]

        if not pending:
            await ctx.author.send(
                f"⚠️  No pending payments found for `{buyer_id}` in **{guild.name}**."
            )
            return

        for p in pending:
            p["confirmed"] = True
        total_confirmed = sum(p["amount"] for p in pending)

        spawn(update_all_live_boards(guild_id))
        await db_update_user_claim_confirmed(guild_id, buyer_id)

        # Notify the buyer from the bot, exactly like a manager confirm.
        member = guild.get_member(buyer_id)
        buyer_name = member.display_name if member else str(buyer_id)
        if member:
            try:
                await member.send(
                    f"✅  Your payment of **${total_confirmed:.2f}** has been confirmed! "
                    f"Thanks so much — enjoy your order! 🎉"
                )
            except discord.Forbidden:
                pass
        await ctx.author.send(
            f"✅  Confirmed **${total_confirmed:.2f}** from **{buyer_name}** in **{guild.name}**."
        )
        return

    # ── !creator paymentboard <guild_id> ─────────────────────────────────────
    if sub == "paymentboard":
        if not args:
            await ctx.author.send("Usage: `!creator paymentboard <guild_id>`")
            return
        try:
            guild_id = int(args[0])
        except ValueError:
            await ctx.author.send("⚠️  Invalid guild ID.")
            return
        guild = bot.get_guild(guild_id)
        if not guild:
            await ctx.author.send(f"⚠️  Bot is not in a server with ID `{guild_id}`.")
            return
        drop_channel = get_drop_channel(guild)
        if not drop_channel:
            drop_channel = next(
                (c for c in guild.text_channels if c.permissions_for(guild.me).send_messages),
                None,
            )
            if not drop_channel:
                await ctx.author.send(
                    f"⚠️  No drop channel set for **{guild.name}** and no accessible channel found."
                )
                return
        result = await post_or_refresh_payment_board(guild, drop_channel)
        await ctx.author.send(f"✅  Payment board {result} on **{guild.name}**.")
        return

    await ctx.author.send(f"⚠️  Unknown subcommand `{subcommand}`. Type `!creator` for a list of commands.")


@bot.command(name="help")
async def cmd_help(ctx):
    if not ctx.guild:
        return
    embed = discord.Embed(
        title="📖  VaultDrop Commands",
        color=discord.Color.gold(),
        timestamp=datetime.datetime.utcnow()
    )
    embed.add_field(
        name="During a Drop",
        value=(
            "`!claim <item> <qty>` — Claim an item\n"
            "`!unclaim <item> <qty>` — Remove a claim\n"
            "`!stock` — See what's available\n"
            "`!myclaims` — See your claims and total\n"
            "`!waitlist <item>` — Join waitlist for sold out items"
        ),
        inline=False
    )
    embed.add_field(
        name="Payments",
        value=(
            "`!paid <method> <amount>` — Report your payment\n"
            "e.g. `!paid venmo $125`\n"
            "Methods: venmo, zelle, cashapp, applepay"
        ),
        inline=False
    )
    embed.add_field(
        name="Your Order History",
        value=(
            "`!myhistory` — View all your past claims and payment status\n"
            "Works anytime — even after a drop closes!"
        ),
        inline=False
    )
    embed.add_field(
        name="Raffles",
        value=(
            "`/raffles` — See all active raffles\n"
            "Tap a button on the raffle post to claim a spot — no command needed!"
        ),
        inline=False
    )
    embed.set_footer(text="VaultDrop — First come, first served!")
    await ctx.send(embed=embed)


@bot.command(name="myhistory")
async def cmd_myhistory(ctx):
    if not ctx.guild:
        await ctx.author.send("⚠️  Please use `!myhistory` in your server channel.")
        return
    guild_id = ctx.guild.id
    await silent(ctx)

    async with db_pool.acquire() as conn:
        rows = await conn.fetch("""
            SELECT drop_number, closed_at, item_display, qty, price, subtotal, confirmed, tracking
            FROM user_claims
            WHERE guild_id = $1 AND user_id = $2
            ORDER BY closed_at DESC, drop_number DESC
        """, guild_id, ctx.author.id)

    if not rows:
        await ctx.author.send("You don't have any claim history yet. Claims are saved when a drop closes.")
        return

    # Group by drop_number
    drops = {}
    for row in rows:
        dn = row["drop_number"]
        if dn not in drops:
            drops[dn] = {
                "drop_number": dn,
                "closed_at": row["closed_at"],
                "items": [],
                "total": 0.0,
                "confirmed": row["confirmed"],
                "tracking": row["tracking"],
            }
        drops[dn]["items"].append({
            "display": row["item_display"],
            "qty": row["qty"],
            "subtotal": float(row["subtotal"]),
        })
        drops[dn]["total"] += float(row["subtotal"])
        # confirmed = True if any item in drop is confirmed
        if row["confirmed"]:
            drops[dn]["confirmed"] = True
        # tracking is per drop; keep the first non-empty value seen
        if row["tracking"] and not drops[dn].get("tracking"):
            drops[dn]["tracking"] = row["tracking"]

    embed = discord.Embed(
        title=f"📋  {ctx.author.display_name}'s Claim History",
        color=discord.Color.blurple(),
        timestamp=datetime.datetime.utcnow()
    )

    all_time_total = 0.0
    for dn, drop in sorted(drops.items(), reverse=True):
        date_str = drop["closed_at"].strftime("%b %d, %Y")
        status = "✅  Paid & Confirmed" if drop["confirmed"] else "⏳  Payment Pending"
        item_lines = []
        for item in drop["items"]:
            item_lines.append(f"- {item['display']}  x{item['qty']}  - ${item['subtotal']:.2f}")
        lines_str = "\n".join(item_lines)
        field_value = f"{lines_str}\n**Total: ${drop['total']:.2f}**\n{status}"
        if drop.get("tracking"):
            field_value += f"\n📦  Shipped — Tracking: `{drop['tracking']}`"
        if len(field_value) > 1024:
            field_value = field_value[:1020] + "..."
        embed.add_field(
            name=f"Drop #{dn}  •  {date_str}",
            value=field_value,
            inline=False
        )
        all_time_total += drop["total"]

    embed.set_footer(text=f"All-time total: ${all_time_total:.2f}")

    try:
        await ctx.author.send(embed=embed)
    except discord.Forbidden:
        await ctx.send("⚠️  I couldn't DM you — please open your DMs and try again.")



# ══════════════════════════════════════════════════════════════════════════════
# RAFFLE MODULE — Button UI + Slash Commands
# ══════════════════════════════════════════════════════════════════════════════

# ── RAFFLE DB HELPERS ─────────────────────────────────────────────────────────

async def _db_save_raffle(guild_id: int, name: str):
    r = server_raffles[guild_id][name]
    async with db_pool.acquire() as conn:
        await conn.execute("""
            INSERT INTO raffles (guild_id, name, spots, price, channel_id, message_id, status, host_num)
            VALUES ($1, $2, $3, $4, $5, $6, $7, $8)
            ON CONFLICT (guild_id, name) DO UPDATE SET
                spots      = EXCLUDED.spots,
                price      = EXCLUDED.price,
                channel_id = EXCLUDED.channel_id,
                message_id = EXCLUDED.message_id,
                status     = EXCLUDED.status,
                host_num   = EXCLUDED.host_num
        """, guild_id, name,
            r["spots"], r["price"], r["channel_id"], r["message_id"], r["status"],
            r.get("host_num", 0))


async def _db_save_slot(guild_id: int, name: str, spot_num: int):
    s = server_raffles[guild_id][name]["slots"][spot_num]
    async with db_pool.acquire() as conn:
        await conn.execute("""
            INSERT INTO raffle_slots (guild_id, raffle_name, spot_num, user_id, username, paid)
            VALUES ($1, $2, $3, $4, $5, $6)
            ON CONFLICT (guild_id, raffle_name, spot_num) DO UPDATE SET
                user_id  = EXCLUDED.user_id,
                username = EXCLUDED.username,
                paid     = EXCLUDED.paid
        """, guild_id, name, spot_num, s["user_id"], s["username"], s["paid"])


async def _db_delete_raffle(guild_id: int, name: str):
    async with db_pool.acquire() as conn:
        await conn.execute(
            "DELETE FROM raffle_slots WHERE guild_id = $1 AND raffle_name = $2",
            guild_id, name)
        await conn.execute(
            "DELETE FROM raffles WHERE guild_id = $1 AND name = $2",
            guild_id, name)


async def _db_save_raffle_channel(guild_id: int, channel_id: int):
    if guild_id not in server_settings:
        server_settings[guild_id] = {}
    server_settings[guild_id]["raffle_channel_id"] = channel_id
    await db_save_settings(guild_id)


# ── RAFFLE EMBED BUILDER ──────────────────────────────────────────────────────

def _raffle_embed(name: str, raffle: dict) -> discord.Embed:
    slots     = raffle["slots"]
    total     = raffle["spots"]
    claimed   = sum(1 for s in slots.values() if s["user_id"] is not None)
    remaining = total - claimed

    if raffle["status"] == "complete":
        color = discord.Color.gold()
    elif raffle["status"] == "closed":
        color = discord.Color.red()
    else:
        color = discord.Color.green()

    embed = discord.Embed(title=f"Raffle -- {name}", color=color)
    embed.add_field(name="Price per Spot", value=raffle["price"], inline=True)
    embed.add_field(name="Total Spots",    value=str(total),      inline=True)
    embed.add_field(name="Remaining",      value=str(remaining),  inline=True)

    lines = []
    for num in sorted(slots.keys()):
        s = slots[num]
        if s["user_id"] is None:
            lines.append(f"`{num:>2}` Open")
        elif s["paid"]:
            lines.append(f"`{num:>2}` Paid: {s['username']}")
        else:
            lines.append(f"`{num:>2}` Pending: {s['username']}")

    mid = (len(lines) + 1) // 2
    embed.add_field(name="Spots",  value="\n".join(lines[:mid]) or "--", inline=True)
    embed.add_field(name="\u200b", value="\n".join(lines[mid:]) or "--", inline=True)

    status_map = {
        "open":     "Open -- tap a button to claim a spot!",
        "closed":   "All spots claimed -- awaiting payment confirmations",
        "complete": "Raffle complete",
    }
    embed.set_footer(text=status_map.get(raffle["status"], ""))
    return embed


def _build_raffle_payment_dm(guild_id: int, host_num: int = 0) -> str:
    """Build payment info string. Uses host-specific info if host_num is 1 or 2."""
    if host_num in (1, 2):
        h = raffle_hosts.get(guild_id, {}).get(host_num, {})
        if h:
            lines = []
            if h.get("venmo"):    lines.append(f"Venmo: {h['venmo']}")
            if h.get("zelle"):    lines.append(f"Zelle: {h['zelle']}")
            if h.get("cashapp"):  lines.append(f"Cash App: {h['cashapp']}")
            if h.get("applepay"): lines.append(f"Apple Pay: {h['applepay']}")
            host_label = h.get("name") or f"Host {host_num}"
            prefix = f"Send payment to **{host_label}**:"
            return prefix + "\n" + ("\n".join(lines) if lines else "(no payment methods set for this host)")
    # Fall back to server default
    s = server_settings.get(guild_id, {})
    lines = []
    if s.get("venmo"):    lines.append(f"Venmo: {s['venmo']}")
    if s.get("zelle"):    lines.append(f"Zelle: {s['zelle']}")
    if s.get("cashapp"):  lines.append(f"Cash App: {s['cashapp']}")
    if s.get("applepay"): lines.append(f"Apple Pay: {s['applepay']}")
    return "\n".join(lines) if lines else "(payment info not set -- contact the server owner)"


def _build_raffle_view(guild_id: int, name: str, raffle: dict) -> discord.ui.View:
    view = discord.ui.View(timeout=None)
    for num in sorted(raffle["slots"].keys()):
        s     = raffle["slots"][num]
        taken = s["user_id"] is not None
        label = f"Spot {num}" if not taken else f"#{num} -- {s['username'][:12]}"
        btn   = discord.ui.Button(
            label     = label,
            style     = discord.ButtonStyle.danger if taken else discord.ButtonStyle.success,
            disabled  = taken or raffle["status"] != "open",
            custom_id = f"raffle:{guild_id}:{name}:{num}",
            row       = (num - 1) // 5,
        )
        view.add_item(btn)
    return view


# ── RAFFLE AUTOCOMPLETE ───────────────────────────────────────────────────────

async def _raffle_name_autocomplete(
    interaction: discord.Interaction,
    current: str,
) -> list[discord.app_commands.Choice[str]]:
    guild_id = interaction.guild_id
    raffles  = server_raffles.get(guild_id, {})
    return [
        discord.app_commands.Choice(name=n, value=n)
        for n in raffles
        if current.lower() in n.lower()
    ][:25]


# ── BUTTON INTERACTION HANDLER ────────────────────────────────────────────────

@bot.event
async def on_interaction(interaction: discord.Interaction):
    if interaction.type != discord.InteractionType.component:
        return
    custom_id = interaction.data.get("custom_id", "")
    if not custom_id.startswith("raffle:"):
        return
    parts = custom_id.split(":", 3)
    if len(parts) != 4:
        return
    try:
        guild_id = int(parts[1])
        name     = parts[2]
        spot_num = int(parts[3])
    except (ValueError, IndexError):
        return

    await interaction.response.defer(ephemeral=True)

    if guild_id not in server_raffles or name not in server_raffles[guild_id]:
        await interaction.followup.send("This raffle no longer exists.", ephemeral=True)
        return

    raffle = server_raffles[guild_id][name]

    if raffle["status"] != "open":
        await interaction.followup.send("This raffle is no longer accepting claims.", ephemeral=True)
        return

    slot = raffle["slots"].get(spot_num)
    if slot is None or slot["user_id"] is not None:
        await interaction.followup.send(
            f"Spot #{spot_num} was just taken! Pick another open spot.",
            ephemeral=True,
        )
        return

    username = str(interaction.user)
    raffle["slots"][spot_num] = {"user_id": interaction.user.id, "username": username, "paid": False}
    await _db_save_slot(guild_id, name, spot_num)

    all_claimed = all(s["user_id"] is not None for s in raffle["slots"].values())
    if all_claimed:
        raffle["status"] = "closed"
        await _db_save_raffle(guild_id, name)

    channel = bot.get_channel(raffle["channel_id"])
    if channel and raffle["message_id"]:
        try:
            msg  = await channel.fetch_message(raffle["message_id"])
            view = _build_raffle_view(guild_id, name, raffle)
            await msg.edit(embed=_raffle_embed(name, raffle), view=view)
        except (discord.NotFound, discord.HTTPException):
            pass

    if all_claimed and channel:
        await channel.send(
            f"All spots in **{name}** are claimed! "
            f"Waiting on payment confirmations before the spin."
        )

    await interaction.followup.send(
        f"You claimed Spot #{spot_num} in **{name}**! Check your DMs for payment details.",
        ephemeral=True,
    )

    payment_info = _build_raffle_payment_dm(guild_id, raffle.get("host_num", 0))
    try:
        await interaction.user.send(
            f"You claimed Spot #{spot_num} in the **{name}** raffle!\n\n"
            f"Price: {raffle['price']}\n\n"
            f"Send payment using one of these methods:\n{payment_info}\n\n"
            f"Once sent, go back to the server and type:\n"
            f"`!paid venmo $25` (or whichever method you used)\n\nGood luck!"
        )
    except discord.Forbidden:
        if channel:
            await channel.send(
                f"{interaction.user.mention} -- I couldn't DM you! "
                f"Please open your DMs and contact the server owner for payment details.",
                delete_after=25,
            )


# ── SLASH COMMAND GROUP ───────────────────────────────────────────────────────

raffle_group = discord.app_commands.Group(
    name="raffle",
    description="Raffle commands for VaultDrop"
)


@raffle_group.command(name="setchannel", description="Set the channel where raffles are posted (one-time setup)")
@discord.app_commands.describe(channel="The channel to post raffles in")
async def slash_raffle_setchannel(interaction: discord.Interaction, channel: discord.TextChannel):
    await interaction.response.defer(ephemeral=True)
    if interaction.guild.owner_id != interaction.user.id:
        await interaction.followup.send("Only the server owner can set the raffle channel.", ephemeral=True)
        return
    guild_id = interaction.guild_id
    server_raffle_channel[guild_id] = channel.id
    await _db_save_raffle_channel(guild_id, channel.id)
    await interaction.followup.send(f"Raffle channel set to **#{channel.name}**.", ephemeral=True)


@raffle_group.command(name="create", description="Create a new raffle with button-based spot claiming")
@discord.app_commands.describe(
    name="Raffle name (e.g. ScarletVault)",
    spots="Number of spots (2-10)",
    price="Price per spot (e.g. $25)",
    host="Which host is collecting payment (1 or 2) — leave blank to use server default",
)
async def slash_raffle_create(interaction: discord.Interaction, name: str, spots: int, price: str, host: int = 0):
    await interaction.response.defer(ephemeral=True)
    if interaction.guild.owner_id != interaction.user.id:
        await interaction.followup.send("Only the server owner can create raffles.", ephemeral=True)
        return
    guild_id = interaction.guild_id
    if spots < 2 or spots > 10:
        await interaction.followup.send("Spots must be between 2 and 10.", ephemeral=True)
        return
    price = price if price.startswith("$") else f"${price}"
    if name in server_raffles[guild_id]:
        await interaction.followup.send(
            f"A raffle named **{name}** already exists. Cancel it first or use a different name.",
            ephemeral=True,
        )
        return
    if guild_id not in server_raffle_channel:
        await interaction.followup.send(
            "One-time setup required.\nRun `/raffle setchannel #channel` first, then create your raffle.",
            ephemeral=True,
        )
        return

    if host not in (0, 1, 2):
        await interaction.followup.send("Host must be 1 or 2 (or leave blank for server default).", ephemeral=True)
        return

    if host in (1, 2) and host not in raffle_hosts.get(guild_id, {}):
        await interaction.followup.send(
            f"⚠️  Host {host} has no payment info set. Run `/raffle sethost host:{host}` first.",
            ephemeral=True
        )
        return

    raffle = {
        "spots":      spots,
        "price":      price,
        "channel_id": server_raffle_channel[guild_id],
        "message_id": None,
        "status":     "open",
        "host_num":   host,
        "slots":      {n: {"user_id": None, "username": None, "paid": False}
                       for n in range(1, spots + 1)},
    }
    server_raffles[guild_id][name] = raffle

    channel = bot.get_channel(server_raffle_channel[guild_id])
    if channel is None:
        await interaction.followup.send(
            "Raffle channel not found. Use `/raffle setchannel` to set a new one.",
            ephemeral=True,
        )
        del server_raffles[guild_id][name]
        return

    payment_hint = _build_raffle_payment_dm(guild_id, raffle.get("host_num", 0))
    embed        = _raffle_embed(name, raffle)
    embed.description = (
        f"Tap a button below to claim your spot!\n"
        f"The bot will DM you payment details instantly.\n\n"
        f"Payment accepted via:\n{payment_hint}"
    )
    view = _build_raffle_view(guild_id, name, raffle)
    msg  = await channel.send(embed=embed, view=view)

    raffle["message_id"] = msg.id
    await _db_save_raffle(guild_id, name)
    for spot_num in raffle["slots"]:
        await _db_save_slot(guild_id, name, spot_num)

    host_label = ""
    if host in (1, 2):
        h = raffle_hosts.get(guild_id, {}).get(host, {})
        host_label = f" — payments to **{h.get('name', f'Host {host}')}**"
    await interaction.followup.send(
        f"Raffle **{name}** is live -- **{spots} spots** at **{price}** each{host_label}!",
        ephemeral=True,
    )


@raffle_group.command(name="confirm", description="Confirm a user's payment for a raffle spot")
@discord.app_commands.autocomplete(name=_raffle_name_autocomplete)
@discord.app_commands.describe(name="Raffle name", user="The user to confirm")
async def slash_raffle_confirm(interaction: discord.Interaction, name: str, user: discord.Member):
    await interaction.response.defer(ephemeral=True)
    guild_id = interaction.guild_id
    # Allow server owner OR any drop manager to confirm raffle payments
    if (interaction.guild.owner_id != interaction.user.id
            and not is_manager(guild_id, interaction.user.id)):
        await interaction.followup.send(
            "Only the server owner or a drop manager can confirm raffle payments.",
            ephemeral=True
        )
        return
    if name not in server_raffles[guild_id]:
        await interaction.followup.send(f"No raffle named **{name}**.", ephemeral=True)
        return

    raffle       = server_raffles[guild_id][name]
    unpaid_spots = [
        num for num, s in raffle["slots"].items()
        if s["user_id"] == user.id and not s["paid"]
    ]
    if not unpaid_spots:
        await interaction.followup.send(
            f"**{user.display_name}** has no unpaid spots in **{name}**.",
            ephemeral=True,
        )
        return

    for num in unpaid_spots:
        raffle["slots"][num]["paid"] = True
        await _db_save_slot(guild_id, name, num)

    spots_str = ", ".join(f"#{n}" for n in sorted(unpaid_spots))

    channel = bot.get_channel(raffle["channel_id"])
    if channel and raffle["message_id"]:
        try:
            msg  = await channel.fetch_message(raffle["message_id"])
            view = _build_raffle_view(guild_id, name, raffle)
            await msg.edit(embed=_raffle_embed(name, raffle), view=view)
        except (discord.NotFound, discord.HTTPException):
            pass
        await channel.send(
            f"Payment confirmed -- **{user.display_name}** "
            f"(Spot{'s' if len(unpaid_spots) > 1 else ''} {spots_str}) in **{name}**!"
        )

    try:
        await user.send(
            f"Your payment for Spot{'s' if len(unpaid_spots) > 1 else ''} {spots_str} "
            f"in the **{name}** raffle is confirmed! Watch for the live spin announcement."
        )
    except discord.Forbidden:
        pass

    await interaction.followup.send(
        f"Confirmed payment for **{user.display_name}** -- Spot{'s' if len(unpaid_spots) > 1 else ''} {spots_str}.",
        ephemeral=True,
    )


@raffle_group.command(name="wheel", description="Generate Wheel of Names entry list for the live spin")
@discord.app_commands.autocomplete(name=_raffle_name_autocomplete)
@discord.app_commands.describe(
    name="Raffle name",
    force="Spin even if some payments are not confirmed yet",
)
async def slash_raffle_wheel(interaction: discord.Interaction, name: str, force: bool = False):
    await interaction.response.defer(ephemeral=True)
    if interaction.guild.owner_id != interaction.user.id:
        await interaction.followup.send("Only the server owner can start the wheel.", ephemeral=True)
        return
    guild_id = interaction.guild_id
    if name not in server_raffles[guild_id]:
        await interaction.followup.send(f"No raffle named **{name}**.", ephemeral=True)
        return

    raffle     = server_raffles[guild_id][name]
    paid_slots = [(num, s) for num, s in sorted(raffle["slots"].items()) if s["paid"]]

    if not paid_slots:
        await interaction.followup.send(
            "No confirmed payments yet. Use `/raffle confirm` first.",
            ephemeral=True,
        )
        return

    unpaid_claimed = [
        (num, s) for num, s in raffle["slots"].items()
        if s["user_id"] is not None and not s["paid"]
    ]
    if unpaid_claimed and not force:
        names_list = ", ".join(f"**{s['username']}** (#{num})" for num, s in unpaid_claimed)
        await interaction.followup.send(
            f"Unconfirmed payments: {names_list}\n\n"
            f"Confirm them first, or use `/raffle wheel name:{name} force:True` to spin without them.",
            ephemeral=True,
        )
        return

    entries      = [f"{s['username']} - Spot {num}" for num, s in paid_slots]
    entries_text = "\n".join(entries)

    embed = discord.Embed(
        title=f"SPIN TIME -- {name} Raffle!",
        description=(
            f"**{len(paid_slots)} entries** are ready!\n\n"
            f"Go to **wheelofnames.com**, paste the names below, share your screen and spin!\n\n"
            f"After the spin use `/raffle winner {name} <spot>`"
        ),
        color=discord.Color.gold(),
    )
    embed.add_field(
        name="Copy these names into Wheel of Names",
        value=f"```\n{entries_text}\n```",
        inline=False,
    )
    await interaction.followup.send(embed=embed, ephemeral=True)


@raffle_group.command(name="winner", description="Announce the raffle winner by winning spot number")
@discord.app_commands.autocomplete(name=_raffle_name_autocomplete)
@discord.app_commands.describe(
    name="Raffle name",
    spot="The winning spot number from the wheel spin",
)
async def slash_raffle_winner(interaction: discord.Interaction, name: str, spot: int):
    await interaction.response.defer(ephemeral=True)
    if interaction.guild.owner_id != interaction.user.id:
        await interaction.followup.send("Only the server owner can record the winner.", ephemeral=True)
        return
    guild_id = interaction.guild_id
    if name not in server_raffles[guild_id]:
        await interaction.followup.send(f"No raffle named **{name}**.", ephemeral=True)
        return

    raffle = server_raffles[guild_id][name]
    if spot not in raffle["slots"]:
        await interaction.followup.send(
            f"Spot #{spot} doesn't exist in **{name}**. Valid spots: 1-{raffle['spots']}.",
            ephemeral=True,
        )
        return

    winning_slot = raffle["slots"][spot]
    if winning_slot["user_id"] is None:
        await interaction.followup.send(
            f"Spot #{spot} was never claimed. Double-check the winning spot number.",
            ephemeral=True,
        )
        return

    winner         = interaction.guild.get_member(winning_slot["user_id"])
    winner_mention = winner.mention if winner else winning_slot["username"]
    winner_name    = winning_slot["username"]

    raffle["status"] = "complete"
    await _db_save_raffle(guild_id, name)

    channel = bot.get_channel(raffle["channel_id"])
    if channel and raffle["message_id"]:
        try:
            msg  = await channel.fetch_message(raffle["message_id"])
            view = _build_raffle_view(guild_id, name, raffle)
            await msg.edit(embed=_raffle_embed(name, raffle), view=view)
        except (discord.NotFound, discord.HTTPException):
            pass
        winner_embed = discord.Embed(
            title=f"We Have a Winner -- {name} Raffle!",
            description=(
                f"Congratulations to {winner_mention}!\n\n"
                f"**Winner:** {winner_name}\n"
                f"**Winning Spot:** #{spot}\n\n"
                f"Thanks to everyone who participated!"
            ),
            color=discord.Color.gold(),
        )
        await channel.send(embed=winner_embed)

    if winner:
        try:
            await winner.send(
                f"You won the **{name}** raffle with Spot #{spot}! Congratulations!\n"
                f"The server owner will reach out shortly with your prize details."
            )
        except discord.Forbidden:
            pass

    await interaction.followup.send(
        f"Winner recorded -- **{winner_name}**, Spot #{spot}.",
        ephemeral=True,
    )


@raffle_group.command(name="cancel", description="Cancel and remove a raffle")
@discord.app_commands.autocomplete(name=_raffle_name_autocomplete)
@discord.app_commands.describe(name="Raffle name")
async def slash_raffle_cancel(interaction: discord.Interaction, name: str):
    await interaction.response.defer(ephemeral=True)
    if interaction.guild.owner_id != interaction.user.id:
        await interaction.followup.send("Only the server owner can cancel raffles.", ephemeral=True)
        return
    guild_id = interaction.guild_id
    if name not in server_raffles[guild_id]:
        await interaction.followup.send(f"No raffle named **{name}**.", ephemeral=True)
        return

    raffle  = server_raffles[guild_id][name]
    channel = bot.get_channel(raffle["channel_id"])
    if channel and raffle["message_id"]:
        try:
            msg = await channel.fetch_message(raffle["message_id"])
            await msg.edit(
                embed=discord.Embed(
                    title=f"Raffle Cancelled -- {name}",
                    color=discord.Color.dark_gray(),
                ),
                view=None,
            )
        except (discord.NotFound, discord.HTTPException):
            pass

    await _db_delete_raffle(guild_id, name)
    del server_raffles[guild_id][name]
    await interaction.followup.send(f"Raffle **{name}** cancelled and removed.", ephemeral=True)


@raffle_group.command(name="status", description="Show the current state of a raffle")
@discord.app_commands.autocomplete(name=_raffle_name_autocomplete)
@discord.app_commands.describe(name="Raffle name")
async def slash_raffle_status(interaction: discord.Interaction, name: str):
    await interaction.response.defer(ephemeral=True)
    guild_id = interaction.guild_id
    if name not in server_raffles[guild_id]:
        await interaction.followup.send(f"No raffle named **{name}**.", ephemeral=True)
        return
    await interaction.followup.send(
        embed=_raffle_embed(name, server_raffles[guild_id][name]),
        ephemeral=True,
    )


@bot.tree.command(name="raffles", description="List all active raffles")
async def slash_raffles(interaction: discord.Interaction):
    await interaction.response.defer(ephemeral=True)
    guild_id = interaction.guild_id
    active   = {
        k: v for k, v in server_raffles.get(guild_id, {}).items()
        if v["status"] != "complete"
    }
    if not active:
        await interaction.followup.send("No active raffles right now.", ephemeral=True)
        return
    embed = discord.Embed(title="Active Raffles", color=discord.Color.blurple())
    for rname, r in active.items():
        claimed = sum(1 for s in r["slots"].values() if s["user_id"] is not None)
        paid    = sum(1 for s in r["slots"].values() if s["paid"])
        embed.add_field(
            name=rname,
            value=(
                f"Price: {r['price']} | "
                f"Claimed: {claimed}/{r['spots']} | "
                f"Paid: {paid} | "
                f"Status: {r['status'].capitalize()}"
            ),
            inline=False,
        )
    await interaction.followup.send(embed=embed, ephemeral=True)



@raffle_group.command(name="release", description="Release your claimed raffle spot back to open")
@discord.app_commands.autocomplete(name=_raffle_name_autocomplete)
@discord.app_commands.describe(name="Raffle name")
async def slash_raffle_release(interaction: discord.Interaction, name: str):
    await interaction.response.defer(ephemeral=True)
    guild_id = interaction.guild_id

    if name not in server_raffles.get(guild_id, {}):
        await interaction.followup.send(f"No raffle named **{name}**.", ephemeral=True)
        return

    raffle = server_raffles[guild_id][name]

    if raffle["status"] not in ("open", "closed"):
        await interaction.followup.send(
            "This raffle is no longer accepting changes.", ephemeral=True
        )
        return

    # Find spots held by this user
    user_spots = [
        num for num, s in raffle["slots"].items()
        if s["user_id"] == interaction.user.id
    ]

    if not user_spots:
        await interaction.followup.send(
            f"You don't have a spot in **{name}**.", ephemeral=True
        )
        return

    # Block release if payment already confirmed
    paid_spots = [num for num in user_spots if raffle["slots"][num]["paid"]]
    if paid_spots:
        spots_str = ", ".join(f"#{n}" for n in paid_spots)
        await interaction.followup.send(
            f"Spot{'s' if len(paid_spots) > 1 else ''} {spots_str} "
            f"{'have' if len(paid_spots) > 1 else 'has'} already been paid and confirmed — "
            f"contact the server owner to make changes.",
            ephemeral=True
        )
        return

    # Release all unpaid spots for this user
    released = []
    for num in user_spots:
        if not raffle["slots"][num]["paid"]:
            raffle["slots"][num] = {"user_id": None, "username": None, "paid": False}
            await _db_save_slot(guild_id, name, num)
            released.append(num)

    # If raffle was closed (all spots taken), reopen it
    if raffle["status"] == "closed":
        raffle["status"] = "open"
        await _db_save_raffle(guild_id, name)

    # Update the raffle embed
    channel = bot.get_channel(raffle["channel_id"])
    if channel and raffle["message_id"]:
        try:
            msg  = await channel.fetch_message(raffle["message_id"])
            view = _build_raffle_view(guild_id, name, raffle)
            await msg.edit(embed=_raffle_embed(name, raffle), view=view)
        except (discord.NotFound, discord.HTTPException):
            pass

    spots_str = ", ".join(f"#{n}" for n in sorted(released))
    await interaction.followup.send(
        f"Spot{'s' if len(released) > 1 else ''} {spots_str} in **{name}** "
        f"{'have' if len(released) > 1 else 'has'} been released back to open. "
        f"Tap a button to reclaim a different spot!",
        ephemeral=True
    )

    # Notify owner
    guild = interaction.guild
    if guild:
        owner = guild.get_member(guild.owner_id)
        if owner:
            try:
                await owner.send(
                    f"ℹ️  **{interaction.user.display_name}** released "
                    f"Spot{'s' if len(released) > 1 else ''} {spots_str} "
                    f"in raffle **{name}** on **{guild.name}**."
                )
            except discord.Forbidden:
                pass


@raffle_group.command(name="swap", description="Swap or clear a raffle spot (owner only)")
@discord.app_commands.autocomplete(name=_raffle_name_autocomplete)
@discord.app_commands.describe(
    name="Raffle name",
    spot="The spot number to reassign",
    user="New user to assign this spot to (leave empty to clear the spot)",
)
async def slash_raffle_swap(
    interaction: discord.Interaction,
    name: str,
    spot: int,
    user: discord.Member = None,
):
    await interaction.response.defer(ephemeral=True)

    if interaction.guild.owner_id != interaction.user.id:
        await interaction.followup.send(
            "Only the server owner can swap raffle spots.", ephemeral=True
        )
        return

    guild_id = interaction.guild_id

    if name not in server_raffles.get(guild_id, {}):
        await interaction.followup.send(f"No raffle named **{name}**.", ephemeral=True)
        return

    raffle = server_raffles[guild_id][name]

    if spot not in raffle["slots"]:
        await interaction.followup.send(
            f"Spot #{spot} doesn't exist. Valid spots: 1-{raffle['spots']}.",
            ephemeral=True
        )
        return

    current_slot = raffle["slots"][spot]
    current_holder = current_slot.get("username") or "empty"
    was_paid = current_slot.get("paid", False)

    if was_paid and user is None:
        # Clearing a paid spot — warn owner
        await interaction.followup.send(
            f"⚠️  Spot #{spot} is already confirmed as paid by **{current_holder}**. "
            f"Are you sure? Run the command again to confirm, or use `/raffle confirm` to fix payment.",
            ephemeral=True
        )
        return

    if user is None:
        # Clear the spot
        old_user_id = current_slot.get("user_id")
        raffle["slots"][spot] = {"user_id": None, "username": None, "paid": False}
        await _db_save_slot(guild_id, name, spot)

        # Reopen raffle if it was closed
        if raffle["status"] == "closed":
            raffle["status"] = "open"
            await _db_save_raffle(guild_id, name)

        # Notify previous holder if they exist
        if old_user_id:
            prev_member = interaction.guild.get_member(old_user_id)
            if prev_member:
                try:
                    await prev_member.send(
                        f"ℹ️  Your Spot #{spot} in the **{name}** raffle on **{interaction.guild.name}** "
                        f"has been cleared by the server owner. "
                        f"Please contact them if you have questions."
                    )
                except discord.Forbidden:
                    pass

        result_msg = f"✅  Spot #{spot} cleared (was held by **{current_holder}**). It's now open."

    else:
        # Assign spot to new user
        # Check if new user already has a spot in this raffle
        existing_spots = [
            num for num, s in raffle["slots"].items()
            if s["user_id"] == user.id
        ]
        if existing_spots:
            existing_str = ", ".join(f"#{n}" for n in existing_spots)
            await interaction.followup.send(
                f"⚠️  **{user.display_name}** already holds Spot{'s' if len(existing_spots) > 1 else ''} "
                f"{existing_str} in **{name}**. Swap anyway by reassigning.",
                ephemeral=True
            )
            return

        old_user_id = current_slot.get("user_id")
        old_was_paid = current_slot.get("paid", False)
        raffle["slots"][spot] = {
            "user_id":  user.id,
            "username": str(user),
            "paid":     False,  # reset paid status on swap
        }
        await _db_save_slot(guild_id, name, spot)

        # DM new holder with payment info
        payment_info = _build_raffle_payment_dm(guild_id, raffle.get("host_num", 0))
        try:
            await user.send(
                f"You have been assigned Spot #{spot} in the {name} raffle"
                f" on {interaction.guild.name}.\n\n"
                f"Price: {raffle['price']}\n\n"
                f"Send payment using:\n{payment_info}\n\n"
                "Once sent, run `!paid <method> <amount>` in the server."
            )
        except discord.Forbidden:
            pass

        # Notify previous holder they were moved
        if old_user_id and old_user_id != user.id:
            prev_member = interaction.guild.get_member(old_user_id)
            if prev_member:
                try:
                    await prev_member.send(
                        f"ℹ️  Your Spot #{spot} in the **{name}** raffle on **{interaction.guild.name}** "
                        f"has been reassigned by the server owner. "
                        f"Please contact them if you have questions."
                    )
                except discord.Forbidden:
                    pass

        result_msg = (
            f"✅  Spot #{spot} reassigned from **{current_holder}** to **{user.display_name}**."
            + (" *(was paid — payment status reset)*" if old_was_paid else "")
        )

    # Update raffle embed
    channel = bot.get_channel(raffle["channel_id"])
    if channel and raffle["message_id"]:
        try:
            msg  = await channel.fetch_message(raffle["message_id"])
            view = _build_raffle_view(guild_id, name, raffle)
            await msg.edit(embed=_raffle_embed(name, raffle), view=view)
        except (discord.NotFound, discord.HTTPException):
            pass

    await interaction.followup.send(result_msg, ephemeral=True)



@raffle_group.command(name="close", description="Clean up a completed raffle — removes it from the list and deletes the embed")
@discord.app_commands.autocomplete(name=_raffle_name_autocomplete)
@discord.app_commands.describe(name="Raffle name")
async def slash_raffle_close(interaction: discord.Interaction, name: str):
    await interaction.response.defer(ephemeral=True)

    if interaction.guild.owner_id != interaction.user.id:
        await interaction.followup.send(
            "Only the server owner can close raffles.", ephemeral=True
        )
        return

    guild_id = interaction.guild_id

    if name not in server_raffles.get(guild_id, {}):
        await interaction.followup.send(f"No raffle named **{name}**.", ephemeral=True)
        return

    raffle = server_raffles[guild_id][name]

    # Warn if trying to close an incomplete raffle
    if raffle["status"] not in ("complete",):
        claimed = sum(1 for s in raffle["slots"].values() if s["user_id"] is not None)
        paid    = sum(1 for s in raffle["slots"].values() if s["paid"])
        await interaction.followup.send(
            f"⚠️  **{name}** is not marked complete yet "
            f"(status: {raffle['status']}, {claimed}/{raffle['spots']} claimed, {paid} paid).\n"
            f"Use `/raffle winner` to record the winner first, or `/raffle cancel` to remove it entirely.",
            ephemeral=True
        )
        return

    # Remove from DB and memory — embed stays in channel for reference
    await _db_delete_raffle(guild_id, name)
    del server_raffles[guild_id][name]

    await interaction.followup.send(
        f"✅  Raffle **{name}** has been archived and removed from the active list. "
        f"The embed remains in the channel for reference.",
        ephemeral=True
    )



@raffle_group.command(name="sethost", description="Set payment info for a raffle host (owner only)")
@discord.app_commands.describe(
    host="Which host to configure (1 or 2)",
    name="Display name for this host (shown in payment DMs)",
    venmo="Venmo handle",
    zelle="Zelle phone or email",
    cashapp="Cash App handle",
    applepay="Apple Pay phone number",
)
async def slash_raffle_sethost(
    interaction: discord.Interaction,
    host: int,
    name: str,
    venmo: str = "",
    zelle: str = "",
    cashapp: str = "",
    applepay: str = "",
):
    await interaction.response.defer(ephemeral=True)

    if interaction.guild.owner_id != interaction.user.id:
        await interaction.followup.send(
            "Only the server owner can configure raffle hosts.", ephemeral=True
        )
        return

    if host not in (1, 2):
        await interaction.followup.send(
            "Host must be 1 or 2.", ephemeral=True
        )
        return

    guild_id = interaction.guild_id
    raffle_hosts[guild_id][host] = {
        "name":     name,
        "venmo":    venmo    or None,
        "zelle":    zelle    or None,
        "cashapp":  cashapp  or None,
        "applepay": applepay or None,
    }
    await db_save_raffle_host(guild_id, host)

    lines = [f"**Host {host} — {name}**"]
    if venmo:    lines.append(f"Venmo: {venmo}")
    if zelle:    lines.append(f"Zelle: {zelle}")
    if cashapp:  lines.append(f"Cash App: {cashapp}")
    if applepay: lines.append(f"Apple Pay: {applepay}")
    if not any([venmo, zelle, cashapp, applepay]):
        lines.append("⚠️  No payment methods set — add at least one.")

    await interaction.followup.send(
        "✅  Host saved:\n" + "\n".join(lines),
        ephemeral=True
    )

# Register slash command group
bot.tree.add_command(raffle_group)



async def post_or_refresh_payment_board(guild, channel):
    """Update the tracked payment board if one exists, else post a fresh one in
    `channel`. Returns a short status string."""
    guild_id = guild.id
    embed = await render_payment_board(guild_id)
    drop_number = await _current_drop_number(guild_id)
    pb_msg = payment_board_message.get(guild_id)

    # After a restart the in-memory message is gone but the board is still
    # tracked in the DB — reuse it so we edit the existing board instead of
    # posting a duplicate.
    if pb_msg is None:
        ref = payment_board_ref.get(guild_id)
        if ref:
            ch = bot.get_channel(ref["channel_id"])
            if ch is None:
                try:
                    ch = await bot.fetch_channel(ref["channel_id"])
                except (discord.NotFound, discord.Forbidden, discord.HTTPException):
                    ch = None
            if ch:
                try:
                    pb_msg = await ch.fetch_message(ref["message_id"])
                    payment_board_message[guild_id] = pb_msg
                except (discord.NotFound, discord.Forbidden, discord.HTTPException):
                    pb_msg = None

    if pb_msg:
        try:
            await pb_msg.edit(embed=embed)
            await db_save_payment_board(guild_id, pb_msg.channel.id, pb_msg.id, drop_number)
            return "updated"
        except (discord.NotFound, discord.HTTPException):
            payment_board_message.pop(guild_id, None)
    msg = await channel.send(embed=embed)
    payment_board_message[guild_id] = msg
    await db_save_payment_board(guild_id, msg.channel.id, msg.id, drop_number)
    return f"posted in #{msg.channel.name}"


@bot.command(name="paymentboard")
async def cmd_paymentboard(ctx):
    """Post or refresh the payment board in the drop channel."""
    if not ctx.guild:
        await ctx.author.send("⚠️  Please run `!paymentboard` in your server channel.")
        return
    guild_id = ctx.guild.id
    if not is_manager(guild_id, ctx.author.id):
        return
    await silent(ctx)
    drop_channel = get_drop_channel(ctx.guild) or ctx.channel
    result = await post_or_refresh_payment_board(ctx.guild, drop_channel)
    await dm(ctx, f"✅  Payment board {result}.")


@bot.command(name="payments")
async def cmd_payments(ctx):
    """DM the manager a full payment summary across all active drops and raffles."""
    if not ctx.guild:
        await ctx.author.send("⚠️  Please run `!payments` in your server channel.")
        return
    guild_id = ctx.guild.id
    if not is_manager(guild_id, ctx.author.id):
        return
    await silent(ctx)

    stock_ref    = stock[guild_id]   if stock[guild_id]   else last_drop_snapshot.get(guild_id, {}).get("stock",  {})
    claims_ref   = claims[guild_id]  if claims[guild_id]  else last_drop_snapshot.get(guild_id, {}).get("claims", {})
    archived     = archived_payments.get(guild_id, {})
    arch_claims  = archived.get("claims", {})
    arch_stock   = archived.get("stock",  {})
    arch_pmts    = archived.get("payments", {})

    # ── Build per-user totals across live + archived drops ────────────────────
    def build_user_totals(claims_data, stock_data, payments_data, label):
        rows = []
        seen = set()
        for key, claim_list in claims_data.items():
            for c in claim_list:
                uid = c["user"].id
                if uid in seen or key not in stock_data:
                    continue
                seen.add(uid)
                owed = sum(
                    cc["qty"] * stock_data[k]["price"]
                    for k, cl in claims_data.items()
                    for cc in cl
                    if cc["user"].id == uid and k in stock_data
                )
                confirmed = sum(
                    p["amount"] for p in payments_data.get(uid, [])
                    if p["confirmed"]
                )
                pending = sum(
                    p["amount"] for p in payments_data.get(uid, [])
                    if not p["confirmed"]
                )
                outstanding = owed - confirmed
                rows.append({
                    "name":        c["user"].display_name,
                    "owed":        owed,
                    "confirmed":   confirmed,
                    "pending":     pending,
                    "outstanding": outstanding,
                    "label":       label,
                })
        return rows

    live_rows     = build_user_totals(claims_ref,  stock_ref,  payments[guild_id], "Current Drop")
    archived_rows = build_user_totals(arch_claims, arch_stock, arch_pmts,          "Previous Drop")

    # ── Raffle rows ───────────────────────────────────────────────────────────
    raffle_rows = []
    for r_name, raffle in server_raffles.get(guild_id, {}).items():
        if raffle["status"] == "complete":
            continue
        for spot_num, s in raffle["slots"].items():
            if s["user_id"] is None:
                continue
            raffle_rows.append({
                "name":    s["username"],
                "raffle":  r_name,
                "spot":    spot_num,
                "price":   raffle["price"],
                "paid":    s["paid"],
            })

    # ── Build embed ───────────────────────────────────────────────────────────
    embed = discord.Embed(
        title="💳  Payment Summary — All Drops",
        color=discord.Color.gold(),
        timestamp=datetime.datetime.utcnow()
    )

    total_confirmed_all   = 0.0
    total_outstanding_all = 0.0

    def format_rows(rows, title):
        nonlocal total_confirmed_all, total_outstanding_all
        if not rows:
            return
        lines = []
        for r in rows:
            if r["outstanding"] <= 0.01 and r["confirmed"] > 0:
                status = "✅"
            elif r["pending"] > 0:
                status = f"⏳ ${r['pending']:.2f} pending"
            else:
                status = f"owes ${r['outstanding']:.2f}"
            lines.append(
                f"**{r['name']}** — owed ${r['owed']:.2f} | "
                f"confirmed ${r['confirmed']:.2f} | {status}"
            )
            total_confirmed_all   += r["confirmed"]
            total_outstanding_all += max(r["outstanding"], 0)
        field_value = "\n".join(lines)
        if len(field_value) > 1024:
            field_value = field_value[:1020] + "..."
        embed.add_field(name=title, value=field_value, inline=False)

    format_rows(live_rows,     "🟢  Current Drop")
    format_rows(archived_rows, "🕐  Previous Drop")

    if raffle_rows:
        r_lines = []
        for r in raffle_rows:
            status = "✅  Paid" if r["paid"] else "⏳  Awaiting payment"
            r_lines.append(
                f"**{r['name']}** — {r['raffle']} Spot #{r['spot']} "
                f"({r['price']}) — {status}"
            )
        r_value = "\n".join(r_lines)
        if len(r_value) > 1024:
            r_value = r_value[:1020] + "..."
        embed.add_field(name="🎟️  Active Raffles", value=r_value, inline=False)

    if not live_rows and not archived_rows and not raffle_rows:
        embed.description = "No active drops or raffles with claims right now."
    else:
        embed.set_footer(
            text=f"Total confirmed: ${total_confirmed_all:.2f}  |  "
                 f"Total outstanding: ${total_outstanding_all:.2f}"
        )

    try:
        await ctx.author.send(embed=embed)
    except discord.Forbidden:
        await ctx.send("⚠️  I couldn't DM you — please open your DMs and try again.")


@bot.command(name="addtracking")
async def cmd_addtracking(ctx, *, args=""):
    """Attach a tracking number to a buyer. Usage: !addtracking @user <tracking#>"""
    if not ctx.guild:
        await ctx.author.send("⚠️  Please run `!addtracking` in your server channel.")
        return
    guild_id = ctx.guild.id
    if not is_manager(guild_id, ctx.author.id):
        return
    await silent(ctx)
    if not ctx.message.mentions:
        await dm(ctx, "Usage: `!addtracking @user <tracking number>`")
        return
    user = ctx.message.mentions[0]
    # Strip the mention from args to get tracking number
    tracking = args.replace(f"<@{user.id}>", "").replace(f"<@!{user.id}>", "").strip()
    if not tracking:
        await dm(ctx, "Usage: `!addtracking @user <tracking number>`\nExample: `!addtracking @SpacemanG 1Z999AA10123456784`")
        return
    # Persist the tracking number to the buyer's most recent drop in the DB so
    # it becomes a permanent part of their order history (survives restarts).
    drop_no = await db_set_user_claim_tracking(guild_id, user.id, tracking)
    if drop_no is None:
        await dm(
            ctx,
            f"⚠️  **{user.display_name}** has no saved orders yet, so there's nothing to "
            f"attach tracking to. Orders are saved once a drop closes (`!enddrop`)."
        )
        return
    await dm(
        ctx,
        f"✅  Tracking number **{tracking}** saved for **{user.display_name}** "
        f"on **Drop #{drop_no}**."
    )
    # DM the buyer their tracking number
    try:
        await user.send(
            f"📦  Your order from **Drop #{drop_no}** has shipped! "
            f"Here is your tracking number:\n"
            f"**{tracking}**\n\n"
            f"You can use `!myhistory` to view your full order history."
        )
    except discord.Forbidden:
        await dm(ctx, f"⚠️  Saved, but couldn't DM **{user.display_name}** — their DMs may be closed.")


@bot.command(name="export")
async def cmd_export(ctx):
    """Generate and DM an Excel export of the current drop."""
    if not ctx.guild:
        await ctx.author.send("⚠️  Please run `!export` in your server channel.")
        return
    guild_id = ctx.guild.id
    if not is_manager(guild_id, ctx.author.id):
        return
    await silent(ctx)

    stock_ref  = stock[guild_id]  if stock[guild_id]  else last_drop_snapshot.get(guild_id, {}).get("stock",  {})
    claims_ref = claims[guild_id] if claims[guild_id] else last_drop_snapshot.get(guild_id, {}).get("claims", {})
    archived   = archived_payments.get(guild_id, {})

    # Most recent tracking number per buyer, pulled from persistent history.
    tracking_map = {}
    if db_pool is not None:
        async with db_pool.acquire() as conn:
            trows = await conn.fetch("""
                SELECT DISTINCT ON (user_id) user_id, tracking
                FROM user_claims
                WHERE guild_id = $1 AND tracking IS NOT NULL AND tracking <> ''
                ORDER BY user_id, drop_number DESC
            """, guild_id)
        tracking_map = {r["user_id"]: r["tracking"] for r in trows}

    # ── Styles ────────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()

    HEADER_FILL   = PatternFill("solid", fgColor="1E1E2E")
    HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11)
    PAID_FILL     = PatternFill("solid", fgColor="D4EDDA")
    PENDING_FILL  = PatternFill("solid", fgColor="FFF3CD")
    UNPAID_FILL   = PatternFill("solid", fgColor="F8D7DA")
    ALT_FILL      = PatternFill("solid", fgColor="F5F5F5")
    BORDER_SIDE   = Side(style="thin", color="CCCCCC")
    THIN_BORDER   = Border(left=BORDER_SIDE, right=BORDER_SIDE,
                           top=BORDER_SIDE, bottom=BORDER_SIDE)
    CENTER        = Alignment(horizontal="center", vertical="center")
    LEFT          = Alignment(horizontal="left", vertical="center")

    def style_header_row(ws, row, cols):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill   = HEADER_FILL
            cell.font   = HEADER_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER

    def style_data_row(ws, row, cols, fill=None):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            if fill:
                cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = LEFT

    def auto_width(ws, min_width=10, max_width=40):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except (TypeError, AttributeError):
                    pass
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)

    # ── SHEET 1: Orders ───────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Orders"
    ws1.freeze_panes = "A2"

    headers1 = ["Buyer", "Item", "Qty", "Unit Price", "Subtotal",
                 "Total Owed", "Confirmed", "Outstanding", "Status", "Tracking #"]
    for col, h in enumerate(headers1, 1):
        ws1.cell(row=1, column=col, value=h)
    style_header_row(ws1, 1, len(headers1))

    row = 2
    # Build per-user order data
    user_totals = {}
    for key, claim_list in claims_ref.items():
        if key not in stock_ref:
            continue
        for c in claim_list:
            uid = c["user"].id
            if uid not in user_totals:
                user_totals[uid] = {
                    "name":   c["user"].display_name,
                    "items":  [],
                    "owed":   0.0,
                }
            subtotal = c["qty"] * stock_ref[key]["price"]
            user_totals[uid]["items"].append({
                "display":  stock_ref[key]["display"],
                "qty":      c["qty"],
                "price":    stock_ref[key]["price"],
                "subtotal": subtotal,
            })
            user_totals[uid]["owed"] += subtotal

    for uid, udata in user_totals.items():
        confirmed = sum(p["amount"] for p in payments[guild_id].get(uid, []) if p["confirmed"])
        outstanding = max(udata["owed"] - confirmed, 0)
        tracking = tracking_map.get(uid, "")
        if outstanding <= 0.01 and confirmed > 0:
            status = "✅ Paid"
            fill = PAID_FILL
        elif confirmed > 0:
            status = "⏳ Partial"
            fill = PENDING_FILL
        else:
            status = "❌ Unpaid"
            fill = UNPAID_FILL

        first = True
        for item in udata["items"]:
            ws1.cell(row=row, column=1, value=udata["name"] if first else "")
            ws1.cell(row=row, column=2, value=item["display"])
            ws1.cell(row=row, column=3, value=item["qty"])
            ws1.cell(row=row, column=4, value=round(item["price"], 2))
            ws1.cell(row=row, column=5, value=round(item["subtotal"], 2))
            ws1.cell(row=row, column=6, value=round(udata["owed"], 2) if first else "")
            ws1.cell(row=row, column=7, value=round(confirmed, 2) if first else "")
            ws1.cell(row=row, column=8, value=round(outstanding, 2) if first else "")
            ws1.cell(row=row, column=9, value=status if first else "")
            ws1.cell(row=row, column=10, value=tracking if first else "")
            style_data_row(ws1, row, len(headers1), fill)
            ws1.cell(row=row, column=4).number_format = "$#,##0.00"
            ws1.cell(row=row, column=5).number_format = "$#,##0.00"
            if first:
                ws1.cell(row=row, column=6).number_format = "$#,##0.00"
                ws1.cell(row=row, column=7).number_format = "$#,##0.00"
                ws1.cell(row=row, column=8).number_format = "$#,##0.00"
            first = False
            row += 1

    auto_width(ws1)

    # ── SHEET 2: Payment Summary ───────────────────────────────────────────────
    ws2 = wb.create_sheet("Payment Summary")
    ws2.freeze_panes = "A2"

    headers2 = ["Buyer", "Total Owed", "Confirmed", "Outstanding",
                 "Payment Methods", "Status", "Tracking #"]
    for col, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=col, value=h)
    style_header_row(ws2, 1, len(headers2))

    row2 = 2
    for uid, udata in user_totals.items():
        uid_pmts    = payments[guild_id].get(uid, [])
        confirmed   = sum(p["amount"] for p in uid_pmts if p["confirmed"])
        outstanding = max(udata["owed"] - confirmed, 0)
        methods     = ", ".join(
            f"{p['method'].title()} ${p['amount']:.2f}"
            for p in uid_pmts if p["confirmed"]
        ) or "None"
        tracking = tracking_map.get(uid, "")
        if outstanding <= 0.01 and confirmed > 0:
            status = "✅ Paid"
            fill = PAID_FILL
        elif confirmed > 0:
            status = "⏳ Partial"
            fill = PENDING_FILL
        else:
            status = "❌ Unpaid"
            fill = UNPAID_FILL

        ws2.cell(row=row2, column=1, value=udata["name"])
        ws2.cell(row=row2, column=2, value=round(udata["owed"], 2))
        ws2.cell(row=row2, column=3, value=round(confirmed, 2))
        ws2.cell(row=row2, column=4, value=round(outstanding, 2))
        ws2.cell(row=row2, column=5, value=methods)
        ws2.cell(row=row2, column=6, value=status)
        ws2.cell(row=row2, column=7, value=tracking)
        style_data_row(ws2, row2, len(headers2), fill)
        for col in [2, 3, 4]:
            ws2.cell(row=row2, column=col).number_format = "$#,##0.00"
        row2 += 1

    # Totals row
    if user_totals:
        ws2.cell(row=row2, column=1, value="TOTAL")
        ws2.cell(row=row2, column=1).font = Font(bold=True)
        total_owed = sum(u["owed"] for u in user_totals.values())
        total_conf = sum(
            sum(p["amount"] for p in payments[guild_id].get(uid, []) if p["confirmed"])
            for uid in user_totals
        )
        ws2.cell(row=row2, column=2, value=round(total_owed, 2))
        ws2.cell(row=row2, column=3, value=round(total_conf, 2))
        ws2.cell(row=row2, column=4, value=round(max(total_owed - total_conf, 0), 2))
        for col in [2, 3, 4]:
            ws2.cell(row=row2, column=col).number_format = "$#,##0.00"
            ws2.cell(row=row2, column=col).font = Font(bold=True)
        style_data_row(ws2, row2, len(headers2))

    auto_width(ws2)

    # ── SHEET 3: Raffles ──────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Raffles")
    ws3.freeze_panes = "A2"

    headers3 = ["Raffle Name", "Spot #", "Buyer", "Price", "Paid", "Host"]
    for col, h in enumerate(headers3, 1):
        ws3.cell(row=1, column=col, value=h)
    style_header_row(ws3, 1, len(headers3))

    row3 = 2
    for r_name, raffle in server_raffles.get(guild_id, {}).items():
        host_num  = raffle.get("host_num", 0)
        host_data = raffle_hosts.get(guild_id, {}).get(host_num, {})
        host_name = host_data.get("name", "Server Default") if host_num in (1, 2) else "Server Default"
        for spot_num in sorted(raffle["slots"].keys()):
            s    = raffle["slots"][spot_num]
            fill = PAID_FILL if s["paid"] else (PENDING_FILL if s["user_id"] else ALT_FILL)
            ws3.cell(row=row3, column=1, value=r_name)
            ws3.cell(row=row3, column=2, value=spot_num)
            ws3.cell(row=row3, column=3, value=s["username"] or "Open")
            ws3.cell(row=row3, column=4, value=raffle["price"])
            ws3.cell(row=row3, column=5, value="✅ Yes" if s["paid"] else ("⏳ Pending" if s["user_id"] else "Open"))
            ws3.cell(row=row3, column=6, value=host_name)
            style_data_row(ws3, row3, len(headers3), fill)
            row3 += 1

    auto_width(ws3)

    # ── SHEET 4: Previous Drop ────────────────────────────────────────────────
    arch_claims_data = archived.get("claims", {})
    arch_stock_data  = archived.get("stock",  {})
    arch_pmts_data   = archived.get("payments", {})

    if arch_claims_data:
        ws4 = wb.create_sheet("Previous Drop")
        ws4.freeze_panes = "A2"

        headers4 = ["Buyer", "Item", "Qty", "Unit Price", "Subtotal",
                     "Total Owed", "Confirmed", "Outstanding", "Status"]
        for col, h in enumerate(headers4, 1):
            ws4.cell(row=1, column=col, value=h)
        style_header_row(ws4, 1, len(headers4))

        arch_user_totals = {}
        for key, claim_list in arch_claims_data.items():
            if key not in arch_stock_data:
                continue
            for c in claim_list:
                uid = c["user"].id
                if uid not in arch_user_totals:
                    arch_user_totals[uid] = {
                        "name":  c["user"].display_name,
                        "items": [],
                        "owed":  0.0,
                    }
                subtotal = c["qty"] * arch_stock_data[key]["price"]
                arch_user_totals[uid]["items"].append({
                    "display":  arch_stock_data[key]["display"],
                    "qty":      c["qty"],
                    "price":    arch_stock_data[key]["price"],
                    "subtotal": subtotal,
                })
                arch_user_totals[uid]["owed"] += subtotal

        row4 = 2
        for uid, udata in arch_user_totals.items():
            confirmed   = sum(p["amount"] for p in arch_pmts_data.get(uid, []) if p["confirmed"])
            outstanding = max(udata["owed"] - confirmed, 0)
            if outstanding <= 0.01 and confirmed > 0:
                status = "✅ Paid"
                fill = PAID_FILL
            elif confirmed > 0:
                status = "⏳ Partial"
                fill = PENDING_FILL
            else:
                status = "❌ Unpaid"
                fill = UNPAID_FILL
            first = True
            for item in udata["items"]:
                ws4.cell(row=row4, column=1, value=udata["name"] if first else "")
                ws4.cell(row=row4, column=2, value=item["display"])
                ws4.cell(row=row4, column=3, value=item["qty"])
                ws4.cell(row=row4, column=4, value=round(item["price"], 2))
                ws4.cell(row=row4, column=5, value=round(item["subtotal"], 2))
                ws4.cell(row=row4, column=6, value=round(udata["owed"], 2) if first else "")
                ws4.cell(row=row4, column=7, value=round(confirmed, 2) if first else "")
                ws4.cell(row=row4, column=8, value=round(outstanding, 2) if first else "")
                ws4.cell(row=row4, column=9, value=status if first else "")
                style_data_row(ws4, row4, len(headers4), fill)
                first = False
                row4 += 1
        auto_width(ws4)

    # ── Save to bytes and DM ──────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    guild   = ctx.guild
    date_str = datetime.datetime.utcnow().strftime("%Y-%m-%d")
    filename = f"VaultDrop_{guild.name.replace(' ', '_')}_{date_str}.xlsx"

    try:
        await ctx.author.send(
            f"📊  **Drop Export — {guild.name}** ({date_str})\\n"
            f"Sheets: Orders | Payment Summary | Raffles"
            + (" | Previous Drop" if arch_claims_data else ""),
            file=discord.File(buf, filename=filename)
        )
    except discord.Forbidden:
        await ctx.send("⚠️  I couldn\\'t DM you the file — please open your DMs and try again.")


bot.run(BOT_TOKEN)
